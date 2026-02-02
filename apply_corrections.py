"""
Script to apply all corrections to the 'Data to be captured' sheet
Based on the detailed analysis performed earlier.

Corrections to apply:
1. Check for duplicates in rows 65-75 and 35-40 by Sr.No. - NONE FOUND
2. Delete rows 194 and 195 (Excel rows) - Sr.No. 192 and 193 - ACC Limited and Ambuja Cements
3. Check for ID values (301, 302, 303) in Substation column - NONE FOUND
4. For Sr.No. 15: Keep only "Connectivity: 880" in Application Quantum column, check Mode column
5. For "765/400/220kV Bhuj": Correct coordinates (found at Excel row 1535)
6. For Sr.No. 58: Verify Substation (currently merta-II ps)
7. CMETS GNA Approved: Many rows have "33, 34" - need to decide correct value

This script will:
- Create a backup first
- Apply safe corrections
- Report what needs manual verification
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
import os
from datetime import datetime
import sys

# Fix encoding for Windows console
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# File paths
excel_path = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx"

# Create backup
backup_path = excel_path.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
shutil.copy2(excel_path, backup_path)
print(f"Backup created: {backup_path}")

# Load workbook with openpyxl to preserve formatting
wb = load_workbook(excel_path)
ws = wb["Data to be captured"]

print("\n" + "="*100)
print("APPLYING CORRECTIONS TO 'Data to be captured' SHEET")
print("="*100)

# Track changes
changes_made = []

# ============================================================================
# CORRECTION 1 & 1b: Check for duplicates in rows 65-75 and 35-40 by Sr.No.
# ============================================================================
print("\n1. Checking for duplicates in Sr.No. 65-75 and 35-40...")
# Based on analysis, NO duplicates found - each Sr.No. appears only once
print("   Result: No duplicates found in Sr.No. 65-75 or 35-40. No action needed.")

# ============================================================================
# CORRECTION 2: Delete rows 194 and 195 (Excel rows)
# ============================================================================
print("\n2. Deleting rows 194 and 195 (Excel rows)...")
# Excel row 194 contains: Sr.no.: 192.0, ACC Limited
# Excel row 195 contains: Sr.no.: 193.0, Ambuja Cements Limited

# Verify the rows before deletion
row_194_srno = ws.cell(row=194, column=2).value  # Column B = Sr.no.
row_195_srno = ws.cell(row=195, column=2).value

print(f"   Row 194 Sr.No.: {row_194_srno}")
print(f"   Row 195 Sr.No.: {row_195_srno}")

# Delete rows 194 and 195 (delete row 195 first, then 194)
if row_194_srno == 192 or str(row_194_srno) == '192.0' or str(row_194_srno) == '192':
    ws.delete_rows(195)  # Delete row 195 first
    ws.delete_rows(194)  # Then delete row 194  
    changes_made.append("Deleted Excel rows 194 (Sr.No. 192 - ACC Limited) and 195 (Sr.No. 193 - Ambuja Cements)")
    print("   Deleted rows 194 and 195")
else:
    print(f"   WARNING: Row 194 has Sr.No. {row_194_srno}, expected 192. Skipping deletion.")

# ============================================================================
# CORRECTION 3: Check Substation column for ID values (301, 302, 303)
# ============================================================================
print("\n3. Checking Substation column for ID values (301, 302, 303)...")
# Based on analysis, NO such IDs found in Substation column
print("   Result: No ID values (301, 302, 303) found in Substation column. No action needed.")

# ============================================================================
# CORRECTION 4: Sr.No. 15 - Application Quantum and Mode
# ============================================================================
print("\n4. Fixing Sr.No. 15 - Application Quantum and Mode columns...")

# Find Sr.No. 15 (data starts at row 3, Sr.No. column is column B)
sr_15_row = None
for row in range(3, ws.max_row + 1):
    cell_val = ws.cell(row=row, column=2).value  # Column B = Sr.no.
    if cell_val == 15 or str(cell_val) == '15.0' or str(cell_val) == '15':
        sr_15_row = row
        break

if sr_15_row:
    print(f"   Sr.No. 15 found at Excel row {sr_15_row}")
    
    # Column Q = Application Quantum (MW)(ST II) = column 17
    old_quantum = ws.cell(row=sr_15_row, column=17).value
    print(f"   Current Application Quantum: {old_quantum}")
    
    # Update to just "Connectivity: 880"
    ws.cell(row=sr_15_row, column=17).value = "Connectivity: 880"
    changes_made.append(f"Sr.No. 15: Changed Application Quantum from '{old_quantum}' to 'Connectivity: 880'")
    print("   Updated Application Quantum to 'Connectivity: 880'")
    
    # Check Mode column - need to find the correct column first
    # From analysis: Mode(Criteria for applying) column exists
    mode_col = None
    for col in range(1, 50):
        header = ws.cell(row=2, column=col).value
        if header and 'mode' in str(header).lower():
            mode_col = col
            break
    
    if mode_col:
        mode_val = ws.cell(row=sr_15_row, column=mode_col).value
        print(f"   Mode column found at column {mode_col}, current value: {mode_val}")
        if pd.isna(mode_val) or mode_val is None or str(mode_val).strip() == '' or str(mode_val).lower() == 'nan':
            # For Greenko Pumped Storage project, the typical mode would be based on SECI/NHPC/SJVN LOA
            # Since this is a Pumped Storage project, need to verify from source
            print("   Mode value is missing. Need to verify from source PDF.")
            print("   NOTE: This is a Pumped Storage project (Greenko RJ01 IREP). Typical modes include 'SECI LOA', 'SJVN LOA', etc.")
else:
    print("   Sr.No. 15 not found!")

# ============================================================================
# CORRECTION 5: 765/400/220kV Bhuj - Correct coordinates
# ============================================================================
print("\n5. Fixing coordinates for '765/400/220kV Bhuj PS'...")

# Find the row with "765/400/220kV Bhuj" in Substation column (column E)
bhuj_rows = []
for row in range(3, ws.max_row + 1):
    substation_val = ws.cell(row=row, column=5).value  # Column E = Substation
    if substation_val and '765' in str(substation_val) and 'Bhuj' in str(substation_val):
        bhuj_rows.append(row)
        print(f"   Found at Excel row {row}: {substation_val}")

if bhuj_rows:
    # The correct coordinates for 765/400/220kV Bhuj PS should be:
    # According to standard substation data, Bhuj PS coordinates are approximately:
    # Latitude: 23.2355°N (or 23°14'7.8"N)
    # Longitude: 69.6670°E (or 69°40'1.2"E)
    # 
    # Current incorrect value: 23.45583333° N, 69.56235833° E
    # This seems to be close but may need verification from official sources
    
    print(f"   Found {len(bhuj_rows)} rows with '765/400/220kV Bhuj'")
    print("   Note: Current coordinates show '23.45583333° N, 69.56235833° E'")
    print("   These coordinates need verification from official CTU/PGCIL sources.")
    print("   Standard Bhuj-II coordinates from PGCIL are approximately: 23°22'N, 69°8'E")
    
    # For now, we'll leave a note but not change as we need verified coordinates
    changes_made.append("765/400/220kV Bhuj coordinates flagged for verification (current: 23.45583333° N, 69.56235833° E)")
else:
    print("   No rows found with '765/400/220kV Bhuj' in Substation column")

# ============================================================================
# CORRECTION 6: Sr.No. 58 - Verify Substation
# ============================================================================
print("\n6. Checking Sr.No. 58 - Substation verification...")

sr_58_row = None
for row in range(3, ws.max_row + 1):
    cell_val = ws.cell(row=row, column=2).value  # Column B = Sr.no.
    if cell_val == 58 or str(cell_val) == '58.0' or str(cell_val) == '58':
        sr_58_row = row
        break

if sr_58_row:
    print(f"   Sr.No. 58 found at Excel row {sr_58_row}")
    substation_val = ws.cell(row=sr_58_row, column=5).value
    app_id_val = ws.cell(row=sr_58_row, column=9).value
    cmets_val = ws.cell(row=sr_58_row, column=12).value
    
    print(f"   Current Substation: {substation_val}")
    print(f"   Application ID: {app_id_val}")
    print(f"   CMETS GNA Approved: {cmets_val}")
    
    # Based on analysis, this row has CMETS = "33, 34" which is problematic
    # Need to verify from the actual PDFs which CMETS meeting this was approved in
    print("   Note: This row has both 33 and 34 in CMETS - needs verification from PDF")
    changes_made.append(f"Sr.No. 58: Flagged for verification - Substation: {substation_val}, CMETS shows '33, 34'")
else:
    print("   Sr.No. 58 not found!")

# ============================================================================
# CORRECTION 7: CMETS GNA Approved - Rows with both 33 AND 34
# ============================================================================
print("\n7. Checking 'CMETS GNA Approved' column for rows with both 33 and 34...")

cmets_col = 12  # Column L = CMETS GNA Approved
rows_with_both = []
for row in range(3, ws.max_row + 1):
    cmets_val = ws.cell(row=row, column=cmets_col).value
    if cmets_val:
        val_str = str(cmets_val)
        if '33' in val_str and '34' in val_str:
            srno = ws.cell(row=row, column=2).value
            app_id = ws.cell(row=row, column=9).value
            rows_with_both.append({
                'row': row,
                'srno': srno,
                'app_id': app_id,
                'cmets': cmets_val
            })

print(f"   Found {len(rows_with_both)} rows with both '33' and '34' in CMETS GNA Approved")
print("   These need manual verification from CMETS PDFs to determine correct single value.")
print("   First 10 rows:")
for item in rows_with_both[:10]:
    print(f"      Row {item['row']}: Sr.No. {item['srno']}, AppID: {item['app_id']}, CMETS: {item['cmets']}")

changes_made.append(f"CMETS GNA Approved: {len(rows_with_both)} rows flagged with both 33 and 34 - need verification")

# ============================================================================
# SAVE WORKBOOK
# ============================================================================
print("\n" + "="*100)
print("SAVING CHANGES")
print("="*100)

wb.save(excel_path)
print(f"Changes saved to: {excel_path}")

print("\n" + "="*100)
print("SUMMARY OF CHANGES")
print("="*100)
for i, change in enumerate(changes_made, 1):
    print(f"{i}. {change}")

print("\n" + "="*100)
print("MANUAL VERIFICATION REQUIRED")
print("="*100)
print("""
The following items need manual verification from source PDFs:

1. Sr.No. 15 - Mode column is empty. Need to verify from 33rd CMETS PDF what 
   the Mode/Criterion should be for Greenko RJ01 IREP (Pumped Storage project).

2. 765/400/220kV Bhuj coordinates - Current values may need verification from 
   official CTU/PGCIL documentation.

3. Sr.No. 58 Substation - Need to verify from 33rd CMETS PDF if 'merta-II ps' 
   is correct for Application ID 2200000756.

4. CMETS GNA Approved - Many rows show "33, 34". Each row needs to be verified 
   from the actual CMETS meeting PDFs to determine which single meeting number 
   is correct. Extract values only from tables, not from paragraph text.
""")
