"""Verify no duplicates exist and data is preserved"""
import openpyxl
from collections import Counter

wb = openpyxl.load_workbook(r'c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx')
ws = wb['Data to be captured']

print('=== Verifying Unique IDs ===')
ids = []
for row_idx in range(3, ws.max_row + 1):
    val = ws.cell(row=row_idx, column=9).value
    if val:
        val_str = str(val).strip()
        ids.append(val_str)

counts = Counter(ids)
duplicates = {k: v for k, v in counts.items() if v > 1}

if duplicates:
    print(f"FAILED: Found {len(duplicates)} duplicates!")
    print(list(duplicates.items())[:5])
else:
    print("SUCCESS: No duplicates found in GNA/ST II Application ID column.")

print('\n=== Verifying Specific ID: 2200000860 ===')
target_id = '2200000860'
found = []
for row_idx in range(3, ws.max_row + 1):
    val = ws.cell(row=row_idx, column=9).value
    if val and str(val).strip() == target_id:
        found.append(row_idx)

print(f"ID {target_id} found in rows: {found}")
if len(found) == 1:
    r = found[0]
    element_codes = [ws.cell(row=r, column=c).value for c in [40, 41, 42]]
    print(f"Row {r} Data (Element Codes): {element_codes}")
    if any(element_codes):
        print("SUCCESS: Preserved the row with element data.")
    else:
        print("WARNING: Row has no element data (maybe none existed originally?)")
else:
    print(f"ERROR: Expected exactly 1 row, found {len(found)}")
