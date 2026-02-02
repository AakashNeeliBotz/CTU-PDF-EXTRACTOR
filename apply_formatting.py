import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def apply_comprehensive_formatting():
    """Apply comprehensive formatting to the Excel file"""
    file_path = 'Connectivity_Application_Data_TEST_ALL_SHEETS38.xlsx'
    
    print("Applying comprehensive formatting...")
    
    # Load the workbook
    wb = load_workbook(file_path)
    
    # Define formatting styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # For each sheet, format the header rows
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        print(f"Formatting sheet: {sheet_name}")
        
        # Apply formatting to row 1 (main headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # If there's a second header row, format it too
        if ws.max_row >= 2:
            for cell in ws[2]:
                if cell.value is not None and str(cell.value).strip() != '':
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
    
    # Save the workbook
    wb.save(file_path)
    print(f"Formatting completed for: {file_path}")

def verify_formatting():
    """Verify that formatting was applied"""
    file_path = 'Connectivity_Application_Data_TEST_ALL_SHEETS38.xlsx'
    
    print("\\nVerifying formatting application...")
    
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Sheet '{sheet_name}':")
        
        # Check if row 1 has bold formatting
        sample_cell = ws['A1']
        is_bold = sample_cell.font.bold if hasattr(sample_cell.font, 'bold') else False
        print(f"  - Row 1 A1 cell is bold: {is_bold}")
        print(f"  - Row 1 A1 font size: {sample_cell.font.sz if hasattr(sample_cell.font, 'sz') else 'N/A'}")
        break  # Just check first sheet for verification
    
    print("Verification complete!")

def main():
    print("Starting comprehensive formatting process...")
    apply_comprehensive_formatting()
    verify_formatting()
    print("Process completed! Please open the Excel file to see the formatting changes.")

if __name__ == "__main__":
    main()