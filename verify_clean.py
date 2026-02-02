"""Verify the cleaned Excel data"""
import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx')
ws = wb['Data to be captured']
print('=== VERIFICATION: Sample rows after cleaning ===')
print('='*100)

for row_idx in range(3, 25):
    sr = ws.cell(row=row_idx, column=2).value
    region = ws.cell(row=row_idx, column=3).value or ""
    state = ws.cell(row=row_idx, column=4).value or ""
    subs = str(ws.cell(row=row_idx, column=5).value or "")[:25]
    dev = str(ws.cell(row=row_idx, column=7).value or "")[:30]
    gna_id = str(ws.cell(row=row_idx, column=9).value or "")[:15]
    lta_id = str(ws.cell(row=row_idx, column=10).value or "")[:20]
    col_l = str(ws.cell(row=row_idx, column=12).value or "")
    col_m = str(ws.cell(row=row_idx, column=13).value or "")
    col_n = str(ws.cell(row=row_idx, column=14).value or "")
    col_o = str(ws.cell(row=row_idx, column=15).value or "")
    
    print(f"Row {row_idx}: Sr.no={sr}")
    print(f"  Region={region}, State={state}")
    print(f"  Substation={subs}")
    print(f"  Developer={dev}")
    print(f"  GNA_ID={gna_id}, LTA_ID={lta_id}")
    print(f"  L(CMETS GNA)={col_l}, M(CMETS LTA)={col_m}")
    print(f"  N(GNA Date)={col_n}, O(LTA Date)={col_o}")
    print("-"*80)

# Check final serial number continuity
print("\n=== SERIAL NUMBER VERIFICATION ===")
print("Checking first 20 and last 20 rows...")
sr_issues = []
for row_idx in range(3, min(23, ws.max_row + 1)):
    sr = ws.cell(row=row_idx, column=2).value
    expected = row_idx - 2
    if sr != expected:
        sr_issues.append(f"Row {row_idx}: expected {expected}, got {sr}")

# Check last 20 rows
last_20_start = max(ws.max_row - 20, 3)
for row_idx in range(last_20_start, ws.max_row + 1):
    sr = ws.cell(row=row_idx, column=2).value
    if sr is not None:
        print(f"Row {row_idx}: Sr.no = {sr}")

print(f"\nTotal rows with data: checked")
if sr_issues:
    print("Serial number issues found:")
    for issue in sr_issues[:10]:
        print(f"  {issue}")
else:
    print("First 20 serial numbers are correct!")
