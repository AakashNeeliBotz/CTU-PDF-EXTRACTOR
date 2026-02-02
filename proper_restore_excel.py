import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

def restore_excel_properly():
    """
    Properly restore the Excel file maintaining original formatting and structure
    """
    original_file = 'Connectivity_Application_Data_backup.xlsx'
    current_file = 'Connectivity_Application_Data_TEST_ALL_SHEETS38.xlsx'
    output_file = 'Connectivity_Application_Data_TEST_ALL_SHEETS38.xlsx'
    
    print("Restoring Excel file with proper structure...")
    
    try:
        # Load the original file structure
        original_xl = pd.ExcelFile(original_file)
        
        # Create new workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Process each sheet
        for sheet_name in original_xl.sheet_names:
            print(f"Processing sheet: {sheet_name}")
            
            if sheet_name == 'Data to be captured':
                # Handle the main data sheet specially
                restore_data_sheet(original_file, current_file, wb, sheet_name)
            else:
                # Copy other sheets as-is from original
                copy_sheet_as_is(original_file, wb, sheet_name)
        
        # Save the workbook
        wb.save(output_file)
        print(f"\nExcel file restored successfully: {output_file}")
        
        # Verify the restoration
        verify_restoration(output_file)
        
    except Exception as e:
        print(f"Error restoring Excel file: {e}")
        import traceback
        traceback.print_exc()

def restore_data_sheet(original_file, current_file, workbook, sheet_name):
    """Restore the Data to be captured sheet properly"""
    print(f"  Restoring {sheet_name} with proper headers...")
    
    # Load original structure (headers in row 1, data from row 2)
    original_df = pd.read_excel(original_file, sheet_name=sheet_name, header=None)
    
    # Load current file to get our additions
    current_df = pd.read_excel(current_file, sheet_name=sheet_name, header=None)
    
    # Extract headers (rows 0 and 1)
    headers = original_df.iloc[:2]  # First 2 rows contain headers
    
    # Extract original data (from row 2 onwards, excluding our additions)
    original_data = original_df.iloc[2:len(original_df)-5]  # Remove last 5 rows which might be our additions
    
    # Extract our new data (last 5 rows from current file)
    new_data = current_df.tail(5)  # Get last 5 rows which contain our extracted data
    
    # Combine: headers + original data + new data
    final_df = pd.concat([headers, original_data, new_data], ignore_index=True)
    
    print(f"    Original data rows: {len(original_data)}")
    print(f"    New data rows: {len(new_data)}")
    print(f"    Final rows: {len(final_df)}")
    
    # Create worksheet
    ws = workbook.create_sheet(title=sheet_name)
    
    # Write data to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(final_df, index=False, header=False), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Auto-adjust column widths
    auto_adjust_columns(ws)
    
    print(f"    Sheet '{sheet_name}' restored successfully")

def copy_sheet_as_is(file_path, workbook, sheet_name):
    """Copy a sheet from file as-is"""
    print(f"  Copying {sheet_name} unchanged...")
    
    # Load the sheet
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    
    # Create worksheet
    ws = workbook.create_sheet(title=sheet_name)
    
    # Write data to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Auto-adjust column widths
    auto_adjust_columns(ws)
    
    print(f"    Sheet '{sheet_name}' copied ({df.shape[0]} rows, {df.shape[1]} columns)")

def auto_adjust_columns(worksheet):
    """Auto-adjust column widths"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        worksheet.column_dimensions[column_letter].width = adjusted_width

def verify_restoration(file_path):
    """Verify that the restoration worked properly"""
    print("\nVerifying restoration...")
    
    try:
        xl_file = pd.ExcelFile(file_path)
        print(f"Total sheets: {len(xl_file.sheet_names)}")
        print("Sheet names:", xl_file.sheet_names)
        
        # Check Data to be captured sheet specifically
        df_data = pd.read_excel(file_path, sheet_name='Data to be captured', header=1, nrows=3)
        print("\nData to be captured sheet (with proper headers):")
        print("Shape:", df_data.shape)
        print("First few rows:")
        print(df_data.head(2))
        print("Columns:", df_data.columns[:10].tolist(), "...")
        
        # Check that headers are preserved
        df_raw = pd.read_excel(file_path, sheet_name='Data to be captured', header=None, nrows=3)
        print("\nHeader rows verification:")
        print("Row 0 (should be mostly NaN):", df_raw.iloc[0, :5].tolist())
        print("Row 1 (actual headers):", df_raw.iloc[1, :5].tolist())
        
        print("Restoration verification completed successfully!")
        
    except Exception as e:
        print(f"Error during verification: {e}")

def main():
    print("Starting proper Excel restoration...")
    restore_excel_properly()
    print("Process completed!")

if __name__ == "__main__":
    main()