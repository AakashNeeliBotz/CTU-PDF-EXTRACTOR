
import os
import shutil
from element_status_processor import ElementStatusProcessor

PDF_PATH = r"c:\Users\Admin\Documents\CTU-automated-pdf-extraction\downloaded_pdfs\SN_TBCB\Report_TBCB_UC.pdf"
OUTPUT_FILE = r"c:\Users\Admin\Documents\CTU-automated-pdf-extraction\debug_output.xlsx"
TEMPLATE_FILE = r"c:\Users\Admin\Documents\CTU-automated-pdf-extraction\Connectivity Application Data.xlsx"

def run_debug():
    print(f"Debug Script for Element Status Processor")
    print(f"----------------------------------------")
    
    if not os.path.exists(PDF_PATH):
        print(f"ERROR: PDF not found at {PDF_PATH}")
        return

    # Create output from template
    if os.path.exists(TEMPLATE_FILE):
        shutil.copy2(TEMPLATE_FILE, OUTPUT_FILE)
        print(f"Created output file from template: {OUTPUT_FILE}")
    else:
        print(f"ERROR: Template not found at {TEMPLATE_FILE}")
        return

    # Run Processor
    try:
        proc = ElementStatusProcessor()
        print("Calling process_and_write...")
        proc.process_and_write(PDF_PATH, OUTPUT_FILE)
        print("process_and_write completed.")
        
        # Verify Output matches
        import openpyxl
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        if "Element Status" in wb.sheetnames:
            ws = wb["Element Status"]
            print(f"Sheet 'Element Status' exists.")
            print(f"Max Row: {ws.max_row}")
            print(f"Value at E4: {ws['E4'].value}")
        else:
            print("ERROR: Sheet 'Element Status' NOT found in output.")
            
    except Exception as e:
        print(f"EXCEPTION: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    import sys
    with open("debug_log.txt", "w") as f:
        sys.stdout = f
        sys.stderr = f
        run_debug()
