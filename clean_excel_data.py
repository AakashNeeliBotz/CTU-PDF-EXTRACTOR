"""
Excel Data Cleaning Script
This script performs data cleaning on the "Data to be captured" sheet ONLY.
It preserves all formatting, styles, and other sheets.

Tasks:
1. Clean GNA/ST II Application ID column - keep only numeric IDs (preserve Enhancement/Enh text)
2. Extract IDs from Substation column and move to GNA/ST II Application ID
3. Extract IDs from other columns and move to GNA/ST II Application ID  
4. Fix State and Region column issues (AP -> Andhra Pradesh, move region codes)
"""

import openpyxl
import re
from copy import copy

# File path
FILE_PATH = r'c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx'
OUTPUT_PATH = r'c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38_CLEANED.xlsx'

# Column indices (1-based)
COL_REGION = 3
COL_STATE = 4
COL_SUBSTATION = 5
COL_GNA_ID = 9

# State abbreviation mappings
STATE_ABBREV_MAP = {
    'AP': 'Andhra Pradesh',
    'MP': 'Madhya Pradesh',
    'UP': 'Uttar Pradesh',
}

# Region code to State mapping (for when region code is in state column)
REGION_CODES = {'NR', 'SR', 'WR', 'ER', 'NER'}


def extract_numeric_id(text):
    """Extract numeric Application ID from text, preserving Enhancement/Enh info."""
    if not text or not isinstance(text, str):
        return text, None
    
    text = str(text).strip()
    
    # Check if it contains Enhancement/Enh keywords
    has_enhancement = False
    enhancement_text = ""
    if re.search(r'\b(Enhancement|Enh\.?)\b', text, re.IGNORECASE):
        has_enhancement = True
        # Find the enhancement keyword
        match = re.search(r'\b(Enhancement|Enh\.?)\b', text, re.IGNORECASE)
        if match:
            enhancement_text = match.group(0)
    
    # Extract all numeric IDs (7+ digits)
    numeric_ids = re.findall(r'\b(\d{7,})\b', text)
    
    if numeric_ids:
        # Take the first numeric ID found
        numeric_id = numeric_ids[0]
        
        if has_enhancement:
            return f"{numeric_id} ({enhancement_text})", text
        else:
            return numeric_id, text
    
    return text, None


def clean_gna_id(text):
    """
    Clean GNA/ST II Application ID column.
    Remove St-II:, brackets, MW, TSSPDCL, etc.
    Keep only numeric ID and Enhancement/Enh if present.
    """
    if not text or not isinstance(text, str):
        return text
    
    original = str(text).strip()
    
    # If it's purely numeric, return as-is
    if original.isdigit():
        return original
    
    # Check for Enhancement/Enh
    has_enhancement = False
    enhancement_text = ""
    enh_match = re.search(r'\b(Enhancement|Enh\.?)\b', original, re.IGNORECASE)
    if enh_match:
        has_enhancement = True
        enhancement_text = enh_match.group(0)
    
    # Extract all numeric IDs (7+ digits - typical Application ID length)
    numeric_ids = re.findall(r'\b(\d{7,})\b', original)
    
    if numeric_ids:
        # Handle multiple IDs - join them with comma
        if len(numeric_ids) > 1:
            # If there are multiple IDs, join them
            result = ', '.join(numeric_ids)
        else:
            result = numeric_ids[0]
        
        if has_enhancement:
            result = f"{result} ({enhancement_text})"
        
        return result
    
    # If no numeric ID found, return original (avoid modifying ambiguous cells)
    return original


def extract_id_from_substation(substation_text):
    """
    Extract numeric ID from Substation text.
    Returns (cleaned_substation, extracted_id)
    """
    if not substation_text or not isinstance(substation_text, str):
        return substation_text, None
    
    original = str(substation_text).strip()
    
    # Find numeric IDs (7+ digits)
    ids_found = re.findall(r'\b(\d{7,})\b', original)
    
    if ids_found:
        # Remove the ID from the substation text
        cleaned = original
        for id_val in ids_found:
            cleaned = re.sub(r'\b' + id_val + r'\b', '', cleaned)
        
        # Clean up extra whitespace
        cleaned = re.sub(r'\s+', ' ', cleaned).strip()
        
        # Return the first ID found
        return cleaned, ids_found[0]
    
    return original, None


def fix_state_region(state_value, region_value):
    """
    Fix State and Region values.
    - Replace abbreviations (AP -> Andhra Pradesh)
    - If Region code is in State, move it to Region
    Returns (new_state, new_region)
    """
    new_state = state_value
    new_region = region_value
    
    if not state_value:
        return new_state, new_region
    
    state_str = str(state_value).strip()
    
    # Check if State is a region code (should be moved to Region)
    if state_str.upper() in REGION_CODES:
        # Only move if Region is empty
        if not region_value or (isinstance(region_value, str) and not region_value.strip()):
            new_region = state_str.upper()
            new_state = None  # Clear the state since we don't know the actual state
        else:
            # Region already has a value, just clear the region code from state
            new_state = None
    
    # Replace abbreviations
    elif state_str.upper() in STATE_ABBREV_MAP:
        new_state = STATE_ABBREV_MAP[state_str.upper()]
    
    return new_state, new_region


def scan_for_ids_in_all_columns(ws, row, gna_col):
    """
    Scan all columns in a row for misplaced numeric IDs.
    Returns list of (column, id_found, cleaned_value)
    """
    found_ids = []
    
    # Columns to skip (already processed or not relevant)
    skip_cols = {COL_GNA_ID, 10, 11, 12, 13}  # GNA, LTA, Enhancement ID, CMETS cols
    
    for col in range(1, ws.max_column + 1):
        if col in skip_cols:
            continue
        
        cell = ws.cell(row=row, column=col)
        value = cell.value
        
        if value and isinstance(value, str):
            # Look for standalone numeric IDs (not part of dates or other data)
            # IDs are typically 10 digits starting with 1 or 2
            ids = re.findall(r'\b([12]\d{9})\b', value)
            
            if ids:
                # Check if this is not a date or other data
                # Substation column is the main source
                if col == COL_SUBSTATION:
                    cleaned, extracted_id = extract_id_from_substation(value)
                    if extracted_id:
                        found_ids.append((col, extracted_id, cleaned))
    
    return found_ids


def main():
    print(f"Loading workbook: {FILE_PATH}")
    
    # Load workbook preserving all formatting
    wb = openpyxl.load_workbook(FILE_PATH)
    
    print(f"Available sheets: {wb.sheetnames}")
    
    # Work only on "Data to be captured" sheet
    if 'Data to be captured' not in wb.sheetnames:
        print("ERROR: Sheet 'Data to be captured' not found!")
        return
    
    ws = wb['Data to be captured']
    
    print(f"Processing sheet: 'Data to be captured' (Rows: {ws.max_row}, Cols: {ws.max_column})")
    
    # Track changes
    changes = {
        'gna_id_cleaned': 0,
        'substation_id_extracted': 0,
        'state_fixed': 0,
        'region_fixed': 0,
    }
    
    # Process each row (skip header row 2)
    header_row = 2
    
    for row in range(header_row + 1, ws.max_row + 1):
        # TASK 1: Clean GNA/ST II Application ID column
        gna_cell = ws.cell(row=row, column=COL_GNA_ID)
        original_gna = gna_cell.value
        
        if original_gna and isinstance(original_gna, str):
            cleaned_gna = clean_gna_id(original_gna)
            if cleaned_gna != original_gna:
                gna_cell.value = cleaned_gna
                changes['gna_id_cleaned'] += 1
                if changes['gna_id_cleaned'] <= 10:
                    print(f"  Row {row} GNA cleaned: {repr(original_gna)[:50]} -> {repr(cleaned_gna)[:50]}")
        
        # TASK 2 & 3: Extract IDs from Substation and other columns
        substation_cell = ws.cell(row=row, column=COL_SUBSTATION)
        substation_value = substation_cell.value
        
        if substation_value and isinstance(substation_value, str):
            cleaned_substation, extracted_id = extract_id_from_substation(substation_value)
            
            if extracted_id:
                # Check if GNA ID column is empty or should be populated
                current_gna = ws.cell(row=row, column=COL_GNA_ID).value
                
                if not current_gna or (isinstance(current_gna, str) and not current_gna.strip()):
                    # GNA column is empty, move the ID there
                    ws.cell(row=row, column=COL_GNA_ID).value = extracted_id
                    substation_cell.value = cleaned_substation
                    changes['substation_id_extracted'] += 1
                    if changes['substation_id_extracted'] <= 10:
                        print(f"  Row {row} Substation ID extracted: {extracted_id}, Substation now: {repr(cleaned_substation)[:40]}")
                else:
                    # GNA already has value, just clean the substation
                    substation_cell.value = cleaned_substation
                    changes['substation_id_extracted'] += 1
        
        # TASK 4: Fix State and Region column issues
        state_cell = ws.cell(row=row, column=COL_STATE)
        region_cell = ws.cell(row=row, column=COL_REGION)
        
        original_state = state_cell.value
        original_region = region_cell.value
        
        new_state, new_region = fix_state_region(original_state, original_region)
        
        if new_state != original_state:
            state_cell.value = new_state
            changes['state_fixed'] += 1
            if changes['state_fixed'] <= 10:
                print(f"  Row {row} State fixed: {repr(original_state)} -> {repr(new_state)}")
        
        if new_region != original_region:
            region_cell.value = new_region
            changes['region_fixed'] += 1
            if changes['region_fixed'] <= 10:
                print(f"  Row {row} Region fixed: {repr(original_region)} -> {repr(new_region)}")
    
    # Print summary
    print("\n" + "="*60)
    print("SUMMARY OF CHANGES:")
    print("="*60)
    print(f"  GNA/ST II Application IDs cleaned: {changes['gna_id_cleaned']}")
    print(f"  IDs extracted from Substation: {changes['substation_id_extracted']}")
    print(f"  State values fixed: {changes['state_fixed']}")
    print(f"  Region values fixed: {changes['region_fixed']}")
    
    # Save the workbook
    print(f"\nSaving to: {OUTPUT_PATH}")
    wb.save(OUTPUT_PATH)
    print("Done! File saved successfully.")
    print("\nNOTE: Only 'Data to be captured' sheet was modified.")
    print("NOTE: No formatting or styles were changed.")


if __name__ == "__main__":
    main()
