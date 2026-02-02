import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

def update_excel_with_meeting_dates():
    '''Update Excel with meeting dates from PDFs'''
    
    # Excel file path
    excel_path = 'c:/Users/Sree Charan/Desktop/fold2/CTU-PDF-EXTRACTOR/Connectivity_Application_Data_TEST_ALL_SHEETS38.xlsx'
    
    # Meeting dates from PDF filenames
    dates_map = {33: '05.08.2024', 34: '20.09.2024'}
    
    print('=== Updating Excel with Meeting Dates ===')
    print(f'PDF #33 Date: {dates_map[33]}')
    print(f'PDF #34 Date: {dates_map[34]}')
    
    # Load Excel file
    wb = load_workbook(excel_path)
    ws = wb['Data to be captured']
    
    # Find column indices (assuming row 2 contains headers)
    gna_col_idx = None
    lta_col_idx = None
    cmets_gna_date_col_idx = None
    cmets_lta_date_col_idx = None
    cmets_gna_approved_col_idx = 12  # CMETS GNA Approved is column 12
    cmets_lta_approved_col_idx = 13  # CMETS LTA Approved is column 13
    
    # Check row 2 for headers
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=2, column=col).value
        if cell_value:
            header_text = str(cell_value).strip().lower()
            if 'gna/st ii application id' in header_text:
                gna_col_idx = col
            elif 'lta application id' in header_text:
                lta_col_idx = col
            elif 'cmets gna meeting date' in header_text:
                cmets_gna_date_col_idx = col
            elif 'cmets lta meeting date' in header_text:
                cmets_lta_date_col_idx = col
    
    print(f'Column indices found:')
    print(f'  GNA/ST II Application ID: {gna_col_idx}')
    print(f'  LTA Application ID: {lta_col_idx}')
    print(f'  CMETS GNA Meeting Date: {cmets_gna_date_col_idx}')
    print(f'  CMETS LTA Meeting Date: {cmets_lta_date_col_idx}')
    print(f'  CMETS GNA Approved: {cmets_gna_approved_col_idx}')
    print(f'  CMETS LTA Approved: {cmets_lta_approved_col_idx}')
    
    # Process each row in the Excel sheet
    updated_gna_dates = 0
    updated_lta_dates = 0
    
    for row in range(3, ws.max_row + 1):  # Start from row 3 (data rows)
        # Get CMETS approval values from current row
        cmets_gna_approved = ws.cell(row=row, column=cmets_gna_approved_col_idx).value
        cmets_lta_approved = ws.cell(row=row, column=cmets_lta_approved_col_idx).value
        
        # Update meeting dates based on which PDF the application was approved in
        if cmets_gna_approved is not None:
            pdf_num = int(cmets_gna_approved)  # Convert to int (e.g., 34.0 -> 34)
            if pdf_num in dates_map and cmets_gna_date_col_idx:
                gna_date_cell = ws.cell(row=row, column=cmets_gna_date_col_idx)
                gna_date_cell.value = dates_map[pdf_num]
                updated_gna_dates += 1
        
        if cmets_lta_approved is not None:
            pdf_num = int(cmets_lta_approved)  # Convert to int (e.g., 34.0 -> 34)
            if pdf_num in dates_map and cmets_lta_date_col_idx:
                lta_date_cell = ws.cell(row=row, column=cmets_lta_date_col_idx)
                lta_date_cell.value = dates_map[pdf_num]
                updated_lta_dates += 1
    
    print(f'\nUpdates made:')
    print(f'  CMETS GNA Meeting Date updated: {updated_gna_dates} rows')
    print(f'  CMETS LTA Meeting Date updated: {updated_lta_dates} rows')
    
    # Save the workbook
    wb.save(excel_path)
    print(f'\nExcel file updated successfully: {excel_path}')

if __name__ == "__main__":
    update_excel_with_meeting_dates()