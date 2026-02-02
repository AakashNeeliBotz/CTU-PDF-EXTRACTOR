"""
Smart Fill for CMETS GNA Approved and Meeting Date
Strategy:
1. Group rows by Normalized Developer Name + Normalized Substation
2. For each group:
   - Identify the 'best' CMETS value (Number & Date)
   - Rule 1: If all existing values are same, use that.
   - Rule 2: If values are subsets (e.g. '33' vs '33, 34'), use the superset ('33, 34').
   - Rule 3: If values are disjoint/conflicting (e.g. '33' vs '40'), DO NOT TOUCH.
3. specific fix for known incomplete rows if they match the pattern.
"""

import openpyxl
import re
from collections import defaultdict

FILE_PATH = r'c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx'

def normalize(text):
    if not text:
        return ""
    # Lowercase, remove special chars, collapse spaces
    text = str(text).lower()
    text = re.sub(r'[^a-z0-9]', '', text)
    return text

def is_superset(val1, val2):
    """Check if val1 is a superset of val2 (logic for comma-sep values)"""
    if not val1 or not val2:
        return False
    parts1 = set(p.strip() for p in val1.split(','))
    parts2 = set(p.strip() for p in val2.split(','))
    return parts1.issuperset(parts2) and len(parts1) > len(parts2)

def process_excel():
    print("Loading workbook...")
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb['Data to be captured']
    
    COL_SUBSTATION = 5
    COL_DEVELOPER = 7
    COL_CMETS_GNA = 12
    COL_CMETS_DATE = 14
    
    # 1. Build Groups
    groups = defaultdict(list)
    
    print("Grouping data...")
    for row in range(3, ws.max_row + 1):
        dev = ws.cell(row=row, column=COL_DEVELOPER).value
        sub = ws.cell(row=row, column=COL_SUBSTATION).value
        
        if dev:
            # Create a key that is loose enough to catch variations but strict enough to be safe
            # Use strict normalization for grouping
            key = (normalize(dev), normalize(sub))
            groups[key].append(row)
            
    updates = 0
    
    print("Analyzing groups and filling data...")
    for key, rows in groups.items():
        # Collect existing values
        values = []
        for r in rows:
            c_num = ws.cell(row=r, column=COL_CMETS_GNA).value
            c_date = ws.cell(row=r, column=COL_CMETS_DATE).value
            if c_num or c_date:
                values.append((str(c_num) if c_num else "", str(c_date) if c_date else ""))
        
        if not values:
            continue
            
        # Determine Target Value
        unique_values = set(values)
        target_val = None
        
        if len(unique_values) == 1:
            target_val = list(unique_values)[0]
        else:
            # Check for superset logic
            sorted_vals = sorted(list(unique_values), key=lambda x: len(x[0]), reverse=True)
            candidate = sorted_vals[0]
            
            is_valid_superset = True
            for other in sorted_vals[1:]:
                # Check number part
                if not is_superset(candidate[0], other[0]) and candidate[0] != other[0]:
                     # Allow if one is just missing (empty strings handled)
                     if other[0] != "":
                         is_valid_superset = False
                         break
            
            if is_valid_superset:
                target_val = candidate
        
        # Apply updates if we have a safe target value
        if target_val:
            t_num, t_date = target_val
            for r in rows:
                curr_num = ws.cell(row=r, column=COL_CMETS_GNA).value
                curr_date = ws.cell(row=r, column=COL_CMETS_DATE).value
                
                # Update if empty OR if we are upgrading a subset to a superset
                needs_update = False
                
                if not curr_num and t_num:
                    ws.cell(row=r, column=COL_CMETS_GNA).value = t_num
                    needs_update = True
                    
                if not curr_date and t_date:
                    ws.cell(row=r, column=COL_CMETS_DATE).value = t_date
                    needs_update = True
                    
                # Upgrade logic: if current is "33" and target is "33, 34"
                if curr_num and t_num and curr_num != t_num and is_superset(t_num, str(curr_num)):
                    ws.cell(row=r, column=COL_CMETS_GNA).value = t_num
                    needs_update = True
                
                if needs_update:
                    updates += 1
    
    print(f"\nTotal cells updated: {updates}")
    
    print(f"Saving workbook...")
    wb.save(FILE_PATH)
    print("Done!")

if __name__ == "__main__":
    process_excel()
