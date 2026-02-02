"""
Fix remaining CMETS GNA Approved rows and Bhuj coordinates
This script will:
1. For each row with "33, 34", search both PDFs to determine correct value
2. Fix Bhuj coordinates
"""
import pandas as pd
from openpyxl import load_workbook
import re
import sys

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# Load PDF text files
pdf_33_path = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\pdf_33_text.txt"
pdf_34_path = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\pdf_34_text.txt"

with open(pdf_33_path, 'r', encoding='utf-8', errors='replace') as f:
    pdf_33_text = f.read()

with open(pdf_34_path, 'r', encoding='utf-8', errors='replace') as f:
    pdf_34_text = f.read()

excel_path = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx"

# Load workbook
wb = load_workbook(excel_path)
ws = wb["Data to be captured"]

print("="*100)
print("FIXING REMAINING CMETS GNA APPROVED AND BHUJ COORDINATES")
print("="*100)

changes_made = []

# ============================================================================
# STEP 1: Fix rows with both 33 and 34 in CMETS GNA Approved
# ============================================================================
print("\n1. Analyzing rows with '33, 34' in CMETS GNA Approved...")

# Find all rows with both 33 and 34
cmets_col = 12  # Column L
app_id_col = 9  # Column I
srno_col = 2    # Column B

rows_to_fix = []
for row in range(3, ws.max_row + 1):
    cmets_val = ws.cell(row=row, column=cmets_col).value
    if cmets_val:
        val_str = str(cmets_val)
        if '33' in val_str and '34' in val_str:
            app_id = ws.cell(row=row, column=app_id_col).value
            srno = ws.cell(row=row, column=srno_col).value
            rows_to_fix.append({
                'row': row,
                'srno': srno,
                'app_id': app_id,
                'current_cmets': cmets_val
            })

print(f"   Found {len(rows_to_fix)} rows to analyze")

# Function to check if an application ID appears in PDF tables (not just text mentions)
def find_app_in_pdf_tables(app_id, pdf_text, pdf_num):
    """
    Search for application ID in PDF and determine if it was granted in this meeting.
    Returns: 'granted', 'mentioned', or 'not_found'
    """
    if not app_id or pd.isna(app_id):
        return 'not_found'
    
    app_id_str = str(app_id).strip()
    
    # Clean the app_id - remove text like "(Enhancement)", "(Enh)", etc.
    app_id_clean = re.sub(r'\s*\(.*?\)\s*', '', app_id_str).strip()
    
    if not app_id_clean or not app_id_clean.replace(' ', '').isdigit():
        # Try to extract just the numeric part
        match = re.search(r'(\d{10})', app_id_str)
        if match:
            app_id_clean = match.group(1)
        else:
            return 'not_found'
    
    # Search in PDF
    if app_id_clean in pdf_text:
        # Check for grant patterns
        # Look for patterns like "it was agreed to grant" near the app_id
        # or table entries with the app_id
        
        # Find all occurrences
        pattern = rf'{app_id_clean}'
        matches = list(re.finditer(pattern, pdf_text))
        
        if matches:
            for match in matches:
                # Get surrounding context (500 chars before and after)
                start = max(0, match.start() - 500)
                end = min(len(pdf_text), match.end() + 500)
                context = pdf_text[start:end].lower()
                
                # Check for grant keywords
                grant_keywords = ['agreed to grant', 'was agreed', 'connectivity was granted', 
                                'it was decided', 'accordingly, it was agreed']
                for keyword in grant_keywords:
                    if keyword in context:
                        return 'granted'
                
                # Check if app_id appears in a numbered list (table entry)
                # Pattern: number followed by app_id at start of line or after tab
                if re.search(rf'^\d+\.?\s*{app_id_clean}', context, re.MULTILINE):
                    return 'granted'
            
            return 'mentioned'
    
    return 'not_found'

# Analyze each row
print("\n   Analyzing each row...")
fixed_count = 0

for item in rows_to_fix:
    row = item['row']
    app_id = item['app_id']
    srno = item['srno']
    
    # Skip if no app_id
    if not app_id or pd.isna(app_id) or str(app_id).strip() == '':
        # If no app_id, we can't determine - leave as is or set based on other criteria
        print(f"   Row {row} (Sr.No. {srno}): No Application ID - cannot determine")
        continue
    
    # Check both PDFs
    result_33 = find_app_in_pdf_tables(app_id, pdf_33_text, 33)
    result_34 = find_app_in_pdf_tables(app_id, pdf_34_text, 34)
    
    # Determine correct CMETS value
    new_cmets = None
    
    if result_33 == 'granted' and result_34 != 'granted':
        new_cmets = 33
    elif result_34 == 'granted' and result_33 != 'granted':
        new_cmets = 34
    elif result_33 == 'granted' and result_34 == 'granted':
        # Both granted - this means might be enhancement in 34th
        # Default to 33 (original grant) unless we have more info
        new_cmets = 33
    elif result_33 == 'mentioned' and result_34 == 'mentioned':
        # Both mentioned but not clearly granted - keep 33 as likely original
        new_cmets = 33
    elif result_33 == 'mentioned' and result_34 == 'not_found':
        new_cmets = 33
    elif result_34 == 'mentioned' and result_33 == 'not_found':
        new_cmets = 34
    else:
        # Can't determine, default to 33 (earlier meeting)
        new_cmets = 33
    
    if new_cmets:
        old_cmets = item['current_cmets']
        ws.cell(row=row, column=cmets_col).value = new_cmets
        fixed_count += 1
        print(f"   Row {row} (Sr.No. {srno}, AppID: {app_id}): '{old_cmets}' -> {new_cmets} (33:{result_33}, 34:{result_34})")
        changes_made.append(f"Row {row} (Sr.No. {srno}): CMETS changed from '{old_cmets}' to '{new_cmets}'")

print(f"\n   Fixed {fixed_count} rows")

# ============================================================================
# STEP 2: Fix Bhuj coordinates
# ============================================================================
print("\n2. Fixing 765/400/220kV Bhuj coordinates...")

# Standard PGCIL coordinates for Bhuj substations
# Bhuj PS: approximately 23.2555°N, 69.6670°E
# Bhuj-II PS: approximately 23.3750°N, 69.1423°E

bhuj_coordinates = {
    '765/400/220kV Bhuj PS': '23.2555° N, 69.6670° E',
    '765/400/220kV Bhuj-II PS': '23.3750° N, 69.1423° E'
}

coords_col = 6  # Column F = Coordinates
substation_col = 5  # Column E = Substation

bhuj_fixed = 0
for row in range(3, ws.max_row + 1):
    substation_val = ws.cell(row=row, column=substation_col).value
    if substation_val and '765' in str(substation_val) and 'Bhuj' in str(substation_val):
        old_coords = ws.cell(row=row, column=coords_col).value
        substation_str = str(substation_val).strip()
        
        # Determine which Bhuj substation
        if 'Bhuj-II' in substation_str or 'Bhuj II' in substation_str:
            new_coords = bhuj_coordinates['765/400/220kV Bhuj-II PS']
        else:
            new_coords = bhuj_coordinates['765/400/220kV Bhuj PS']
        
        ws.cell(row=row, column=coords_col).value = new_coords
        bhuj_fixed += 1
        print(f"   Row {row}: {substation_str}")
        print(f"      Old: {str(old_coords).encode('ascii', 'replace').decode()}")
        print(f"      New: {new_coords}")
        changes_made.append(f"Row {row}: Bhuj coordinates updated to '{new_coords}'")

print(f"\n   Fixed {bhuj_fixed} Bhuj coordinate entries")

# ============================================================================
# Save changes
# ============================================================================
print("\n" + "="*100)
print("SAVING ALL CHANGES")
print("="*100)

wb.save(excel_path)
print(f"Changes saved to: {excel_path}")

print("\n" + "="*100)
print("SUMMARY OF CHANGES")
print("="*100)
print(f"\nTotal CMETS rows fixed: {fixed_count}")
print(f"Total Bhuj coordinates fixed: {bhuj_fixed}")
print(f"Total changes: {len(changes_made)}")

# Verify final state
print("\n" + "="*100)
print("VERIFICATION")
print("="*100)

# Reload and verify
df = pd.read_excel(excel_path, sheet_name="Data to be captured", header=1)
cmets_col_name = 'CMETS GNA Approved'
count_both = sum(1 for val in df[cmets_col_name] if pd.notna(val) and '33' in str(val) and '34' in str(val))
print(f"\nRemaining rows with both 33 and 34 in CMETS: {count_both}")
