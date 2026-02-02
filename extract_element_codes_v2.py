"""
Improved extraction script with better Element Status parsing and PDF section extraction
"""
import pandas as pd
import fitz  # PyMuPDF
import re
from openpyxl import load_workbook
from difflib import SequenceMatcher
import warnings
warnings.filterwarnings('ignore')

# File paths
EXCEL_PATH = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx"
PDF_33 = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\downloaded_pdfs\SN1\172381548953Minutes of 33rd CMETS NR meeting held on 05.08.2024.pdf"
PDF_34 = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\downloaded_pdfs\SN1\172838877090Minutes of meeting 34th CMETS NR Meeting held on 20-9-24.pdf"

# Column positions (0-indexed, will add +1 for openpyxl 1-indexed)
COL_APP_ID = 8   # GNA/ST II Application ID
COL_CTS = 39     # CTS Element Unique Code
COL_ATS = 40     # ATS Element Unique Code
COL_DTL = 41     # DTL Element Code Unique
DATA_START_ROW = 3  # Data starts at row 3 in Excel (1-indexed)

def extract_pdf_text(pdf_path):
    """Extract text from PDF"""
    doc = fitz.open(pdf_path)
    full_text = ""
    for page in doc:
        full_text += page.get_text("text")
    doc.close()
    return full_text

def load_element_status():
    """Load Element Status sheet with correct parsing"""
    print("[*] Loading Element Status sheet...")
    df = pd.read_excel(EXCEL_PATH, sheet_name='Element Status', header=None)
    
    element_map = {}
    code_to_desc = {}
    
    # Scan all rows to find element codes
    for i in range(len(df)):
        for j in range(min(5, len(df.columns))):  # Check first 5 columns
            cell_val = df.iloc[i, j]
            if pd.notna(cell_val):
                cell_str = str(cell_val).strip()
                # Element code pattern: EL-XXXXX (5 alphanumeric chars after EL-)
                if re.match(r'^EL-[A-Z0-9]{5}$', cell_str):
                    # This is an element code, get description from column 3 or 4
                    desc_col = 3 if j == 0 else 4
                    if desc_col < len(df.columns):
                        desc = df.iloc[i, desc_col]
                        if pd.notna(desc):
                            desc_str = str(desc).strip()
                            element_map[normalize_text(desc_str)] = cell_str
                            code_to_desc[cell_str] = desc_str
    
    print(f"[+] Loaded {len(code_to_desc)} Element Codes")
    
    if code_to_desc:
        print("    Sample mappings:")
        for code, desc in list(code_to_desc.items())[:5]:
            print(f"      {code}: {desc[:60]}...")
    
    return element_map, code_to_desc

def normalize_text(text):
    """Normalize text for matching"""
    if not text:
        return ""
    text = str(text).lower()
    text = ' '.join(text.split())
    text = re.sub(r'[^\w\s\-/]', ' ', text)
    return text.strip()

def similarity_score(s1, s2):
    """Calculate similarity between two strings"""
    return SequenceMatcher(None, s1, s2).ratio()

def find_element_codes(description, element_map, code_to_desc, threshold=0.65):
    """Find matching Element Codes for a description"""
    if not description or pd.isna(description):
        return []
    
    desc_norm = normalize_text(description)
    matches = []
    
    # Try exact match first
    if desc_norm in element_map:
        matches.append(element_map[desc_norm])
        return matches
    
    # Try substring matching - both directions
    for key, code in element_map.items():
        if len(key) > 20:  # Only for substantial descriptions
            if desc_norm in key or key in desc_norm:
                matches.append(code)
    
    if matches:
        return list(set(matches))
    
    # Fuzzy matching
    best_matches = []
    for key, code in element_map.items():
        score = similarity_score(desc_norm, key)
        if score >= threshold:
            best_matches.append((score, code))
    
    best_matches.sort(reverse=True)
    return [code for score, code in best_matches[:3]]  # Return top 3 matches

def extract_transmission_sections(pdf_text):
    """
    Extract ATS, DTL, CTS sections for each Application ID from PDF text.
    Returns: dict {app_id: {'ATS': str, 'DTL': str, 'CTS': str}}
    """
    results = {}
    
    # Split by "Details of Transmission system for Connectivity under GNA:"
    sections = re.split(
        r'Details of Transmission system for Connectivity under GNA:',
        pdf_text, 
        flags=re.IGNORECASE
    )
    
    for i, section in enumerate(sections[1:], 1):  # Skip first (before any section)
        # Find application IDs in context around this section
        # Look at text BEFORE this section to find which App IDs this applies to
        prev_text = sections[i-1] if i-1 < len(sections) else ""
        
        # Find application IDs in previous 2000 chars
        context = prev_text[-2000:] if len(prev_text) > 2000 else prev_text
        app_ids = set(re.findall(r'\b22\d{8}\b', context))
        
        # Also check beginning of this section
        app_ids.update(re.findall(r'\b22\d{8}\b', section[:1500]))
        
        # Extract ATS
        ats_text = ""
        ats_match = re.search(
            r'A\.\s*Associated\s*Transmission\s*System\s*\(ATS\)[:\s]*(.+?)(?=B\.\s*Transmission|$)',
            section, re.DOTALL | re.IGNORECASE
        )
        if ats_match:
            ats_text = clean_text(ats_match.group(1))
        
        # Extract DTL (Section B)
        dtl_text = ""
        dtl_match = re.search(
            r'B\.\s*Transmission\s*System\s*under\s*applicant\s*scope[:\s]*(.+?)(?=C\.\s*Transmission|$)',
            section, re.DOTALL | re.IGNORECASE
        )
        if dtl_match:
            dtl_raw = dtl_match.group(1)
            # Extract the actual line/substation descriptions
            items = re.findall(
                r'\([ivx]+\)\.*\s*(.+?)(?=\([ivx]+\)\.|C\.|Minutes of|$)',
                dtl_raw, re.DOTALL | re.IGNORECASE
            )
            if items:
                dtl_text = clean_text(items[0])
            else:
                dtl_text = clean_text(dtl_raw[:800])
        
        # Extract CTS (Section C)
        cts_text = ""
        cts_match = re.search(
            r'C\.\s*Transmission\s*system\s*for\s*Connectivity\s*under\s*GNA[:\s]*(.+?)(?=Start Date|Sl\.\s*No|Minutes of|$)',
            section, re.DOTALL | re.IGNORECASE
        )
        if cts_match:
            cts_text = clean_text(cts_match.group(1))
        
        # Store for each applicable App ID
        for app_id in app_ids:
            if app_id not in results:
                results[app_id] = {'ATS': '', 'DTL': '', 'CTS': ''}
            
            if ats_text and not results[app_id]['ATS']:
                results[app_id]['ATS'] = ats_text
            if dtl_text and not results[app_id]['DTL']:
                results[app_id]['DTL'] = dtl_text
            if cts_text and not results[app_id]['CTS']:
                results[app_id]['CTS'] = cts_text
    
    return results

def clean_text(text):
    """Clean extracted text"""
    if not text:
        return ""
    # Remove page headers
    text = re.sub(r'Minutes of \d+\w* Consultation.*?Page \d+ of \d+', '', text, flags=re.DOTALL | re.IGNORECASE)
    # Remove table headers
    text = re.sub(r'Sl\.\s*No\..*?Conn BGs requirement', '', text, flags=re.DOTALL | re.IGNORECASE)
    # Clean whitespace
    text = ' '.join(text.split())
    return text.strip()

def update_excel(pdf_sections, element_map, code_to_desc):
    """Update Excel with Element Codes"""
    print("\n[*] Opening Excel workbook...")
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Data to be captured']
    
    updates = {'CTS': 0, 'ATS': 0, 'DTL': 0}
    processed_ids = set()
    
    # Track seen Application IDs for duplicate handling
    seen_app_ids = set()
    
    for row in range(DATA_START_ROW, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=COL_APP_ID + 1).value
        
        if not cell_val:
            continue
        
        # Extract clean Application ID
        app_id_raw = str(cell_val).strip()
        app_id = re.sub(r'\D', '', app_id_raw.split()[0])[:10]
        
        if not app_id or len(app_id) != 10:
            continue
        
        # Handle duplicates - skip if already seen
        if app_id in seen_app_ids:
            continue
        seen_app_ids.add(app_id)
        
        # Check if this App ID has data from PDFs
        if app_id not in pdf_sections:
            continue
        
        section_data = pdf_sections[app_id]
        
        # Get existing values
        cts_val = ws.cell(row=row, column=COL_CTS + 1).value
        ats_val = ws.cell(row=row, column=COL_ATS + 1).value
        dtl_val = ws.cell(row=row, column=COL_DTL + 1).value
        
        # Update CTS if empty
        if not cts_val and section_data['CTS']:
            cts_codes = find_element_codes(section_data['CTS'], element_map, code_to_desc)
            if cts_codes:
                ws.cell(row=row, column=COL_CTS + 1).value = ','.join(sorted(set(cts_codes)))
                updates['CTS'] += 1
                print(f"  Row {row} (App {app_id}): CTS = {','.join(cts_codes)}")
        
        # Update ATS if empty
        if not ats_val:
            ats_text = section_data['ATS']
            if ats_text and 'NIL' not in ats_text.upper():
                ats_codes = find_element_codes(ats_text, element_map, code_to_desc)
                if ats_codes:
                    ws.cell(row=row, column=COL_ATS + 1).value = ','.join(sorted(set(ats_codes)))
                    updates['ATS'] += 1
                    print(f"  Row {row} (App {app_id}): ATS = {','.join(ats_codes)}")
            # Leave blank if NIL (per requirements)
        
        # Update DTL if empty
        if not dtl_val and section_data['DTL']:
            dtl_codes = find_element_codes(section_data['DTL'], element_map, code_to_desc)
            if dtl_codes:
                ws.cell(row=row, column=COL_DTL + 1).value = ','.join(sorted(set(dtl_codes)))
                updates['DTL'] += 1
                print(f"  Row {row} (App {app_id}): DTL = {','.join(dtl_codes)}")
    
    print(f"\n[*] Saving workbook...")
    wb.save(EXCEL_PATH)
    print("[+] Saved successfully!")
    
    return updates

def main():
    print("=" * 70)
    print("CTU PDF Extractor - Element Code Mapping (v2)")
    print("=" * 70)
    
    # Step 1: Load Element Status mapping
    element_map, code_to_desc = load_element_status()
    
    if not element_map:
        print("[!] ERROR: No element codes loaded. Check Element Status sheet.")
        return
    
    # Step 2: Extract from PDFs
    print("\n[*] Extracting from PDF 33...")
    text_33 = extract_pdf_text(PDF_33)
    sections_33 = extract_transmission_sections(text_33)
    print(f"[+] Found sections for {len(sections_33)} Application IDs")
    
    print("\n[*] Extracting from PDF 34...")
    text_34 = extract_pdf_text(PDF_34)
    sections_34 = extract_transmission_sections(text_34)
    print(f"[+] Found sections for {len(sections_34)} Application IDs")
    
    # Merge PDF data
    all_sections = sections_33.copy()
    for app_id, data in sections_34.items():
        if app_id not in all_sections:
            all_sections[app_id] = data
        else:
            # Merge - prefer non-empty values
            for key in ['ATS', 'DTL', 'CTS']:
                if data[key] and not all_sections[app_id][key]:
                    all_sections[app_id][key] = data[key]
    
    print(f"\n[*] Total unique Application IDs from PDFs: {len(all_sections)}")
    
    # Show sample extracted data
    print("\nSample extracted sections:")
    for app_id in list(all_sections.keys())[:3]:
        print(f"\n  App ID: {app_id}")
        for key in ['ATS', 'DTL', 'CTS']:
            val = all_sections[app_id][key]
            val_preview = val[:80] + '...' if len(val) > 80 else val
            print(f"    {key}: {val_preview}")
    
    # Step 3: Update Excel
    updates = update_excel(all_sections, element_map, code_to_desc)
    
    print("\n" + "=" * 70)
    print("SUMMARY:")
    print(f"  - CTS updates: {updates['CTS']}")
    print(f"  - ATS updates: {updates['ATS']}")
    print(f"  - DTL updates: {updates['DTL']}")
    print(f"  - Total:       {sum(updates.values())}")
    print("=" * 70)

if __name__ == "__main__":
    main()
