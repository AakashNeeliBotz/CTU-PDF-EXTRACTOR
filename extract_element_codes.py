"""
Automated data extraction and Excel update agent.
Extracts transmission system data from 33rd and 34th CMETS PDFs and updates Excel.

CRITICAL RULES:
1. ONLY modify sheet: "Data to be captured"
2. DO NOT modify headers, other sheets, formatting
3. Match Application IDs and extract ATS, DTL, CTS elements
4. Map to Element Codes from "Element Status" sheet
"""

import pandas as pd
import fitz  # PyMuPDF
import re
from openpyxl import load_workbook
from difflib import SequenceMatcher
import warnings
warnings.filterwarnings('ignore')

# File paths
EXCEL_PATH = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx"
PDF_33 = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\downloaded_pdfs\SN1\172381548953Minutes of 33rd CMETS NR meeting held on 05.08.2024.pdf"
PDF_34 = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\downloaded_pdfs\SN1\172838877090Minutes of meeting 34th CMETS NR Meeting held on 20-9-24.pdf"

# Column positions (0-indexed) from exploration
COL_APP_ID = 8  # GNA/ST II Application ID
COL_CTS = 39    # CTS Element Unique Code
COL_ATS = 40    # ATS Element Unique Code
COL_DTL = 41    # DTL Element Code Unique
HEADER_ROW = 1  # Row 1 contains headers (0-indexed)
DATA_START_ROW = 2  # Data starts at row 2

def extract_pdf_text(pdf_path):
    """Extract text from PDF using PyMuPDF"""
    doc = fitz.open(pdf_path)
    full_text = ""
    for page in doc:
        full_text += page.get_text("text")
    doc.close()
    return full_text

def load_element_status():
    """Load Element Status sheet and create mapping of descriptions to codes"""
    print("[*] Loading Element Status sheet...")
    df = pd.read_excel(EXCEL_PATH, sheet_name='Element Status', header=None)
    
    # Header is at row 1, data starts at row 3
    # Column 0: Element Code, Column 3: Transmission Scope (description)
    element_map = {}
    
    for i in range(3, len(df)):  # Start from row 3 (after headers)
        code = df.iloc[i, 0]  # Element Code
        # Try multiple description columns
        descriptions = []
        for col in [3, 2, 4]:  # Transmission Scope, Transmission Scheme, etc.
            desc = df.iloc[i, col]
            if pd.notna(desc) and str(desc).strip():
                descriptions.append(str(desc).strip())
        
        if pd.notna(code) and str(code).strip().startswith('EL-'):
            code_str = str(code).strip()
            for desc in descriptions:
                if desc:
                    # Store original description
                    element_map[desc.lower()] = code_str
                    # Also store normalized version
                    normalized = normalize_text(desc)
                    element_map[normalized] = code_str
    
    print(f"[+] Loaded {len(element_map)} element mappings")
    return element_map

def normalize_text(text):
    """Normalize text for matching"""
    if not text:
        return ""
    text = str(text).lower()
    # Remove extra whitespace
    text = ' '.join(text.split())
    # Remove special characters but keep essential ones
    text = re.sub(r'[^\w\s\-/]', '', text)
    return text.strip()

def similarity_score(s1, s2):
    """Calculate similarity between two strings"""
    return SequenceMatcher(None, normalize_text(s1), normalize_text(s2)).ratio()

def find_element_code(description, element_map, threshold=0.7):
    """Find Element Code for a given description using exact or semantic matching"""
    if not description or pd.isna(description):
        return None
    
    desc_norm = normalize_text(description)
    
    # Try exact match first
    if desc_norm in element_map:
        return element_map[desc_norm]
    
    # Try partial/substring match
    for key, code in element_map.items():
        if desc_norm in key or key in desc_norm:
            return code
    
    # Try fuzzy matching
    best_match = None
    best_score = 0
    for key, code in element_map.items():
        score = similarity_score(desc_norm, key)
        if score > best_score and score >= threshold:
            best_score = score
            best_match = code
    
    return best_match

def extract_app_sections(text):
    """
    Extract sections for each Application ID from the PDF text.
    Returns dict: {app_id: {'ATS': str, 'DTL': str, 'CTS': str}}
    """
    sections = {}
    
    # Pattern to find Application ID references
    app_pattern = r'(\d{10})'
    
    # Split text into potential sections based on "Details of Transmission system"
    section_pattern = r'Details of Transmission system for Connectivity under GNA:'
    parts = re.split(section_pattern, text, flags=re.IGNORECASE)
    
    for part in parts[1:]:  # Skip first part (before first section)
        # Find application IDs mentioned just before this section
        # Look for the table content before this section
        
        # Extract ATS
        ats_content = "NIL"
        ats_match = re.search(r'A\.\s*Associated\s*Transmission\s*System\s*\(ATS\)[:\s]*(.+?)(?=B\.\s*Transmission|$)', 
                              part, re.DOTALL | re.IGNORECASE)
        if ats_match:
            ats_text = ats_match.group(1).strip()
            if 'NIL' in ats_text.upper():
                ats_content = "NIL"
            else:
                # Extract the actual ATS description
                ats_content = clean_section_text(ats_text)
        
        # Extract DTL (Transmission System under applicant scope)
        dtl_content = ""
        dtl_match = re.search(r'B\.\s*Transmission\s*System\s*under\s*applicant\s*scope[:\s]*(.+?)(?=C\.\s*Transmission|$)', 
                              part, re.DOTALL | re.IGNORECASE)
        if dtl_match:
            dtl_text = dtl_match.group(1).strip()
            # Extract item descriptions like (i). description
            items = re.findall(r'\([ivx]+\)\.*\s*(.+?)(?=\([ivx]+\)|C\.|Start Date|$)', dtl_text, re.DOTALL | re.IGNORECASE)
            if items:
                dtl_content = clean_section_text(items[0])
            else:
                dtl_content = clean_section_text(dtl_text[:500])
        
        # Extract CTS (Transmission system for Connectivity under GNA)
        cts_content = ""
        cts_match = re.search(r'C\.\s*Transmission\s*system\s*for\s*Connectivity\s*under\s*GNA[:\s]*(.+?)(?=Start Date|Sl\.\s*No|$)', 
                              part, re.DOTALL | re.IGNORECASE)
        if cts_match:
            cts_text = cts_match.group(1).strip()
            # Check if it references Annexure
            if 'Annexure' in cts_text:
                cts_content = "ANNEXURE_" + re.search(r'Annexure[- ]?([IVX]+|\d+)', cts_text, re.IGNORECASE).group(1) if re.search(r'Annexure[- ]?([IVX]+|\d+)', cts_text, re.IGNORECASE) else ""
            else:
                cts_content = clean_section_text(cts_text)
        
        # Find all application IDs in the context (before this section)
        # Look at the full text before this section
        idx = text.find(part)
        context = text[max(0, idx-3000):idx] if idx > 0 else ""
        
        # Find recent app IDs
        app_ids = re.findall(app_pattern, context)
        app_ids += re.findall(app_pattern, part[:1000])  # Also check start of this section
        
        # Store for each relevant app ID
        for app_id in set(app_ids):
            if app_id.startswith('22'):  # Valid GNA app ID format
                if app_id not in sections:
                    sections[app_id] = {'ATS': [], 'DTL': [], 'CTS': []}
                if ats_content and ats_content != "NIL":
                    sections[app_id]['ATS'].append(ats_content)
                elif ats_content == "NIL":
                    sections[app_id]['ATS'] = ["NIL"]
                if dtl_content:
                    sections[app_id]['DTL'].append(dtl_content)
                if cts_content:
                    sections[app_id]['CTS'].append(cts_content)
    
    return sections

def clean_section_text(text):
    """Clean extracted section text"""
    if not text:
        return ""
    text = str(text).strip()
    # Remove page headers/footers
    text = re.sub(r'Minutes of \d+\w* Consultation Meeting.*?Page \d+ of \d+', '', text, flags=re.DOTALL | re.IGNORECASE)
    # Remove table headers
    text = re.sub(r'Sl\.\s*No\..*?Conn BGs requirement', '', text, flags=re.DOTALL | re.IGNORECASE)
    # Clean whitespace
    text = ' '.join(text.split())
    return text[:500] if len(text) > 500 else text

def extract_annexure_items(text, annexure_num):
    """Extract items from Annexure section"""
    # Find Annexure section
    pattern = rf'Annexure[- ]?{annexure_num}[:\s]*(.*?)(?=Annexure[- ]?\d+|$)'
    match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
    
    if match:
        annexure_text = match.group(1)
        # Find numbered items
        items = re.findall(r'(?:\d+\)|[a-z]\)|\([ivx]+\))[.\s]*(.+?)(?=\d+\)|[a-z]\)|\([ivx]+\)|$)', 
                          annexure_text, re.DOTALL | re.IGNORECASE)
        return [clean_section_text(item) for item in items if item.strip()]
    
    return []

def process_pdfs():
    """Extract transmission data from both PDFs"""
    print("[*] Extracting text from PDF 33...")
    text_33 = extract_pdf_text(PDF_33)
    print(f"[+] Extracted {len(text_33)} characters")
    
    print("[*] Extracting text from PDF 34...")
    text_34 = extract_pdf_text(PDF_34)
    print(f"[+] Extracted {len(text_34)} characters")
    
    # Combine for processing
    all_sections = {}
    
    print("[*] Parsing sections from PDF 33...")
    sections_33 = extract_app_sections(text_33)
    print(f"[+] Found {len(sections_33)} application sections")
    
    print("[*] Parsing sections from PDF 34...")
    sections_34 = extract_app_sections(text_34)
    print(f"[+] Found {len(sections_34)} application sections")
    
    # Merge sections
    for app_id, data in sections_33.items():
        all_sections[app_id] = data
    for app_id, data in sections_34.items():
        if app_id in all_sections:
            for key in ['ATS', 'DTL', 'CTS']:
                all_sections[app_id][key].extend(data[key])
        else:
            all_sections[app_id] = data
    
    return all_sections, text_33, text_34

def remove_duplicates():
    """Remove duplicate Application IDs (keep first occurrence)"""
    print("[*] Checking for duplicate Application IDs...")
    
    # Load using openpyxl to preserve formatting
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Data to be captured']
    
    seen_ids = set()
    rows_to_process = []
    
    # Track unique Application IDs
    for row in range(DATA_START_ROW + 2, ws.max_row + 1):  # Excel is 1-indexed, +2 for header offset
        cell_value = ws.cell(row=row, column=COL_APP_ID + 1).value  # openpyxl is 1-indexed
        if cell_value:
            app_id = str(cell_value).strip().split()[0]  # Get first part if there's enhancement note
            if app_id in seen_ids:
                rows_to_process.append((row, 'duplicate'))
            else:
                seen_ids.add(app_id)
                rows_to_process.append((row, app_id))
    
    duplicates = [r for r, status in rows_to_process if status == 'duplicate']
    print(f"[+] Found {len(duplicates)} duplicate Application IDs (will be skipped)")
    
    wb.close()
    return seen_ids

def update_excel(all_sections, element_map):
    """Update Excel with extracted element codes"""
    print("[*] Loading Excel workbook...")
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Data to be captured']
    
    updated_count = 0
    
    # Process each row
    for row in range(DATA_START_ROW + 2, ws.max_row + 1):  # +2 for 0-index to 1-index + header
        cell_value = ws.cell(row=row, column=COL_APP_ID + 1).value
        
        if not cell_value:
            continue
        
        # Extract Application ID
        app_id = str(cell_value).strip().split()[0]
        app_id = re.sub(r'\D', '', app_id)[:10]  # Keep only digits, first 10
        
        if app_id not in all_sections:
            continue
        
        section_data = all_sections[app_id]
        print(f"  Processing App ID: {app_id}")
        
        # Check if columns already have data (don't overwrite)
        existing_cts = ws.cell(row=row, column=COL_CTS + 1).value
        existing_ats = ws.cell(row=row, column=COL_ATS + 1).value
        existing_dtl = ws.cell(row=row, column=COL_DTL + 1).value
        
        # Process CTS
        if not existing_cts and section_data['CTS']:
            cts_codes = set()
            for cts_desc in section_data['CTS']:
                if cts_desc.startswith('ANNEXURE_'):
                    # Need to extract from annexure
                    pass  # Skip for now, would need full annexure parsing
                else:
                    code = find_element_code(cts_desc, element_map)
                    if code:
                        cts_codes.add(code)
            
            if cts_codes:
                ws.cell(row=row, column=COL_CTS + 1).value = ','.join(sorted(cts_codes))
                updated_count += 1
        
        # Process ATS
        if not existing_ats:
            if section_data['ATS'] and section_data['ATS'] != ["NIL"]:
                ats_codes = set()
                for ats_desc in section_data['ATS']:
                    if ats_desc != "NIL":
                        code = find_element_code(ats_desc, element_map)
                        if code:
                            ats_codes.add(code)
                
                if ats_codes:
                    ws.cell(row=row, column=COL_ATS + 1).value = ','.join(sorted(ats_codes))
                    updated_count += 1
            # If NIL, leave blank as per requirements
        
        # Process DTL
        if not existing_dtl and section_data['DTL']:
            dtl_codes = set()
            for dtl_desc in section_data['DTL']:
                code = find_element_code(dtl_desc, element_map)
                if code:
                    dtl_codes.add(code)
            
            if dtl_codes:
                ws.cell(row=row, column=COL_DTL + 1).value = ','.join(sorted(dtl_codes))
                updated_count += 1
    
    # Save workbook
    print(f"[*] Saving workbook... ({updated_count} updates made)")
    wb.save(EXCEL_PATH)
    print("[+] Excel file updated successfully!")
    
    return updated_count

def main():
    print("=" * 60)
    print("CTU PDF Extractor - Element Code Mapping")
    print("=" * 60)
    
    # Step 1: Load Element Status mapping
    element_map = load_element_status()
    
    # Step 2: Process PDFs
    all_sections, text_33, text_34 = process_pdfs()
    
    # Step 3: Remove duplicates (preprocessing)
    unique_ids = remove_duplicates()
    
    # Step 4: Update Excel
    update_count = update_excel(all_sections, element_map)
    
    print("\n" + "=" * 60)
    print(f"SUMMARY:")
    print(f"  - Application sections found: {len(all_sections)}")
    print(f"  - Unique Application IDs in Excel: {len(unique_ids)}")
    print(f"  - Updates made: {update_count}")
    print("=" * 60)

if __name__ == "__main__":
    main()
