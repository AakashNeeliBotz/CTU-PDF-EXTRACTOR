import re
import openpyxl
from openpyxl.styles import Alignment

def extract_capacity(capacity_str):
    """
    Apply strict capacity splitting logic:
    - If value contains only BESS -> Enter full value into Solar
    - If format: "X (BESS) Y (Solar)" -> Calculate X + Y and enter ONLY in Solar
    - If format: "X (BESS) Y (Wind)" -> Calculate X + Y and enter ONLY in Wind
    - If value contains only Solar -> Enter in Solar
    - If value contains only Wind -> Enter in Wind
    - Populate Hybrid ONLY if the source explicitly mentions "Hybrid"
    """
    solar = None
    wind = None
    hybrid = None
    
    # Clean string
    s = capacity_str.replace('\n', ' ').strip()
    
    # Check for Hybrid
    if 'hybrid' in s.lower():
        m = re.search(r'(\d+\.?\d*)', s)
        if m:
            hybrid = float(m.group(1))
        return solar, wind, hybrid

    # Check for X (BESS) Y (Solar) or vice versa
    if 'bess' in s.lower() and 'solar' in s.lower():
        values = re.findall(r'(\d+\.?\d*)\s*\(?BESS\)?|(\d+\.?\d*)\s*\(?Solar\)?', s, re.IGNORECASE)
        total = 0
        for b, sol in values:
            if b: total += float(b)
            if sol: total += float(sol)
        solar = total
    # Check for X (BESS) Y (Wind) or vice versa
    elif 'bess' in s.lower() and 'wind' in s.lower():
        values = re.findall(r'(\d+\.?\d*)\s*\(?BESS\)?|(\d+\.?\d*)\s*\(?Wind\)?', s, re.IGNORECASE)
        total = 0
        for b, w in values:
            if b: total += float(b)
            if w: total += float(w)
        wind = total
    # Only BESS
    elif 'bess' in s.lower() and 'solar' not in s.lower() and 'wind' not in s.lower():
        m = re.search(r'(\d+\.?\d*)', s)
        if m:
            solar = float(m.group(1))
    # Only Solar
    elif 'solar' in s.lower():
        m = re.search(r'(\d+\.?\d*)', s)
        if m:
            solar = float(m.group(1))
    # Only Wind
    elif 'wind' in s.lower():
        m = re.search(r'(\d+\.?\d*)', s)
        if m:
            wind = float(m.group(1))
    # Special cases for Table A4 entry 5: "150 MW (BESS)"
    elif '150 mw (bess)' in s.lower():
        solar = 150.0
    
    return solar, wind, hybrid

def extract_quantum(quantum_str):
    """Extract MW value from 'ID (XXX MW)'"""
    m = re.search(r'\((\d+\.?\d*)\s*MW\)', quantum_str)
    if m:
        return float(m.group(1))
    # Try just finding the number if brackets are weird
    m = re.search(r'(\d+\.?\d*)', quantum_str)
    if m:
        return float(m.group(1))
    return None

def process_regulation_data():
    # Data from Table A4 (Page 8)
    table_a4 = [
        {
            "id": "2200002123 (05.06.2025)",
            "applicant": "Juniper Green Cosmic Private Limited",
            "location": "Bikaner distt. Rajasthan",
            "nature": "Generating station(s), including REGS(s), without ESS",
            "granted": "Stage-II: 1200003740 (100 MW) LTA: 0412100008(100 MW)",
            "additional": "16 (BESS)",
            "date": "31.12.2025"
        },
        {
            "id": "2200002125 (05.06.2025)",
            "applicant": "Juniper Green Stellar Private Limited",
            "location": "Barmer distt. Rajasthan",
            "nature": "Generating station(s), including REGS(s), with ESS",
            "granted": "0412100010 (150 MW)",
            "additional": "28.75 (BESS)",
            "date": "30.04.2026"
        },
        {
            "id": "2200002126 (05.06.2025)",
            "applicant": "Juniper Green Stellar Private Limited",
            "location": "Barmer distt. Rajasthan",
            "nature": "(Generating station(s), including REGS(s), with ESS",
            "granted": "0412100011 (65 MW)",
            "additional": "17.50 (BESS)",
            "date": "30.04.2026"
        },
        {
            "id": "2200002127 (05.06.2025)",
            "applicant": "Juniper Green Stellar Private Limited",
            "location": "Barmer distt. Rajasthan",
            "nature": "Generating station(s), including REGS(s), with ESS",
            "granted": "0412100009 (150 MW)",
            "additional": "11.25 (BESS)",
            "date": "31.03.2026"
        },
        {
            "id": "2200002140 (11.06.2025)",
            "applicant": "Enren-I Energy Private Limited",
            "location": "Barmer distt. Rajasthan",
            "nature": "Generating station(s), including REGS(s), without ESS",
            "granted": "2200000286 (300 MW)",
            "additional": "150 MW (BESS)",
            "date": "28.02.2026"
        }
    ]

    # Data from Table A7 (Page 9)
    table_a7 = [
        {
            "id": "2200002249 (15.07.2025)",
            "applicant": "Serentica Renewables India Private Limited",
            "location": "Jaisalmer distt. Rajasthan",
            "nature": "Generating station(s), including REGS(s), without ESS",
            "granted": "0212100034 (300 MW)",
            "additional": "200 (Wind)",
            "date": "30.06.2028"
        },
        {
            "id": "2200002250 (15.07.2025)",
            "applicant": "Serentica Renewables India Private Limited",
            "location": "Jaisalmer distt. Rajasthan",
            "nature": "Generating station(s), including REGS(s), without ESS",
            "granted": "0212100036 (300 MW)",
            "additional": "200 (Wind)",
            "date": "30.06.2028"
        },
        {
            "id": "2200002251 (15.07.2025)",
            "applicant": "Serentica Renewables India Private Limited",
            "location": "Jaisalmer distt. Rajasthan",
            "nature": "Generating station(s), including REGS(s), without ESS",
            "granted": "2200000020 (300 MW)",
            "additional": "200 (Wind)",
            "date": "30.06.2028"
        },
        {
            "id": "2200002280 (22.07.2025)",
            "applicant": "ACME Heergarh Powertech Private Limited",
            "location": "Jodhpur distt. Rajasthan",
            "nature": "Generating station(s), including REGS(s), without ESS",
            "granted": "1200002471 (300 MW)",
            "additional": "300 (BESS) 240 (Solar)",
            "date": "01.11.2025, 01.01.2027"
        },
        {
            "id": "2200002294 (30.07.2025)",
            "applicant": "AM Green Energy Private Limited",
            "location": "Bikaner distt. Rajasthan",
            "nature": "Generating station(s), including REGS(s), without ESS",
            "granted": "2200000319 (300 MW)",
            "additional": "100 (Solar) 125 (BESS)",
            "date": "31.12.2027"
        }
    ]

    all_data = table_a4 + table_a7
    
    excel_path = r'c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx'
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Data to be captured']
    
    # Columns
    COL_STATE = 4
    COL_APPLICANT = 7
    COL_ENH_ID = 11
    COL_QUANTUM = 17
    COL_SOLAR = 19
    COL_WIND = 20
    COL_HYBRID = 21
    COL_DATE_ADD = 35
    COL_NATURE = 36
    
    # Find start row (first empty row after existing data)
    start_row = 1795 # Based on previous check max_row was 1794
    
    for i, entry in enumerate(all_data):
        current_row = start_row + i
        
        # Extract fields
        app_id_full = entry['id']
        applicant = entry['applicant']
        
        # Extract State (assuming last word is state or "Rajasthan" is always there)
        state_match = re.search(r'Rajasthan|Uttar Pradesh|Jammu and Kashmir|J&K', entry['location'])
        state = state_match.group(0) if state_match else entry['location']
        if state == "J&K": state = "Jammu and Kashmir"
        
        nature = entry['nature'].replace('\n', ' ').strip()
        quantum = extract_quantum(entry['granted'])
        solar, wind, hybrid = extract_capacity(entry['additional'])
        date_add = entry['date']
        
        # Update cells
        ws.cell(row=current_row, column=COL_APPLICANT).value = applicant
        ws.cell(row=current_row, column=COL_STATE).value = state
        ws.cell(row=current_row, column=COL_NATURE).value = nature
        ws.cell(row=current_row, column=COL_ENH_ID).value = app_id_full
        ws.cell(row=current_row, column=COL_QUANTUM).value = quantum
        ws.cell(row=current_row, column=COL_SOLAR).value = solar
        ws.cell(row=current_row, column=COL_WIND).value = wind
        ws.cell(row=current_row, column=COL_HYBRID).value = hybrid
        ws.cell(row=current_row, column=COL_DATE_ADD).value = date_add
        
        # Set some defaults for consistency (Region = NR)
        ws.cell(row=current_row, column=3).value = "NR"
        
        print(f"Added row {current_row}: {app_id_full}")

    wb.save(excel_path)
    print("Excel update complete.")

if __name__ == "__main__":
    process_regulation_data()
