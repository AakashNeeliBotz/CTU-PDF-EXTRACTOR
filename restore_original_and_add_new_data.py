import pandas as pd
import pdfplumber
from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook, Workbook
import numpy as np


def extract_regulation_52_data():
    """Extract data from both Regulation 5.2 tables"""
    pdf_path = 'c:/Users/Sree Charan/Desktop/fold2/CTU-PDF-EXTRACTOR/175923696873MoM  40th CMETS NR meeting_F.pdf'
    
    all_data = []
    
    with pdfplumber.open(pdf_path) as pdf:
        # Extract from page 8 table 2 (June entries)
        page8 = pdf.pages[7]
        tables8 = page8.extract_tables()
        if len(tables8) > 1:
            june_table = tables8[1]
            
            # Skip header rows and extract data
            for row in june_table[3:]:  # Skip first 3 header rows
                if row and len(row) >= 15:  # Need at least 15 columns to access the data
                    # Parse the row data according to the PDF table structure
                    sl_no = row[0] if row[0] else ""
                    app_no_date = row[1] if row[1] else ""
                    applicant = row[2] if row[2] else ""
                    location = row[3] if row[3] else ""
                    nature = row[4] if row[4] else ""
                    conn_quantum = row[5] if row[5] else ""  # Column 5 for "App. No. & Conn. Quantum"
                    planned_capacity = row[8] if row[8] else ""  # Column 8 for "Planned additional capacity"
                    date_added = row[12] if row[12] else ""  # Column 12 for "Date from which additional capacity will be added"
                    
                    # Extract just the numeric application ID
                    app_id = ""
                    if app_no_date and '(' in str(app_no_date):
                        app_id = str(app_no_date).split('(')[0].strip()
                    
                    all_data.append({
                        'Sr.no.': sl_no,
                        'Application ID under Enhancement 5.2 or revision': app_id,
                        'Name of Developers': applicant,
                        'State': location,
                        'Nature of Applicant': nature,
                        'Application Quantum (MW)(ST II)': conn_quantum,
                        'Planned additional capacity (MW)': planned_capacity,
                        'Date from which additional capacity is to be added': date_added
                    })
        
        # Extract from page 9 table 3 (July entries)
        page9 = pdf.pages[8]
        tables9 = page9.extract_tables()
        if len(tables9) > 2:
            july_table = tables9[2]
            
            # Skip header rows and extract data
            for row in july_table[3:]:  # Skip first 3 header rows
                if row and len(row) >= 13:  # Need at least 13 columns
                    # Parse the row data according to the PDF table structure
                    sl_no = row[0] if row[0] else ""
                    app_no_date = row[1] if row[1] else ""
                    applicant = row[2] if row[2] else ""
                    location = row[3] if row[3] else ""
                    nature = row[4] if row[4] else ""
                    conn_quantum = row[6] if row[6] else ""  # Column 6 for "App. No. & Conn. Quantum"
                    planned_capacity = row[7] if row[7] else ""  # Column 7 for "Planned additional capacity"
                    date_added = row[10] if row[10] else ""  # Column 10 for "Date from which additional capacity will be added"
                    
                    # Extract just the numeric application ID
                    app_id = ""
                    if app_no_date and '(' in str(app_no_date):
                        app_id = str(app_no_date).split('(')[0].strip()
                    
                    all_data.append({
                        'Sr.no.': sl_no,
                        'Application ID under Enhancement 5.2 or revision': app_id,
                        'Name of Developers': applicant,
                        'State': location,
                        'Nature of Applicant': nature,
                        'Application Quantum (MW)(ST II)': conn_quantum,
                        'Planned additional capacity (MW)': planned_capacity,
                        'Date from which additional capacity is to be added': date_added
                    })
    
    return all_data


def parse_planned_capacity(planned_str):
    """Parse planned capacity string to extract solar and wind values"""
    if not planned_str or pd.isna(planned_str):
        return "", ""
    
    solar_val = ""
    wind_val = ""
    
    # Convert to string if it's not already
    planned_str = str(planned_str)
    
    # Check for combined values like "300 (BESS) 240 (Solar)"
    if "Solar" in planned_str and "Wind" in planned_str:
        # Handle cases like "200 (Wind)" or "240 (Solar)"
        import re
        solar_matches = re.findall(r'(\d+(?:\.\d+)?)\s*\(?Solar', planned_str)
        wind_matches = re.findall(r'(\d+(?:\.\d+)?)\s*\(?Wind', planned_str)
        
        solar_val = sum([float(x) for x in solar_matches]) if solar_matches else ""
        wind_val = sum([float(x) for x in wind_matches]) if wind_matches else ""
    elif "Solar" in planned_str:
        import re
        solar_matches = re.findall(r'(\d+(?:\.\d+)?)\s*\(?Solar', planned_str)
        solar_val = sum([float(x) for x in solar_matches]) if solar_matches else ""
    elif "Wind" in planned_str:
        import re
        wind_matches = re.findall(r'(\d+(?:\.\d+)?)\s*\(?Wind', planned_str)
        wind_val = sum([float(x) for x in wind_matches]) if wind_matches else ""
    elif "BESS" in planned_str:
        # BESS values typically go into solar
        import re
        bess_matches = re.findall(r'(\d+(?:\.\d+)?)\s*\(?BESS', planned_str)
        if bess_matches:
            solar_val = sum([float(x) for x in bess_matches])
    
    return solar_val, wind_val


def restore_original_and_add_new_data():
    """Restore original data and add new Regulation 5.2 data"""
    backup_path = 'c:/Users/Sree Charan/Desktop/fold2/CTU-PDF-EXTRACTOR/Connectivity_Application_Data_TEST_ALL_SHEETS38_backup.xlsx'
    target_path = 'c:/Users/Sree Charan/Desktop/fold2/CTU-PDF-EXTRACTOR/Connectivity_Application_Data_TEST_ALL_SHEETS38.xlsx'
    
    # Read the backup Excel file to get all sheets
    all_sheets = pd.read_excel(backup_path, sheet_name=None, header=1)  # Header is in row 1 (0-indexed)
    
    # Get the original 'Data to be captured' sheet
    original_df = all_sheets['Data to be captured']
    
    # Extract new data from PDF
    new_data = extract_regulation_52_data()
    
    if new_data:
        # Create DataFrame from new data
        df_new = pd.DataFrame(new_data)
        
        # Process the planned capacity to fill Solar and Wind columns
        solar_values = []
        wind_values = []
        for idx, row in df_new.iterrows():
            solar, wind = parse_planned_capacity(row['Planned additional capacity (MW)'])
            solar_values.append(solar)
            wind_values.append(wind)
        
        # Add Solar and Wind columns to new data
        df_new['Solar + BESS'] = solar_values
        df_new['Wind'] = wind_values
        
        # Ensure the new data has the same column structure as the original
        # Get the original column names from the original dataframe
        original_columns = list(original_df.columns)
        
        # Reorder and add missing columns to new data
        for col in original_columns:
            if col not in df_new.columns:
                df_new[col] = np.nan  # Add missing columns with NaN values
        
        df_new = df_new.reindex(columns=original_columns)
        
        # Combine original data with new data
        combined_df = pd.concat([original_df, df_new], ignore_index=True)
        
        # Update the 'Data to be captured' sheet with combined data
        all_sheets['Data to be captured'] = combined_df
        print(f"Combined original data ({len(original_df)} rows) with new Regulation 5.2 data ({len(df_new)} rows)")
        print(f"Total rows in updated sheet: {len(combined_df)}")
    else:
        print("No new data extracted from PDF")
    
    # Save the updated Excel file with all sheets
    with pd.ExcelWriter(target_path, engine='openpyxl') as writer:
        for sheet_name, df in all_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    print(f"Excel file updated with all sheets and combined data in 'Data to be captured' sheet.")
    
    # Apply formatting to all sheets
    apply_formatting(target_path)


def apply_formatting(file_path):
    """Apply bold formatting to headers in all sheets"""
    from openpyxl import load_workbook
    
    wb = load_workbook(file_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Make headers bold (headers are in row 2 based on our analysis)
        for cell in ws[2]:  # Second row contains headers
            cell.font = Font(bold=True)
        
        # Apply blue background to headers
        blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
        for cell in ws[2]:
            cell.fill = blue_fill
    
    wb.save(file_path)
    print("Formatting applied to all sheets.")


if __name__ == "__main__":
    restore_original_and_add_new_data()