"""
Remove duplicate rows Sr.No. 65-75 which are copies of Sr.No. 54-64
Also check Sr.No. 35-40 for similar issues
"""
import pandas as pd
from openpyxl import load_workbook
import sys

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

excel_path = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx"

# Load workbook
wb = load_workbook(excel_path)
ws = wb["Data to be captured"]

print("="*100)
print("REMOVING DUPLICATE ROWS Sr.No. 65-75")
print("="*100)

# Find the Excel rows for Sr.No. 65-75
# Check each row in the sheet
srno_col = 2  # Column B

rows_to_delete = []
for row in range(3, ws.max_row + 1):
    srno_val = ws.cell(row=row, column=srno_col).value
    if srno_val is not None:
        try:
            srno = float(srno_val)
            if 65 <= srno <= 75:
                # This row should be deleted - it's a duplicate
                app_id = ws.cell(row=row, column=9).value  # Column I = App ID
                substation = ws.cell(row=row, column=5).value
                developer = ws.cell(row=row, column=7).value
                
                # Only delete if it has NO Application ID (the duplicate ones don't have IDs)
                if app_id is None or str(app_id).strip() == '':
                    rows_to_delete.append({
                        'row': row,
                        'srno': srno,
                        'substation': substation,
                        'developer': str(developer)[:40] if developer else 'N/A',
                        'app_id': app_id
                    })
        except (ValueError, TypeError):
            pass

print(f"\nFound {len(rows_to_delete)} duplicate rows to delete:")
for item in rows_to_delete:
    print(f"  Excel Row {item['row']}: Sr.No. {int(item['srno'])}, {item['substation']}, {item['developer']}, AppID: {item['app_id']}")

# Delete rows from bottom to top to avoid shifting issues
if rows_to_delete:
    print(f"\nDeleting {len(rows_to_delete)} rows...")
    rows_to_delete_sorted = sorted(rows_to_delete, key=lambda x: x['row'], reverse=True)
    
    for item in rows_to_delete_sorted:
        ws.delete_rows(item['row'])
        print(f"  Deleted Excel row {item['row']} (Sr.No. {int(item['srno'])})")

# ============================================================================
# Now check Sr.No. 35-40 for similar issues
# ============================================================================
print("\n" + "="*100)
print("CHECKING Sr.No. 35-40 FOR DUPLICATES")
print("="*100)

rows_35_40 = []
for row in range(3, ws.max_row + 1):
    srno_val = ws.cell(row=row, column=srno_col).value
    if srno_val is not None:
        try:
            srno = float(srno_val)
            if 35 <= srno <= 40:
                app_id = ws.cell(row=row, column=9).value
                substation = ws.cell(row=row, column=5).value
                developer = ws.cell(row=row, column=7).value
                rows_35_40.append({
                    'row': row,
                    'srno': srno,
                    'substation': str(substation)[:25] if substation else 'N/A',
                    'developer': str(developer)[:30] if developer else 'N/A',
                    'app_id': app_id
                })
        except (ValueError, TypeError):
            pass

print(f"\nRows in Sr.No. 35-40: {len(rows_35_40)}")
for item in rows_35_40:
    print(f"  Row {item['row']}: Sr.No. {int(item['srno'])}, {item['substation']}, {item['developer']}, AppID: {item['app_id']}")

# Check for duplicates in 35-40 (same Sr.No. appearing multiple times)
srno_counts = {}
for item in rows_35_40:
    srno = int(item['srno'])
    if srno not in srno_counts:
        srno_counts[srno] = []
    srno_counts[srno].append(item)

duplicates_35_40 = []
for srno, items in srno_counts.items():
    if len(items) > 1:
        print(f"\n  DUPLICATE Sr.No. {srno} found ({len(items)} times):")
        for item in items:
            print(f"    Row {item['row']}: {item['substation']}, {item['developer']}, AppID: {item['app_id']}")
        # Keep the first one (with App ID if possible), delete others
        for item in items[1:]:
            duplicates_35_40.append(item)

if duplicates_35_40:
    print(f"\nDeleting {len(duplicates_35_40)} duplicate rows in 35-40 range...")
    duplicates_sorted = sorted(duplicates_35_40, key=lambda x: x['row'], reverse=True)
    for item in duplicates_sorted:
        ws.delete_rows(item['row'])
        print(f"  Deleted Excel row {item['row']} (Sr.No. {int(item['srno'])})")

# ============================================================================
# Save changes
# ============================================================================
print("\n" + "="*100)
print("SAVING CHANGES")
print("="*100)

wb.save(excel_path)
print(f"Changes saved to: {excel_path}")

# Verify
print("\n" + "="*100)
print("VERIFICATION")
print("="*100)

# Reload and count
df = pd.read_excel(excel_path, sheet_name="Data to be captured", header=1)
print(f"\nTotal rows after deletion: {len(df)}")

# Check rows with Sr.No. 65-75
srno_col_name = 'Sr.no.'
rows_65_75 = df[(df[srno_col_name] >= 65) & (df[srno_col_name] <= 75)]
print(f"Rows with Sr.No. 65-75: {len(rows_65_75)}")

# Check rows with Sr.No. 35-40
rows_35_40_check = df[(df[srno_col_name] >= 35) & (df[srno_col_name] <= 40)]
print(f"Rows with Sr.No. 35-40: {len(rows_35_40_check)}")
