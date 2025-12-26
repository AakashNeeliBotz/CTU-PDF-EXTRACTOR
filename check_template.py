import openpyxl

wb = openpyxl.load_workbook('Connectivity Application Data.xlsx')
ws = wb['Transformation Capacity']

print("="*80)
print("CHECKING TRANSFORMATION CAPACITY TEMPLATE STRUCTURE")
print("="*80)

print("\nRow 4 headers:")
for cell in ws[4]:
    if cell.value:
        print(f"  Column {cell.column}: {cell.value}")

print("\nRow 5 (if multi-row header):")
for cell in ws[5]:
    if cell.value:
        print(f"  Column {cell.column}: {cell.value}")
