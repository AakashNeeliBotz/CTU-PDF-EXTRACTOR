"""
Remove Duplicate GNA/ST II Application IDs
Strategy:
1. Group rows by ID.
2. For meaningful duplicates, identify the 'best' row to keep the ID.
   Criteria:
   - Priority 1: Has Element Codes (Cols AN, AO, AP / 40, 41, 42)
   - Priority 2: Has Developer Name (Col G / 7)
   - Priority 3: First occurrence (Row Index)
3. Clear the ID from all other rows in the group (making them blank in that column).
4. Do NOT delete any rows.
"""

import openpyxl
from collections import defaultdict

FILE_PATH = r'c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx'

def has_element_custom_data(ws, row_idx):
    # Check columns 40, 41, 42 (AN, AO, AP)
    for col in [40, 41, 42]:
        val = ws.cell(row=row_idx, column=col).value
        if val and str(val).strip():
            return True
    return False

def has_developer_data(ws, row_idx):
    val = ws.cell(row=row_idx, column=7).value
    return val and str(val).strip()

def process_excel():
    print("Loading workbook...")
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb['Data to be captured']
    
    # Group rows by ID
    id_groups = defaultdict(list)
    cols_to_map = {} # row -> ID (to keep track)
    
    print("Grouping rows by ID...")
    for row_idx in range(3, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=9).value # Column I
        if val:
            val_str = str(val).strip()
            id_groups[val_str].append(row_idx)
            cols_to_map[row_idx] = val_str
            
    removed_count = 0
    groups_processed = 0
    
    print(f"Found {len(id_groups)} unique IDs.")
    print("Processing duplicates...")
    
    for app_id, rows in id_groups.items():
        if len(rows) > 1:
            groups_processed += 1
            
            # Score rows to find the best one
            scored_rows = []
            for r in rows:
                score = 0
                if has_element_custom_data(ws, r):
                    score += 1000
                if has_developer_data(ws, r):
                    score += 100
                # Subtract small amount for row index to prefer top rows in ties
                score -= (r * 0.001)
                scored_rows.append((score, r))
            
            # Sort by score descending
            scored_rows.sort(key=lambda x: x[0], reverse=True)
            
            best_row = scored_rows[0][1]
            
            # Clear ID for all other rows
            for _, r in scored_rows[1:]:
                ws.cell(row=r, column=9).value = None # Clear GNA ID
                removed_count += 1
                
    print(f"\n=== SUMMARY ===")
    print(f"Processed {groups_processed} groups with duplicates.")
    print(f"Removed {removed_count} repeated ID values.")
    
    print(f"\nSaving workbook...")
    wb.save(FILE_PATH)
    print("Done!")

if __name__ == "__main__":
    process_excel()
