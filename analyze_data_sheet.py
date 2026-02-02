"""
Script to analyze the 'Data to be captured' sheet in the Excel file
"""
import pandas as pd
from openpyxl import load_workbook

# Load the Excel file
excel_path = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx"

# First, let's see all sheet names
wb = load_workbook(excel_path)
print("Sheet names:", wb.sheetnames)

# Load the specific sheet
sheet_name = "Data to be captured"
df = pd.read_excel(excel_path, sheet_name=sheet_name, header=0)

print("\n" + "="*80)
print("Column names and their indices:")
print("="*80)
for i, col in enumerate(df.columns):
    print(f"Column {i} (Excel Col {chr(65+i) if i < 26 else 'A'+chr(65+i-26)}): {col}")

print("\n" + "="*80)
print(f"Total rows: {len(df)}")
print("="*80)

# Check for Sr.No. column
srno_col = None
for col in df.columns:
    if 'sr' in str(col).lower() and 'no' in str(col).lower():
        srno_col = col
        break

if srno_col:
    print(f"\nSr.No. column found: '{srno_col}'")
    
    # Check rows 65-75 for duplicates
    print("\n" + "="*80)
    print("Checking rows 65-75 (by Sr.No.):")
    print("="*80)
    mask_65_75 = (df[srno_col] >= 65) & (df[srno_col] <= 75)
    rows_65_75 = df[mask_65_75]
    print(rows_65_75[[srno_col]].to_string())
    print(f"\nDuplicate Sr.No. values in 65-75: {rows_65_75[srno_col].duplicated().sum()}")
    
    # Check rows 35-40 for duplicates
    print("\n" + "="*80)
    print("Checking rows 35-40 (by Sr.No.):")
    print("="*80)
    mask_35_40 = (df[srno_col] >= 35) & (df[srno_col] <= 40)
    rows_35_40 = df[mask_35_40]
    print(rows_35_40[[srno_col]].to_string())
    print(f"\nDuplicate Sr.No. values in 35-40: {rows_35_40[srno_col].duplicated().sum()}")

# Check Substation column (Column E)
substation_col = None
for col in df.columns:
    if 'substation' in str(col).lower():
        substation_col = col
        break

if substation_col:
    print("\n" + "="*80)
    print(f"Substation column found: '{substation_col}'")
    print("Checking for ID values (301, 302, 303, etc.):")
    print("="*80)
    for idx, val in df[substation_col].items():
        if pd.notna(val) and str(val).strip().isdigit():
            numeric_val = int(str(val).strip())
            if 100 <= numeric_val <= 999:  # Check for 3-digit IDs
                print(f"Row {idx+2} (Excel row): Substation = '{val}'")

# Check Column I for GNA/ST II Application ID
print("\n" + "="*80)
print("Column I (index 8) - GNA/ST II Application ID:")
print("="*80)
if len(df.columns) > 8:
    col_i = df.columns[8]
    print(f"Column I name: '{col_i}'")

# Check Sr.No. 15 data
print("\n" + "="*80)
print("Sr.No. 15 data:")
print("="*80)
if srno_col:
    row_15 = df[df[srno_col] == 15]
    if not row_15.empty:
        for col in row_15.columns:
            val = row_15[col].values[0]
            if pd.notna(val):
                print(f"{col}: {val}")

# Check for "Bhuj" in Substation
print("\n" + "="*80)
print("Rows with 'Bhuj' in Substation:")
print("="*80)
if substation_col:
    bhuj_rows = df[df[substation_col].astype(str).str.contains('Bhuj', case=False, na=False)]
    for idx, row in bhuj_rows.iterrows():
        print(f"\nRow {idx+2} (Excel):")
        for col in ['Sr.no.', 'Sr. no', 'Sr.No', srno_col, substation_col, 'Latitude', 'Longitude', 'latitude', 'longitude']:
            if col in df.columns:
                print(f"  {col}: {row[col]}")

# Check Sr.No. 58
print("\n" + "="*80)
print("Sr.No. 58 data:")
print("="*80)
if srno_col:
    row_58 = df[df[srno_col] == 58]
    if not row_58.empty:
        for col in row_58.columns:
            val = row_58[col].values[0]
            if pd.notna(val):
                print(f"{col}: {val}")

# Check CMETS GNA Approved column
print("\n" + "="*80)
print("Checking 'CMETS GNA Approved' column for rows with both 33 and 34:")
print("="*80)
cmets_col = None
for col in df.columns:
    if 'cmets' in str(col).lower() and 'gna' in str(col).lower() and 'approved' in str(col).lower():
        cmets_col = col
        break

if cmets_col:
    print(f"Found column: '{cmets_col}'")
    for idx, val in df[cmets_col].items():
        if pd.notna(val):
            val_str = str(val)
            if '33' in val_str and '34' in val_str:
                print(f"Row {idx+2} (Excel): {val}")

# Check rows 194 and 195
print("\n" + "="*80)
print("Checking rows 194 and 195 (0-indexed: 193 and 194):")
print("="*80)
if len(df) >= 195:
    for row_idx in [193, 194]:  # 0-indexed
        if row_idx < len(df):
            print(f"\nRow {row_idx+2} (Excel row):")
            for col in df.columns:
                val = df.iloc[row_idx][col]
                if pd.notna(val):
                    print(f"  {col}: {val}")
else:
    print(f"File only has {len(df)} data rows")
