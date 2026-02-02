"""
Fix Sr.No. 301, 302, 303 - Split Substation and AppID values
"""
import pandas as pd
from openpyxl import load_workbook
import sys
import re

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

excel_path = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx"

# Load workbook
wb = load_workbook(excel_path)
ws = wb["Data to be captured"]

print("="*100)
print("FIXING Sr.No. 301, 302, 303 - SEPARATING SUBSTATION AND APPLICATION ID")
print("="*100)

srno_col = 2  # Column B
substation_col = 5  # Column E
app_id_col = 9  # Column I

changes_made = []

for row in range(3, ws.max_row + 1):
    srno_val = ws.cell(row=row, column=srno_col).value
    if srno_val is not None:
        try:
            srno = float(srno_val)
            if srno in [301, 302, 303]:
                substation = ws.cell(row=row, column=substation_col).value
                existing_app_id = ws.cell(row=row, column=app_id_col).value
                
                if substation:
                    sub_str = str(substation).strip()
                    
                    # Pattern: "Bikaner-II PS   212100007" or similar
                    # Split on multiple spaces or extract number at end
                    match = re.match(r'^(.+?)\s{2,}(\d{9,10})$', sub_str)
                    
                    if match:
                        actual_substation = match.group(1).strip()
                        app_id_from_substation = match.group(2).strip()
                        
                        print(f"\nExcel Row {row} - Sr.No. {int(srno)}:")
                        print(f"  Original Substation: '{sub_str}'")
                        print(f"  Extracted Substation: '{actual_substation}'")
                        print(f"  Extracted App ID: '{app_id_from_substation}'")
                        print(f"  Existing App ID (Col I): '{existing_app_id}'")
                        
                        # Update Substation column
                        ws.cell(row=row, column=substation_col).value = actual_substation
                        
                        # Update App ID column (only if empty or contains "(Enh.)")
                        if existing_app_id is None or str(existing_app_id).strip() == '' or '(Enh' in str(existing_app_id):
                            if existing_app_id and '(Enh' in str(existing_app_id):
                                ws.cell(row=row, column=app_id_col).value = f"{app_id_from_substation} (Enh.)"
                            else:
                                ws.cell(row=row, column=app_id_col).value = app_id_from_substation
                        
                        changes_made.append(f"Row {row} (Sr.No. {int(srno)}): Substation='{actual_substation}', AppID='{app_id_from_substation}'")
                        print(f"  ✓ Fixed!")
                    else:
                        print(f"\nExcel Row {row} - Sr.No. {int(srno)}:")
                        print(f"  Substation: '{sub_str}' - Could not parse")
        except (ValueError, TypeError):
            pass

print("\n" + "="*100)
print("SAVING CHANGES")
print("="*100)

wb.save(excel_path)
print(f"Changes saved to: {excel_path}")

print("\n" + "="*100)
print("VERIFICATION")
print("="*100)

# Reload and verify
wb = load_workbook(excel_path)
ws = wb["Data to be captured"]

for row in range(3, ws.max_row + 1):
    srno_val = ws.cell(row=row, column=srno_col).value
    if srno_val is not None:
        try:
            srno = float(srno_val)
            if srno in [301, 302, 303]:
                substation = ws.cell(row=row, column=substation_col).value
                app_id = ws.cell(row=row, column=app_id_col).value
                print(f"\nRow {row} - Sr.No. {int(srno)}:")
                print(f"  Substation: {substation}")
                print(f"  App ID: {app_id}")
        except (ValueError, TypeError):
            pass

print("\n" + "="*100)
print("SUMMARY")
print("="*100)
print(f"\nTotal changes: {len(changes_made)}")
for change in changes_made:
    print(f"  ✓ {change}")
