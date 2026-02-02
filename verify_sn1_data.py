import pandas as pd
from openpyxl import load_workbook

def verify_excel_data():
    """Verify the current state of the Excel file"""
    
    excel_path = r'c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx'
    
    print("="*100)
    print("EXCEL FILE VERIFICATION")
    print("="*100)
    print(f"\nFile: {excel_path}\n")
    
    # Load with openpyxl to preserve formatting
    wb = load_workbook(excel_path)
    ws = wb['Data to be captured']
    
    # Print column headers from row 2
    print("Column Headers (Row 2):")
    print("-"*100)
    headers = []
    for col in range(1, 20):  # First 20 columns
        cell_value = ws.cell(row=2, column=col).value
        headers.append(str(cell_value) if cell_value else "")
        if col in [9, 10, 12, 13, 14, 15]:  # Highlight our target columns
            print(f"  Column {col:2d}: *** {cell_value} ***")
        else:
            print(f"  Column {col:2d}: {cell_value}")
    
    print("\n" + "="*100)
    print("SAMPLE DATA ROWS")
    print("="*100)
    print("\nShowing first 10 data rows with GNA/LTA Application IDs:\n")
    
    # Find rows with GNA or LTA Application IDs
    rows_shown = 0
    for row in range(3, ws.max_row + 1):
        if rows_shown >= 10:
            break
            
        gna_id = ws.cell(row=row, column=9).value  # GNA/ST II Application ID
        lta_id = ws.cell(row=row, column=10).value  # LTA Application ID
        cmets_gna = ws.cell(row=row, column=12).value  # CMETS GNA Approved
        cmets_lta = ws.cell(row=row, column=13).value  # CMETS LTA Approved
        gna_date = ws.cell(row=row, column=14).value  # CMETS GNA Meeting Date
        lta_date = ws.cell(row=row, column=15).value  # CMETS LTA Meeting Date
        
        # Only show rows with at least one Application ID
        if gna_id or lta_id:
            print(f"Row {row}:")
            print(f"  GNA/ST II Application ID: {gna_id}")
            print(f"  LTA Application ID:       {lta_id}")
            print(f"  CMETS GNA Approved:       {cmets_gna}")
            print(f"  CMETS LTA Approved:       {cmets_lta}")
            print(f"  CMETS GNA Meeting Date:   {gna_date}")
            print(f"  CMETS LTA Meeting Date:   {lta_date}")
            print()
            rows_shown += 1
    
    # Statistics
    print("="*100)
    print("STATISTICS")
    print("="*100)
    
    total_rows = ws.max_row - 2  # Subtract header rows
    gna_ids_count = 0
    lta_ids_count = 0
    gna_approved_count = 0
    lta_approved_count = 0
    gna_date_count = 0
    lta_date_count = 0
    
    for row in range(3, ws.max_row + 1):
        gna_id = ws.cell(row=row, column=9).value
        lta_id = ws.cell(row=row, column=10).value
        cmets_gna = ws.cell(row=row, column=12).value
        cmets_lta = ws.cell(row=row, column=13).value
        gna_date = ws.cell(row=row, column=14).value
        lta_date = ws.cell(row=row, column=15).value
        
        if gna_id and str(gna_id).strip() and str(gna_id) != 'nan':
            gna_ids_count += 1
        if lta_id and str(lta_id).strip() and str(lta_id) != 'nan':
            lta_ids_count += 1
        if cmets_gna and str(cmets_gna).strip() and str(cmets_gna) != 'nan':
            gna_approved_count += 1
        if cmets_lta and str(cmets_lta).strip() and str(cmets_lta) != 'nan':
            lta_approved_count += 1
        if gna_date and str(gna_date).strip() and str(gna_date) != 'nan':
            gna_date_count += 1
        if lta_date and str(lta_date).strip() and str(lta_date) != 'nan':
            lta_date_count += 1
    
    print(f"\nTotal data rows: {total_rows}")
    print(f"\nApplication IDs:")
    print(f"  Rows with GNA/ST II Application ID: {gna_ids_count}")
    print(f"  Rows with LTA Application ID:       {lta_ids_count}")
    print(f"\nCMETS Approval Status:")
    print(f"  Rows with CMETS GNA Approved:       {gna_approved_count}")
    print(f"  Rows with CMETS LTA Approved:       {lta_approved_count}")
    print(f"\nMeeting Dates:")
    print(f"  Rows with CMETS GNA Meeting Date:   {gna_date_count}")
    print(f"  Rows with CMETS LTA Meeting Date:   {lta_date_count}")
    
    # Check for empty cells that should be filled
    print(f"\nData Completeness:")
    gna_missing = gna_ids_count - gna_approved_count
    lta_missing = lta_ids_count - lta_approved_count
    gna_date_missing = gna_approved_count - gna_date_count
    lta_date_missing = lta_approved_count - lta_date_count
    
    print(f"  GNA IDs without CMETS Approval:     {gna_missing}")
    print(f"  LTA IDs without CMETS Approval:     {lta_missing}")
    print(f"  GNA Approvals without Meeting Date: {gna_date_missing}")
    print(f"  LTA Approvals without Meeting Date: {lta_date_missing}")
    
    if gna_missing == 0 and lta_missing == 0 and gna_date_missing == 0 and lta_date_missing == 0:
        print("\n[OK] All data is complete!")
    else:
        print("\n[WARNING] Some data is missing and could be filled by the script.")
    
    print("\n" + "="*100)

if __name__ == "__main__":
    verify_excel_data()
