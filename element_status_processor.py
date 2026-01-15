
import pdfplumber
import openpyxl
import pandas as pd
import re
import os
import warnings
from openpyxl.utils import get_column_letter

# Suppress pandas warnings
warnings.filterwarnings("ignore")

class ElementStatusProcessor:
    """
    Processor for 'Element Status' sheet data extraction.
    Integrates functionality from ElementStatusv2/modules/extractor.py and populator.py.
    """
    
    def __init__(self, target_text="Monitoring Report of Under Construction TBCB Projects"):
        self.target_text = target_text
        
        # --- Ported from ExcelPopulator.__init__ ---
        self.src_cols_map = {
            'Scope': ['Name', 'Scope'],
            'SPV': ['SPV', 'Transfe'],
            'Locs': ['Total', 'Locs'],
            'Found': ['Found', 'ation', 'Nos'],
            'Erect': ['Erecti', 'on', 'Nos'],
            'String': ['Stringin', 'g'],
            'Civil': ['Civil', 'works'],
            'EqptRec': ['Eqpt', 'Receive'],
            'EqptEre': ['Eqpt', 'Erectio'],
            'OrgSCOD': ['Target', 'Org'],
            'AntSCOD': ['Target', 'Anticipate'],
            'Remarks': ['Remarks'],
            'Length': ['Lengt', 'h'],
            'AwardedTo': ['Exec', 'Agenc']
        }
        
        # Absolute Column Indices (1-based) for the TARGET Excel
        self.mapping_rules = {
            3: 'InterIntra', 4: 'Scheme',  # New derived columns
            15: 'AwardedTo',  # PDF 'Exec. Agency' -> Excel 'Awarded To'
            16: 'SPV', 17: 'Length', 18: 'Locs', 19: 'Found', 20: 'Erect',
            21: 'String', 22: 'CALC_FOUND', 23: 'CALC_ERECT', 24: 'CALC_STRING',
            25: 'Civil', 26: 'EqptRec', 27: 'EqptEre', 28: 'OrgSCOD',
            29: 'AntSCOD', 30: 'Remarks'
        }

    def extract_scheme_details(self, text):
        """
        Extract Region, Phase, and Part from Project Name.
        Logic:
          - Attempt specific regex extraction first.
          - Fallback to generic cleaning if regex fails to find a Region.
        """
        if not text: return None, None
        
    def extract_scheme_details(self, text):
        """
        Extract Region, Phase, and Part from Project Name.
        Logic:
          - Attempt specific regex extraction for Region/Phase/Part.
          - Construct Scheme from valid components if found.
          - If no Region is found, return empty strings (Strict Mode).
        """
        if not text: return None, None
        
        original_text = text
        text = text.replace('\n', ' ').strip()
        
        # 0. Clean SPV/Suffixes PRE-REGEX (Aggressive)
        text = re.sub(r'\s*\(?SPV\s*[:].*?(?:\)|$)', '', text, flags=re.IGNORECASE)
        text = re.sub(r'\s*,\s*SPV.*', '', text, flags=re.IGNORECASE)
        text = re.sub(r'\s*\(SPV.*?\)', '', text, flags=re.IGNORECASE)
        text = re.sub(r'\s*\([\d\.]+\s*GW\)', '', text, flags=re.IGNORECASE)
        text = re.sub(r'\s*\(CKM.*?\)', '', text, flags=re.IGNORECASE)
        # Remove (Part-X: ...) which usually denotes capacity, keeping (Part-X) if no colon
        text = re.sub(r'\(Part\s*[-–]?\s*\d+\s*:.*?\)', '', text, flags=re.IGNORECASE)

        # 1. Region Regex
        # PRIORITIZE "Rajasthan REZ" / "Gujarat REZ" over subtations like Bhadla/Sikar
        # Sort pattern by length descending to catch specific names first
        regions_list = [
            'Rajasthan REZ', 'Gujarat REZ', 'Khavda RE Park', 'Khavda',
            'Rajasthan', 'Gujarat', 'Madhya Pradesh', 'MP', 'M.P.', 
            'Karnataka', 'Tamil Nadu', 'TN', 'Andhra', 'Telangana',
            'Ananthapuram', 'Kurnool', 'Bhadla', 'Sikar', 'Fatehgarh', 
            'Kadeoni', 'Koppal', 'Gadag', 'Bidar', 'Rajnandgaon', 'Kallam', 
            'KPS[0-9]*', 'KPS', 'Bikaner Complex', 'Bikaner'
        ]
        # Sort by length descending to match "Rajasthan REZ" before "Rajasthan"
        regions_list.sort(key=len, reverse=True)
        region_pattern = r'(' + '|'.join([re.escape(r) for r in regions_list]) + r')(?:[- ]?[IVX]+)?'
        
        # Find ALL matches, prioritize "Rajasthan/Gujarat" if "as part of" context?
        # Actually, simpler: Search for "as part of X" first?
        # Case: "from Bhadla-III PS as part of Rajasthan REZ"
        part_of_match = re.search(r'as part of\s+(.*?Scheme|.*?REZ(?: Phase[- ]?[IVX]*)?)', text, re.IGNORECASE)
        if part_of_match:
             potential_scheme = part_of_match.group(1)
             # Recursively call extract on this snippet to get region? 
             # Or just use this text for region search
             region_match = re.search(region_pattern, potential_scheme, re.IGNORECASE)
        else:
             region_match = re.search(region_pattern, text, re.IGNORECASE)
             
        region = region_match.group(1) if region_match else None
        
        # Normalize Region
        if region:
            if "M.P." in region or "MP" == region: region = "Madhya Pradesh"
            if "KPS" in region.upper(): region = region.upper()
            region = re.sub(r'REZ', 'REZ', region, flags=re.IGNORECASE)

        if not region:
             return "", ""

        # 2. Extract Phase
        # Handles: Ph-IV, Phase-II. 
        # Exclude "REZPhase" fixes (done via pre-clean?)
        text = re.sub(r'REZPhase', 'Phase', text, flags=re.IGNORECASE)
        phase_match = re.search(r'((?:Phase|Ph)\s*[-–]?\s*[IVX0-9]+)', text, re.IGNORECASE)
        phase = phase_match.group(1) if phase_match else ""
        
        # Normalize Phase
        phase = re.sub(r'Ph\s*[-–]', 'Phase-', phase, flags=re.IGNORECASE) # Ph-IV -> Phase-IV
        phase = re.sub(r'\s*[-–]\s*', '-', phase) 
        phase = re.sub(r'\s+', ' ', phase)
        
        # 3. Extract Part
        # Capture Parts outside parens first
        # Regex: Part A, Part B1 & B2
        # EXCLUDE "Part of" (e.g. "as part of Rajasthan...")
        
        # Find "Part X" where X is alphanumeric. 
        # Added negative lookahead (?!\s+of\b) to exclude "Part of"
        # Adjusted to allow "Part-A" (no space)
        parts_candidates = re.finditer(r'(Part(?!\s+of\b)\s*[-–]?\s*[A-Z0-9]+(?:[\s,&]+[A-Z0-9]+)*)', text, re.IGNORECASE)
        
        clean_parts = []
        for p in parts_candidates:
            val = p.group(1)
            # Normalize
            val = re.sub(r'\s*[-–]\s*', ' ', val)
            clean_parts.append(val)
            
        # Deduplication strategy:
        # If text has "(Part-1)... Part A", clean_parts might have ["Part 1", "Part A"].
        # But we want output: "Region Phase (Part-1) Part A".
        # So we should capture "(Part-1)" explicitly as a token.
        
        parts_paren = re.findall(r'(\(Part\s*[-–]?\s*\d+\))', text, re.IGNORECASE)
        
        # Remove matched paren parts from clean_parts if they overlap?
        # e.g. regex for clean_parts catches "Part 1" inside "(Part 1)".
        # We filter out parts that are structurally identical to what we found in parens.
        
        final_parts = []
        # Add paren parts first
        for pp in parts_paren:
            # Normalize pp for output? User kept (Part-1).
            final_parts.append(pp)
            
        # Add non-paren parts if not similar
        for cp in clean_parts:
            # Check if 'cp' (e.g. Part 1) is contained in any paren part (Part-1)
            is_duplicate = False
            cp_norm = re.sub(r'\s', '', cp).lower()
            for pp in parts_paren:
                pp_norm = re.sub(r'[\(\)\-\s]', '', pp).lower()
                if cp_norm in pp_norm: 
                    is_duplicate = True
            
            if not is_duplicate:
                final_parts.append(cp)
        
        part_str = " ".join(final_parts).strip()
        
        # 4. Construct Output
        scheme_parts = [region]
        if phase: scheme_parts.append(phase)
        if part_str: scheme_parts.append(part_str)
        
        scheme = " ".join(scheme_parts).strip()
        
        # Inter/Intra
        inter_intra = f"{region} {phase}".strip()
        
        # Final Whitespace Clean
        scheme = re.sub(r'\s+', ' ', scheme)
        inter_intra = re.sub(r'\s+', ' ', inter_intra)
        
        return inter_intra, scheme

    # --- Methods from PDFExtractor ---
    def is_number(self, s):
        if not isinstance(s, str): return False, None
        try:
            if s.isdigit(): return True, int(s)
            if re.match(r'^-?\d+(\.\d+)?$', s): return True, float(s)
            return False, None
        except: return False, None

    def extract_from_pdf(self, pdf_path):
        """Extract tables from PDF (Ported from PDFExtractor.extract)"""
        print(f"  [Element Status] Opening PDF: {pdf_path}")
        all_tables = []
        started = False
        
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                if not started:
                    # Check first portion of text to see if we reached the section
                    text = page.extract_text()
                    if text and self.target_text.lower() in text.lower():
                        print(f"  [Element Status] Found start on Page {page_num + 1}")
                        started = True
                
                if started:
                    # Extract tables from this page
                    tables = page.extract_tables()
                    for table in tables:
                        # Clean cells
                        cleaned = [[(c.strip() if c else None) for c in row] for row in table]
                        all_tables.append(cleaned)
        return all_tables

    def merge_tables(self, tables):
        """Merge multiple extracted tables (Ported from PDFExtractor.merge_tables)"""
        if not tables: return []
        def is_header(row):
            # Heuristic: headers usually have "SN" in the first column
            if not row or not row[0]: return False
            val = str(row[0]).lower()
            return "sn" in val or "s.n" in val or "sl.no" in val

        # Find the first table that looks like a data table
        start_idx = 0
        start_row_idx = 0
        found_start = False
        
        for i, table in enumerate(tables):
            if not table: continue
            # Check first few rows of the table for header-like content
            for r_idx, row in enumerate(table[:5]): # Check first 5 rows
                # Check for key columns
                row_str = " ".join([str(c).lower() for c in row if c])
                
                has_sn = "sn" in row_str.split() or "s.n" in row_str or "sl.no" in row_str or "sl. no" in row_str
                has_scope = "scope" in row_str or "name" in row_str or "project" in row_str or "element" in row_str
                
                if has_sn and has_scope:
                    start_idx = i
                    start_row_idx = r_idx
                    found_start = True
                    break
            if found_start: break
            
        if not found_start:
            print("  [Element Status] Warning: Could not identify start table with standard headers. Using first table.")
            start_idx = 0
            start_row_idx = 0

        merged = [row[:] for row in tables[start_idx][start_row_idx:]]
        # If the matched row was not the first row, we should probably strip rows before it?
        # But usually headers are at top of table.
        
        # Merge subsequent
        for next_table in tables[start_idx+1:]:
            if not next_table: continue
            if is_header(next_table[0]):
                # If next table has header, skip it (append rows from index 1)
                merged.extend(next_table[1:])
            else:
                # No header, append all rows
                merged.extend(next_table)
        return merged

    def convert_to_dataframe(self, raw_data):
        """Convert list of lists to Pandas DataFrame for easier processing"""
        if not raw_data:
            return pd.DataFrame()
            
        # Assume first row is header if valid
        headers = raw_data[0]
        # Ensure unique headers
        data = raw_data[1:]
        
        # Basic cleanup of headers
        clean_headers = [str(h).strip() if h else f"Col_{i}" for i, h in enumerate(headers)]
        
        df = pd.DataFrame(data, columns=clean_headers)
        return df

    # --- Methods from ExcelPopulator ---
    def clean_text(self, val):
        if pd.isna(val) or val is None: return None
        # Normalize: strip, replace newlines/tabs with space, collapse multiple spaces
        text = str(val).strip().replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
        # Standardize dashes and common special characters
        text = text.replace('û', '-').replace('–', '-').replace('—', '-')
        text = re.sub(r'\s+', ' ', text).strip()
        # Lowercase for consistent semantic matching
        return text.lower()

    def to_float(self, val):
        try: return float(val) if not pd.isna(val) else 0.0
        except: return 0.0

    def find_col(self, df, keywords):
        for col in df.columns:
            if all(k.lower() in str(col).lower() for k in keywords):
                return col
        return None

    def process_and_write(self, pdf_path, target_excel_path):
        """
        Main orchestration method:
         1. Extract tables from PDF
         2. Convert to intermediate DataFrame (Logic from Populator.sync loaded part)
         3. Write to Target Excel (Logic from Populator.sync write part)
        """
        # 1. Extract Data
        raw_tables = self.extract_from_pdf(pdf_path)
        merged_data = self.merge_tables(raw_tables)
        
        if not merged_data:
            print("  [Element Status] No data extracted from PDF.")
            return

        # 2. Prepare Source Data (in-memory)
        # We need to map the raw extracted columns to our internal 'src_data' dict
        # First, find the column names from the headers (first row of merged_data)
        
        if not merged_data: return
        
        # Create a temp DF for column mapping logic
        headers = merged_data[0]
        # print(f"  [DEBUG] Extracted Headers: {headers}")
        df_src = pd.DataFrame(merged_data[1:], columns=headers)
        
        # Identical logic to Populator.sync for finding source columns
        src_cols = {}
        for k, kw in self.src_cols_map.items():
            found = self.find_col(df_src, kw)
            if found: src_cols[k] = found
            else: print(f"  [Element Status] Warning: Missing source col for {k} ({kw})")

        src_data = {}
        
        # Parent Context State Machine
        current_context = {
            'Scheme': None,
            'InterIntra': None,
            'SPV': None
        }
        
        # Source Column Identifiers
        col_sn_name = self.find_col(df_src, ['SN']) 
        if not col_sn_name: col_sn_name = self.find_col(df_src, ['S.N'])
        if not col_sn_name: col_sn_name = self.find_col(df_src, ['Sl', 'No'])

        col_scope_name = src_cols.get('Scope')
        col_spv_name = src_cols.get('SPV')

        for idx, r in df_src.iterrows():
            if not col_scope_name: continue
            
            # Check if this is a Parent/Context Row (Has SN)
            is_parent = False
            sn_val = r[col_sn_name] if col_sn_name else None
            # Check is non-empty SN
            if sn_val and str(sn_val).strip() and str(sn_val).strip().lower() != 'nan':
                 is_parent = True
            
            raw_scope_text = r[col_scope_name]
            clean_scope_text = self.clean_text(raw_scope_text)
            
            if is_parent:
                # Update Context
                # 1. Parse Key Scheme Info
                inter, scheme = self.extract_scheme_details(str(raw_scope_text))
                current_context['InterIntra'] = inter
                current_context['Scheme'] = scheme
                
                # 2. Capture SPV if present in this parent row
                if col_spv_name:
                    spv_val = r[col_spv_name]
                    if spv_val and str(spv_val).strip():
                        current_context['SPV'] = spv_val
                
                # Do NOT add parent row to src_data
                continue
            
            # This is a Child/Data Row
            if not clean_scope_text or len(clean_scope_text) < 3: continue
            
            # Enrich row with Context
            r_enriched = r.copy()
            r_enriched['InterIntra'] = current_context['InterIntra']
            r_enriched['Scheme'] = current_context['Scheme']
            # Only overwrite SPV if child doesn't have it (though usually child has empty SPV)
            if col_spv_name:
                child_spv = r[col_spv_name]
                if not child_spv or pd.isna(child_spv):
                    r_enriched[col_spv_name] = current_context['SPV']
            
            # Store by Scope Key
            src_data[clean_scope_text] = r_enriched
            
        print(f"  [Element Status] Loaded {len(src_data)} records from extracted tables (using hierarchical logic).")

        # 3. Write to Excel
        if not os.path.exists(target_excel_path):
            print(f"  [Element Status] Error: Target Excel {target_excel_path} does not exist.")
            return

        wb = openpyxl.load_workbook(target_excel_path)
        
        # Check if sheet exists, if not create it (though template should have it)
        if "Element Status" not in wb.sheetnames:
            print("  [Element Status] Creating new 'Element Status' sheet.")
            ws = wb.create_sheet("Element Status")
            # You might need to setup headers here if creating from scratch
        else:
            ws = wb["Element Status"]
            
        # Identify Scope Column in Target (Search Row 2)
        # Ported logic:
        scope_col = None
        for cell in ws[2]:
            if cell.value and "Transmission Scope" in str(cell.value):
                scope_col = cell.column
                break
        
        if not scope_col:
            print("  [Element Status] Error: 'Transmission Scope' column not found in row 2 of target sheet.")
            # Fallback or return? Let's return for safety.
            return

        updates, skips = 0, 0
        target_rows = list(ws.iter_rows(min_row=4))
        
        src_keys = list(src_data.keys())
        matched_src_indices = set()
        
        print(f"  [Element Status] Processing {len(target_rows)} rows in target Excel...")

        for r_idx, row_cells in enumerate(target_rows):
            # Safety check for column index
            if scope_col - 1 >= len(row_cells): continue
            
            raw_target_val = row_cells[scope_col-1].value
            target_key = self.clean_text(raw_target_val)
            
            src_row = None
            
            # Logic: If target is empty/too short, try sequential fill
            if not target_key or len(target_key) < 5:
                # Sequential Fallback
                if r_idx < len(src_keys):
                    sk = src_keys[r_idx]
                    src_row = src_data[sk]
                    matched_src_indices.add(r_idx)
            else:
                # Match Logic: Strict -> Fuzzy
                if target_key in src_data:
                    src_row = src_data[target_key]
                    try: matched_src_indices.add(src_keys.index(target_key))
                    except: pass
                else:
                    # Fuzzy match
                    for idx, sk in enumerate(src_keys):
                        if target_key in sk or sk in target_key:
                            src_row = src_data[sk]
                            matched_src_indices.add(idx)
                            break
            
            if src_row is None:
                continue

            # CLEAR specific columns before writing
            source_row_idx = row_cells[0].row
            
            # Clear logic from original code: Cols 3, 4, 5 (C, D, E) and mapped cols
            # Note: openpyxl is 1-based. 3=C, 4=D, 5=E.
            for col_idx in [3, 4, 5]:
                ws.cell(row=source_row_idx, column=col_idx).value = None
            for col_idx in self.mapping_rules.keys():
                ws.cell(row=source_row_idx, column=col_idx).value = None

            # Write Transmission Scope into Col 5 (E)
            src_scope_val = src_row[src_cols['Scope']]
            ws.cell(row=source_row_idx, column=5, value=src_scope_val)

            # Populate Other Mapping Rules
            for col_idx, rule in self.mapping_rules.items():
                cell = ws.cell(row=source_row_idx, column=col_idx)
                val = None
                
                if rule.startswith('CALC_'):
                    try:
                        k = rule.split('_')[1]
                        # Logic: Found/Locs, Erect/Locs, String/Length
                        num_col = src_cols['Found'] if k=='FOUND' else (src_cols['Erect'] if k=='ERECT' else src_cols['String'])
                        den_col = src_cols['Locs'] if k!='STRING' else src_cols['Length']
                        
                        num = self.to_float(src_row[num_col])
                        den = self.to_float(src_row[den_col])
                        
                        if den > 0: val = round(num/den, 2)
                    except: pass
                if rule.startswith('CALC_'):
                    pass # Handled above
                elif rule == 'InterIntra':
                    val = src_row.get('InterIntra')
                elif rule == 'Scheme':
                    val = src_row.get('Scheme')
                else:
                    src_col_name = src_cols.get(rule)
                    if src_col_name:
                        val = src_row[src_col_name]
                
                if val is not None and not pd.isna(val):
                    # Clean the value if string
                    if isinstance(val, str):
                        # Use is_number to check if it's numeric string (Ported from extractor)
                        is_num, num_val = self.is_number(val)
                        cell.value = num_val if is_num else val
                    else:
                        cell.value = val
                    updates += 1

        # APPEND Unmatched Records
        unmatched_indices = [i for i in range(len(src_keys)) if i not in matched_src_indices]
        
        if unmatched_indices:
            print(f"  [Element Status] Appending {len(unmatched_indices)} new records...")
            for idx in unmatched_indices:
                src_row = src_data[src_keys[idx]]
                
                # Create empty row list sufficient to cover max column index
                max_col = max(30, max(self.mapping_rules.keys()))
                new_row = [None] * (max_col + 1) # +1 for safety/padding
                
                # Transmission Scope at index 4 (Col 5)
                new_row[4] = src_row[src_cols['Scope']]
                
                for col_idx, rule in self.mapping_rules.items():
                    val = None
                    if rule.startswith('CALC_'):
                        try:
                            k = rule.split('_')[1]
                            num_col = src_cols['Found'] if k=='FOUND' else (src_cols['Erect'] if k=='ERECT' else src_cols['String'])
                            den_col = src_cols['Locs'] if k!='STRING' else src_cols['Length']
                            
                            num = self.to_float(src_row[num_col])
                            den = self.to_float(src_row[den_col])
                            
                            if den > 0: val = round(num/den, 2)
                        except: pass
                    if rule.startswith('CALC_'):
                        pass # Handled above
                    elif rule == 'InterIntra':
                        val = src_row.get('InterIntra')
                    elif rule == 'Scheme':
                        val = src_row.get('Scheme')
                    else:
                        src_col_name = src_cols.get(rule)
                        if src_col_name: val = src_row[src_col_name]

                    if val is not None and not pd.isna(val):
                         if isinstance(val, str):
                             is_num, num_val = self.is_number(val)
                             new_row[col_idx-1] = num_val if is_num else val # openpyxl append takes list, 0-indexed relative to row? No, list maps to A, B, C...
                         else:
                             new_row[col_idx-1] = val
                
                ws.append(new_row)
                
        wb.save(target_excel_path)
        print(f"  [Element Status] Completed. {updates} updates made.")
