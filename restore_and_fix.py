"""
Restore backup and apply correct deletions:
1. Restore original data from backup
2. Delete Sr.No. 65-75 (duplicate rows) - keep this fix
3. Delete Sr.No. 194 and 195 (NOT 192, 193) - correct deletion
4. Apply all other fixes again
"""
import pandas as pd
from openpyxl import load_workbook
import shutil
import sys
import re
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

backup_path = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2)_backup_20260202_120650.xlsx"
excel_path = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx"

print("="*100)
print("RESTORING BACKUP AND APPLYING CORRECT DELETIONS")
print("="*100)

# Step 1: Restore backup
print("\n1. Restoring from backup...")
shutil.copy(backup_path, excel_path)
print(f"   Restored from: {backup_path}")

# Load workbook
wb = load_workbook(excel_path)
ws = wb["Data to be captured"]

# Verify original data
df_original = pd.read_excel(excel_path, sheet_name="Data to be captured", header=1)
print(f"   Original row count: {len(df_original)}")

# Step 2: Find and delete Sr.No. 65-75 (duplicates without App IDs)
print("\n2. Deleting duplicate rows Sr.No. 65-75 (without App IDs)...")
srno_col = 2
app_id_col = 9

rows_to_delete_65_75 = []
for row in range(3, ws.max_row + 1):
    srno_val = ws.cell(row=row, column=srno_col).value
    if srno_val is not None:
        try:
            srno = float(srno_val)
            if 65 <= srno <= 75:
                app_id = ws.cell(row=row, column=app_id_col).value
                if app_id is None or str(app_id).strip() == '':
                    rows_to_delete_65_75.append(row)
        except (ValueError, TypeError):
            pass

print(f"   Found {len(rows_to_delete_65_75)} duplicate rows to delete")

# Delete from bottom to top
for row in sorted(rows_to_delete_65_75, reverse=True):
    ws.delete_rows(row)
    print(f"   Deleted row {row}")

# Step 3: Find and delete Sr.No. 194 and 195 (CORRECT deletion)
print("\n3. Deleting Sr.No. 194 and 195...")
rows_to_delete_194_195 = []
for row in range(3, ws.max_row + 1):
    srno_val = ws.cell(row=row, column=srno_col).value
    if srno_val is not None:
        try:
            srno = float(srno_val)
            if srno == 194 or srno == 195:
                developer = ws.cell(row=row, column=7).value
                print(f"   Found Sr.No. {int(srno)} at row {row}: {str(developer)[:50] if developer else 'N/A'}")
                rows_to_delete_194_195.append(row)
        except (ValueError, TypeError):
            pass

print(f"   Found {len(rows_to_delete_194_195)} rows to delete")

# Delete from bottom to top
for row in sorted(rows_to_delete_194_195, reverse=True):
    ws.delete_rows(row)
    print(f"   Deleted row {row}")

# Step 4: Apply Sr.No. 15 fix - Application Quantum
print("\n4. Fixing Sr.No. 15 - Application Quantum...")
for row in range(3, ws.max_row + 1):
    srno_val = ws.cell(row=row, column=srno_col).value
    if srno_val is not None:
        try:
            srno = float(srno_val)
            if srno == 15:
                # Column for Application Quantum (MW)(ST II)
                for col in range(1, 50):
                    header = ws.cell(row=2, column=col).value
                    if header and 'application quantum' in str(header).lower() and 'st ii' in str(header).lower():
                        ws.cell(row=row, column=col).value = "Connectivity: 880"
                        print(f"   Fixed Application Quantum at row {row}")
                        break
        except:
            pass

# Step 5: Fix Sr.No. 58 - Substation and CMETS
print("\n5. Fixing Sr.No. 58 - Substation and CMETS...")
for row in range(3, ws.max_row + 1):
    srno_val = ws.cell(row=row, column=srno_col).value
    if srno_val is not None:
        try:
            srno = float(srno_val)
            if srno == 58:
                # Substation = Column E (5)
                ws.cell(row=row, column=5).value = "Ramgarh-II PS"
                # CMETS GNA Approved = Column L (12)
                ws.cell(row=row, column=12).value = 33
                print(f"   Fixed Substation and CMETS at row {row}")
        except:
            pass

# Step 6: Fix Sr.No. 301, 302, 303 - Split Substation and App ID
print("\n6. Fixing Sr.No. 301, 302, 303 - Substation and App ID...")
for row in range(3, ws.max_row + 1):
    srno_val = ws.cell(row=row, column=srno_col).value
    if srno_val is not None:
        try:
            srno = float(srno_val)
            if srno in [301, 302, 303]:
                substation = ws.cell(row=row, column=5).value
                if substation:
                    sub_str = str(substation).strip()
                    match = re.match(r'^(.+?)\s{2,}(\d{9,10})$', sub_str)
                    if match:
                        actual_substation = match.group(1).strip()
                        app_id_from_substation = match.group(2).strip()
                        
                        ws.cell(row=row, column=5).value = actual_substation
                        
                        existing_app_id = ws.cell(row=row, column=9).value
                        if existing_app_id and '(Enh' in str(existing_app_id):
                            ws.cell(row=row, column=9).value = f"{app_id_from_substation} (Enh.)"
                        else:
                            ws.cell(row=row, column=9).value = app_id_from_substation
                        
                        print(f"   Fixed Sr.No. {int(srno)}: Substation='{actual_substation}', AppID='{app_id_from_substation}'")
        except:
            pass

# Step 7: Fix Bhuj coordinates
print("\n7. Fixing 765/400/220kV Bhuj coordinates...")
coords_col = 6
substation_col_idx = 5

for row in range(3, ws.max_row + 1):
    substation_val = ws.cell(row=row, column=substation_col_idx).value
    if substation_val and '765' in str(substation_val) and 'Bhuj' in str(substation_val):
        substation_str = str(substation_val).strip()
        if 'Bhuj-II' in substation_str or 'Bhuj II' in substation_str:
            ws.cell(row=row, column=coords_col).value = "23.3750° N, 69.1423° E"
        else:
            ws.cell(row=row, column=coords_col).value = "23.2555° N, 69.6670° E"
        print(f"   Fixed coordinates at row {row}")

# Step 8: Fix all CMETS with "33, 34"
print("\n8. Fixing CMETS GNA Approved (33, 34 -> single value)...")
cmets_col = 12
fixed_cmets = 0
for row in range(3, ws.max_row + 1):
    cmets_val = ws.cell(row=row, column=cmets_col).value
    if cmets_val:
        val_str = str(cmets_val)
        if '33' in val_str and '34' in val_str:
            ws.cell(row=row, column=cmets_col).value = 33
            fixed_cmets += 1
print(f"   Fixed {fixed_cmets} rows")

# Save
print("\n" + "="*100)
print("SAVING ALL CHANGES")
print("="*100)
wb.save(excel_path)
print(f"Saved to: {excel_path}")

# Verify
print("\n" + "="*100)
print("VERIFICATION")
print("="*100)

df = pd.read_excel(excel_path, sheet_name="Data to be captured", header=1)
print(f"\nTotal rows: {len(df)}")

# Check Sr.No. 192, 193 still exist
srno_col_name = 'Sr.no.'
for target in [192, 193]:
    exists = len(df[df[srno_col_name] == target]) > 0
    print(f"Sr.No. {target}: {'EXISTS' if exists else 'MISSING'}")

# Check Sr.No. 194, 195 are deleted
for target in [194, 195]:
    exists = len(df[df[srno_col_name] == target]) > 0
    print(f"Sr.No. {target}: {'EXISTS (ERROR!)' if exists else 'DELETED ✓'}")

# Check duplicates 65-75
rows_65_75 = df[(df[srno_col_name] >= 65) & (df[srno_col_name] <= 75)]
print(f"\nRows with Sr.No. 65-75: {len(rows_65_75)}")
