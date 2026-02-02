"""
Debug script to verify element mappings and PDF extraction
"""
import pandas as pd
import fitz
import re
from openpyxl import load_workbook

EXCEL_PATH = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx"
PDF_33 = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\downloaded_pdfs\SN1\172381548953Minutes of 33rd CMETS NR meeting held on 05.08.2024.pdf"
PDF_34 = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\downloaded_pdfs\SN1\172838877090Minutes of meeting 34th CMETS NR Meeting held on 20-9-24.pdf"

COL_APP_ID = 8
COL_CTS = 39
COL_ATS = 40
COL_DTL = 41

def check_existing_data():
    """Check what data already exists in the columns"""
    print("=" * 60)
    print("CHECKING EXISTING DATA IN EXCEL")
    print("=" * 60)
    
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Data to be captured']
    
    # Check first 10 data rows
    print("\nFirst 15 rows of data:")
    for row in range(3, 18):  # Rows 3-17 (data rows)
        app_id = ws.cell(row=row, column=COL_APP_ID + 1).value
        cts = ws.cell(row=row, column=COL_CTS + 1).value
        ats = ws.cell(row=row, column=COL_ATS + 1).value
        dtl = ws.cell(row=row, column=COL_DTL + 1).value
        
        print(f"Row {row}: AppID='{app_id}', CTS='{cts}', ATS='{ats}', DTL='{dtl}'")
    
    # Count how many rows have empty CTS, ATS, DTL
    empty_cts = 0
    empty_ats = 0  
    empty_dtl = 0
    
    for row in range(3, ws.max_row + 1):
        if not ws.cell(row=row, column=COL_CTS + 1).value:
            empty_cts += 1
        if not ws.cell(row=row, column=COL_ATS + 1).value:
            empty_ats += 1
        if not ws.cell(row=row, column=COL_DTL + 1).value:
            empty_dtl += 1
    
    print(f"\nEmpty columns summary (out of {ws.max_row - 2} data rows):")
    print(f"  Empty CTS: {empty_cts}")
    print(f"  Empty ATS: {empty_ats}")
    print(f"  Empty DTL: {empty_dtl}")
    
    wb.close()

def check_element_status():
    """Check Element Status sheet structure"""
    print("\n" + "=" * 60)
    print("CHECKING ELEMENT STATUS SHEET")
    print("=" * 60)
    
    df = pd.read_excel(EXCEL_PATH, sheet_name='Element Status', header=None)
    
    # Show first 10 rows
    print("\nFirst 10 rows (all columns):")
    for i in range(10):
        row_data = df.iloc[i].dropna().tolist()
        print(f"Row {i}: {row_data[:5]}...")  # First 5 non-empty values
    
    # Count Element Codes
    codes = []
    for i in range(3, len(df)):
        code = df.iloc[i, 0]
        if pd.notna(code) and str(code).strip().startswith('EL-'):
            codes.append(str(code).strip())
    
    print(f"\n Total Element Codes found: {len(codes)}")
    print(f"Sample codes: {codes[:10]}")

def check_pdf_extraction():
    """Check what we're extracting from PDFs"""
    print("\n" + "=" * 60)
    print("CHECKING PDF EXTRACTION")
    print("=" * 60)
    
    doc = fitz.open(PDF_33)
    text = ""
    for page in doc:
        text += page.get_text("text")
    doc.close()
    
    # Find transmission system sections
    pattern = r'Details of Transmission system for Connectivity under GNA:(.*?)(?=Details of Transmission|$)'
    matches = re.findall(pattern, text, re.DOTALL)
    
    print(f"\nFound {len(matches)} transmission sections in PDF 33")
    
    if matches:
        print("\nFirst section sample:")
        sample = matches[0][:1000].encode('ascii', 'replace').decode('ascii')
        print(sample)
    
    # Find specific Application ID mentioned in Excel
    # App ID 2200000516 is in Excel, let's find its section
    print("\n\nLooking for App ID 2200000516 section...")
    
    idx = text.find("2200000516")
    if idx > 0:
        context = text[max(0, idx-200):idx+2000]
        print("Found context:")
        print(context.encode('ascii', 'replace').decode('ascii'))
    else:
        print("Not found in PDF 33")

def check_specific_app_id():
    """Check a specific Application ID that's in the Excel"""
    print("\n" + "=" * 60)
    print("CHECKING SPECIFIC APPLICATION ID")
    print("=" * 60)
    
    # First, find app IDs in Excel that are in PDF
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Data to be captured']
    
    doc = fitz.open(PDF_33)
    pdf_text = ""
    for page in doc:
        pdf_text += page.get_text("text")
    doc.close()
    
    matching_ids = []
    for row in range(3, min(50, ws.max_row + 1)):
        app_id = ws.cell(row=row, column=COL_APP_ID + 1).value
        if app_id:
            app_id_str = str(app_id).strip().split()[0]
            app_id_clean = re.sub(r'\D', '', app_id_str)[:10]
            if app_id_clean in pdf_text:
                matching_ids.append((row, app_id_clean))
    
    print(f"Found {len(matching_ids)} matching App IDs in first 50 Excel rows")
    
    for row, app_id in matching_ids[:5]:
        cts = ws.cell(row=row, column=COL_CTS + 1).value
        ats = ws.cell(row=row, column=COL_ATS + 1).value
        dtl = ws.cell(row=row, column=COL_DTL + 1).value
        
        print(f"\nRow {row}, AppID {app_id}:")
        print(f"  Current CTS: '{cts}'")
        print(f"  Current ATS: '{ats}'")
        print(f"  Current DTL: '{dtl}'")
    
    wb.close()

if __name__ == "__main__":
    check_existing_data()
    check_element_status()
    check_pdf_extraction()
    check_specific_app_id()
