"""
Check Sr.No. 301, 302, 303 for Application IDs in Substation column
And verify rows 194 and 195 current state
"""
import pandas as pd
from openpyxl import load_workbook
import sys
import re

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

excel_path = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx"

# Load workbook
wb = load_workbook(excel_path)
ws = wb["Data to be captured"]

print("="*100)
print("CHECKING Sr.No. 301, 302, 303")
print("="*100)

# Find rows with Sr.No. 301, 302, 303
srno_col = 2  # Column B
substation_col = 5  # Column E
app_id_col = 9  # Column I

target_srnos = [301, 302, 303]

for row in range(3, ws.max_row + 1):
    srno_val = ws.cell(row=row, column=srno_col).value
    if srno_val is not None:
        try:
            srno = float(srno_val)
            if srno in target_srnos:
                substation = ws.cell(row=row, column=substation_col).value
                app_id = ws.cell(row=row, column=app_id_col).value
                developer = ws.cell(row=row, column=7).value
                
                print(f"\nExcel Row {row} - Sr.No. {int(srno)}:")
                print(f"  Substation (Col E): {substation}")
                print(f"  App ID (Col I): {app_id}")
                print(f"  Developer (Col G): {str(developer)[:50] if developer else 'N/A'}")
                
                # Check if Substation looks like an Application ID (10-digit number)
                if substation:
                    sub_str = str(substation).strip()
                    if re.match(r'^\d{10}$', sub_str):
                        print(f"  >>> SUBSTATION CONTAINS APPLICATION ID: {sub_str}")
        except (ValueError, TypeError):
            pass

print("\n" + "="*100)
print("CHECKING CURRENT ROWS 194 AND 195 (Excel rows)")
print("="*100)

for excel_row in [194, 195]:
    if excel_row <= ws.max_row:
        srno = ws.cell(row=excel_row, column=2).value
        substation = ws.cell(row=excel_row, column=5).value
        developer = ws.cell(row=excel_row, column=7).value
        app_id = ws.cell(row=excel_row, column=9).value
        
        print(f"\nExcel Row {excel_row}:")
        print(f"  Sr.No.: {srno}")
        print(f"  Substation: {substation}")
        print(f"  Developer: {str(developer)[:50] if developer else 'N/A'}")
        print(f"  App ID: {app_id}")
    else:
        print(f"\nExcel Row {excel_row}: Does not exist (max row: {ws.max_row})")

print("\n" + "="*100)
print("SCANNING ALL SUBSTATION COLUMN FOR APPLICATION IDS")
print("="*100)

# Scan all rows to find any that have App IDs in Substation column
app_id_in_substation = []
for row in range(3, ws.max_row + 1):
    substation = ws.cell(row=row, column=substation_col).value
    if substation:
        sub_str = str(substation).strip()
        # Check if it's a 10-digit number (typical App ID format)
        if re.match(r'^\d{10}$', sub_str):
            srno = ws.cell(row=row, column=srno_col).value
            app_id = ws.cell(row=row, column=app_id_col).value
            app_id_in_substation.append({
                'row': row,
                'srno': srno,
                'substation_value': sub_str,
                'existing_app_id': app_id
            })

print(f"\nFound {len(app_id_in_substation)} rows with Application IDs in Substation column:")
for item in app_id_in_substation:
    print(f"  Row {item['row']}, Sr.No. {item['srno']}: Substation='{item['substation_value']}', Existing AppID='{item['existing_app_id']}'")
