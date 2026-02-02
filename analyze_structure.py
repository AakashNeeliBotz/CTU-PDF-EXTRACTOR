"""
Script to analyze the Excel file structure more carefully
"""
import pandas as pd
from openpyxl import load_workbook

# Load the Excel file
excel_path = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx"

# Load workbook with openpyxl to see actual structure
wb = load_workbook(excel_path)
ws = wb["Data to be captured"]

print("="*80)
print("First 5 rows of the sheet (to understand header structure):")
print("="*80)
for row_idx in range(1, 6):
    print(f"\nRow {row_idx}:")
    for col_idx in range(1, 15):  # First 14 columns
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value:
            print(f"  Col {col_idx} ({chr(64+col_idx)}): {cell.value}")

print("\n" + "="*80)
print("Total rows with data:", ws.max_row)
print("Total columns:", ws.max_column)
print("="*80)

# Check which row contains the actual headers
print("\n" + "="*80)
print("Looking for header row with 'Sr' or 'Region' or 'State':")
print("="*80)
header_row = None
for row_idx in range(1, 10):
    row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, 20)]
    row_str = ' '.join([str(v) if v else '' for v in row_values])
    if 'sr' in row_str.lower() and ('no' in row_str.lower() or 'region' in row_str.lower()):
        header_row = row_idx
        print(f"Found header at row {row_idx}: {row_values[:10]}")
        break

# If we found the header row, let's read again with the correct header
if header_row:
    print(f"\n" + "="*80)
    print(f"Reading Excel with header at row {header_row}")
    print("="*80)
    df = pd.read_excel(excel_path, sheet_name="Data to be captured", header=header_row-1)
    print("\nColumn names:")
    for i, col in enumerate(df.columns[:20]):
        print(f"  Col {i} ({chr(65+i) if i < 26 else 'A'+chr(65+i-26)}): {col}")
    
    print(f"\nTotal data rows: {len(df)}")
    
    # Find key columns
    srno_col = None
    substation_col = None
    cmets_gna_col = None
    app_id_col = None
    
    for col in df.columns:
        col_lower = str(col).lower()
        if 'sr' in col_lower and 'no' in col_lower:
            srno_col = col
        if 'substation' in col_lower:
            substation_col = col
        if 'cmets' in col_lower and 'gna' in col_lower and 'approved' in col_lower:
            cmets_gna_col = col
        if 'gna' in col_lower and 'application' in col_lower and 'id' in col_lower:
            app_id_col = col
    
    print(f"\nKey columns found:")
    print(f"  Sr.No. column: {srno_col}")
    print(f"  Substation column: {substation_col}")
    print(f"  CMETS GNA Approved column: {cmets_gna_col}")
    print(f"  Application ID column: {app_id_col}")
