import openpyxl
import pandas as pd

# Check Excel
print("="*60)
print("EXCEL FILE CHECK")
print("="*60)

wb = openpyxl.load_workbook('Connectivity_Application_Data_TEST_SN3_betterPrompts2.xlsx')
ws = wb['Data to be captured']

print(f'\nSheet: {ws.title}')
print(f'Total rows: {ws.max_row}')
print(f'Data rows: {ws.max_row - 2} (excluding header rows)')
print(f'Total columns: {ws.max_column}')

# Check first row with data
print(f'\nFirst data row (row 3) - first 5 columns:')
for col in range(2, 7):
    value = ws.cell(3, col).value
    header = ws.cell(2, col).value
    print(f'  {header}: {value}')

# Count non-empty rows
non_empty_count = 0
for row in range(3, ws.max_row + 1):
    # Check if any cell in the row has data
    has_data = any(ws.cell(row, col).value for col in range(2, ws.max_column + 1))
    if has_data:
        non_empty_count += 1

print(f'\nNon-empty data rows: {non_empty_count}')

# Check CSV for comparison
print(f"\n{'='*60}")
print("CSV FILE CHECK")
print("="*60)

df = pd.read_csv('extraction_output/Data_to_be_captured_extracted_data.csv')
print(f'\nCSV shape: {df.shape}')
print(f'CSV rows: {len(df)}')
print(f'CSV columns: {len(df.columns)}')

# Show column names that got normalized
normalized_cols = [col for col in df.columns if '_' in col and not col.endswith('_1')]
print(f'\nNormalized columns (sample): {normalized_cols[:5]}')
