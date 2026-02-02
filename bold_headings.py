import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

def make_headings_bold():
    """Make the headings bold in the Excel file"""
    file_path = 'Connectivity_Application_Data_TEST_ALL_SHEETS38.xlsx'
    
    print("Making headings bold...")
    
    # Load the workbook
    wb = load_workbook(file_path)
    
    # For each sheet, make the header row bold
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        print(f"Processing sheet: {sheet_name}")
        
        # Assuming headers are in row 1 (and row 2 in some cases)
        # Make row 1 bold (headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # If there's a second header row, make it bold too
        if ws.max_row >= 2:
            for cell in ws[2]:
                cell.font = Font(bold=True)
    
    # Save the workbook
    wb.save(file_path)
    print(f"Heading formatting completed for: {file_path}")

def main():
    print("Starting heading bold formatting process...")
    make_headings_bold()
    print("Process completed!")

if __name__ == "__main__":
    main()