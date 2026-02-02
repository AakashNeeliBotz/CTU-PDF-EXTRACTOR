"""Verify CMETS Fill"""
import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx')
ws = wb['Data to be captured']

def normalize(text):
    if not text: return ""
    import re
    text = str(text).lower()
    return re.sub(r'[^a-z0-9]', '', text)

targets = [
    ('Adani Renewable Energy  Holding Four Limited', 'Ramgarh PS'),
    ('ACC Limited', 'UNKNOWN'),
    ('Renew Solar  Power Private  Li', 'Bhadla-III PS')
]

print('=== Verifying Updates ===')
for t_dev, t_sub in targets:
    print(f'\nChecking Group: {t_dev} @ {t_sub}')
    for row in range(3, ws.max_row + 1):
        dev = ws.cell(row=row, column=7).value
        sub = ws.cell(row=row, column=5).value
        
        n_dev = normalize(dev)
        n_sub = normalize(sub)
        
        target_n_dev = normalize(t_dev)
        target_n_sub = normalize(t_sub)
        
        # approximate match for verification
        if target_n_dev in n_dev and (target_n_sub in n_sub or t_sub == 'UNKNOWN'):
            cmets = ws.cell(row=row, column=12).value
            date = ws.cell(row=row, column=14).value
            print(f'Row {row}: {cmets} | {date}')
            
print('\n=== Verify Empty CMETS GNA check ===')
empty_count = 0
for row in range(3, ws.max_row + 1):
    # Only check if Developer is present (otherwise it's an empty row)
    if ws.cell(row=row, column=7).value and not ws.cell(row=row, column=12).value:
        empty_count += 1
print(f'Rows with Developer but EMPTY CMETS GNA: {empty_count}')
