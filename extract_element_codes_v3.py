"""
Final extraction script with Annexure parsing for CTS Element Code mapping.
Parses transmission system sections from PDFs and matches to Element Codes.
"""
import pandas as pd
import fitz  # PyMuPDF
import re
from openpyxl import load_workbook
from difflib import SequenceMatcher
import warnings
warnings.filterwarnings('ignore')

# File paths
EXCEL_PATH = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx"
PDF_33 = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\downloaded_pdfs\SN1\172381548953Minutes of 33rd CMETS NR meeting held on 05.08.2024.pdf"
PDF_34 = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\downloaded_pdfs\SN1\172838877090Minutes of meeting 34th CMETS NR Meeting held on 20-9-24.pdf"

# Column positions in Excel (0-indexed, add +1 for openpyxl)
COL_APP_ID = 8   # GNA/ST II Application ID
COL_CTS = 39     # CTS Element Unique Code
COL_ATS = 40     # ATS Element Unique Code  
COL_DTL = 41     # DTL Element Code Unique
DATA_START_ROW = 3

def extract_pdf_text(pdf_path):
    """Extract text from PDF"""
    doc = fitz.open(pdf_path)
    full_text = ""
    for page in doc:
        full_text += page.get_text("text")
    doc.close()
    return full_text

def normalize_text(text):
    """Normalize text for matching"""
    if not text:
        return ""
    text = str(text).lower()
    text = ' '.join(text.split())
    # Normalize special characters
    text = text.replace('–', '-').replace('—', '-')
    text = re.sub(r'[^\w\s\-/]', ' ', text)
    return text.strip()

def similarity_score(s1, s2):
    """Calculate string similarity"""
    return SequenceMatcher(None, normalize_text(s1), normalize_text(s2)).ratio()

def load_element_status():
    """Load Element Status sheet and build description -> code mapping"""
    print("[*] Loading Element Status sheet...")
    df = pd.read_excel(EXCEL_PATH, sheet_name='Element Status', header=None)
    
    element_map = {}  # normalized_description -> code
    code_to_desc = {} # code -> original_description
    
    for i in range(len(df)):
        for j in range(min(6, len(df.columns))):
            cell_val = df.iloc[i, j]
            if pd.notna(cell_val):
                cell_str = str(cell_val).strip()
                # Match Element Code pattern: EL-XXXXX
                if re.match(r'^EL-[A-Z0-9]{5}$', cell_str):
                    # Get description from appropriate column
                    desc_col = 3 if j < 3 else 4
                    if desc_col < len(df.columns):
                        desc = df.iloc[i, desc_col]
                        if pd.notna(desc) and len(str(desc).strip()) > 10:
                            desc_str = str(desc).strip()
                            element_map[normalize_text(desc_str)] = cell_str
                            code_to_desc[cell_str] = desc_str
    
    print(f"[+] Loaded {len(code_to_desc)} Element Codes")
    return element_map, code_to_desc

def find_element_codes(description, element_map, threshold=0.60):
    """Find matching Element Codes for a description"""
    if not description:
        return []
    
    desc_norm = normalize_text(description)
    matches = []
    
    # Exact match
    if desc_norm in element_map:
        return [element_map[desc_norm]]
    
    # Partial/substring matching
    for key, code in element_map.items():
        if len(key) > 15:
            # Check if description contains key phrases
            key_words = set(key.split())
            desc_words = set(desc_norm.split())
            # If most key words are in description
            common = key_words & desc_words
            if len(common) >= len(key_words) * 0.6:
                matches.append(code)
                continue
            
            # Substring check
            if key in desc_norm or desc_norm in key:
                matches.append(code)
    
    if matches:
        return list(set(matches))
    
    # Fuzzy matching for remaining
    candidates = []
    for key, code in element_map.items():
        score = similarity_score(desc_norm, key)
        if score >= threshold:
            candidates.append((score, code))
    
    candidates.sort(reverse=True)
    return [code for score, code in candidates[:5]]

def extract_annexures(pdf_text):
    """Extract Annexure sections from PDF text"""
    annexures = {}
    
    # Find all Annexure sections
    pattern = r'Annexure-([IVX]+|\d+)\s*\n+(.*?)(?=Annexure-[IVX\d]+|Minutes of \d+|$)'
    matches = re.findall(pattern, pdf_text, re.DOTALL | re.IGNORECASE)
    
    for num, content in matches:
        # Clean the content
        content = re.sub(r'Minutes of \d+.*?Page \d+ of \d+', '', content, flags=re.DOTALL)
        
        # Extract numbered items
        items = []
        item_pattern = r'(\d+)\.\s*\n?(.+?)(?=\d+\.\s|\n\n|Additional|$)'
        item_matches = re.findall(item_pattern, content, re.DOTALL)
        
        for item_num, item_text in item_matches:
            cleaned = ' '.join(item_text.split())
            if len(cleaned) > 20:
                items.append(cleaned)
        
        # Also try bullet points
        if not items:
            bullet_pattern = r'[•\-]\s*(.+?)(?=[•\-]|$)'
            bullets = re.findall(bullet_pattern, content, re.DOTALL)
            items = [' '.join(b.split()) for b in bullets if len(b.strip()) > 20]
        
        annexures[num.upper()] = items
        # Also store with Roman numeral variations
        if num.upper() == 'II':
            annexures['2'] = items
        elif num.upper() == 'I':
            annexures['1'] = items
    
    return annexures

def extract_app_transmission_data(pdf_text, annexures):
    """
    Extract transmission data (ATS, DTL, CTS) for each Application ID
    """
    results = {}
    
    # Split by "Details of Transmission system for Connectivity under GNA:"
    sections = re.split(
        r'Details of Transmission system for Connectivity under GNA:',
        pdf_text,
        flags=re.IGNORECASE
    )
    
    for i, section in enumerate(sections[1:], 1):
        # Get context to find App IDs
        prev_text = sections[i-1][-2500:] if len(sections[i-1]) > 2500 else sections[i-1]
        app_ids = set(re.findall(r'\b22\d{8}\b', prev_text))
        app_ids.update(re.findall(r'\b22\d{8}\b', section[:2000]))
        
        # Extract ATS (Section A)
        ats_text = ""
        ats_match = re.search(
            r'A\.\s*Associated\s*Transmission\s*System\s*\(ATS\)[:\s]*(.+?)(?=B\.\s*Transmission|$)',
            section, re.DOTALL | re.IGNORECASE
        )
        if ats_match:
            ats_raw = ats_match.group(1).strip()
            # Clean and check for NIL
            ats_clean = ' '.join(ats_raw.split())
            if 'NIL' in ats_clean.upper():
                ats_text = "NIL"
            else:
                ats_text = ats_clean[:500]
        
        # Extract DTL (Section B)
        dtl_text = ""
        dtl_match = re.search(
            r'B\.\s*Transmission\s*System\s*under\s*applicant\s*scope[:\s]*(.+?)(?=C\.\s*Transmission|$)',
            section, re.DOTALL | re.IGNORECASE
        )
        if dtl_match:
            dtl_raw = dtl_match.group(1)
            # Extract (i), (ii), etc. items
            items = re.findall(r'\([ivx]+\)\.*\s*(.+?)(?=\([ivx]+\)|C\.|Minutes|$)',
                             dtl_raw, re.DOTALL | re.IGNORECASE)
            if items:
                dtl_text = ' '.join([' '.join(it.split()) for it in items])[:600]
            else:
                dtl_text = ' '.join(dtl_raw.split())[:600]
        
        # Extract CTS (Section C) - check for Annexure reference
        cts_text = ""
        cts_items = []
        cts_match = re.search(
            r'C\.\s*Transmission\s*system\s*for\s*Connectivity\s*under\s*GNA[:\s]*(.+?)(?=Start Date|Sl\.|Minutes|$)',
            section, re.DOTALL | re.IGNORECASE
        )
        if cts_match:
            cts_raw = cts_match.group(1).strip()
            
            # Check for Annexure reference
            annexure_ref = re.search(r'Annexure[- ]?([IVX]+|\d+)', cts_raw, re.IGNORECASE)
            if annexure_ref:
                annex_num = annexure_ref.group(1).upper()
                if annex_num in annexures:
                    cts_items = annexures[annex_num]
                    cts_text = "ANNEXURE:" + annex_num
            else:
                cts_text = ' '.join(cts_raw.split())[:500]
        
        # Store for each App ID
        for app_id in app_ids:
            if app_id not in results:
                results[app_id] = {
                    'ATS': '',
                    'DTL': '',
                    'CTS': '',
                    'CTS_items': []
                }
            
            # Update with non-empty values (don't overwrite)
            if ats_text and not results[app_id]['ATS']:
                results[app_id]['ATS'] = ats_text
            if dtl_text and not results[app_id]['DTL']:
                results[app_id]['DTL'] = dtl_text
            if cts_text and not results[app_id]['CTS']:
                results[app_id]['CTS'] = cts_text
            if cts_items and not results[app_id]['CTS_items']:
                results[app_id]['CTS_items'] = cts_items
    
    return results

def update_excel(pdf_data, element_map, code_to_desc):
    """Update Excel with Element Codes"""
    print("\n[*] Opening Excel workbook...")
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Data to be captured']
    
    updates = {'CTS': 0, 'ATS': 0, 'DTL': 0}
    seen_app_ids = set()  # For duplicate handling
    
    for row in range(DATA_START_ROW, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=COL_APP_ID + 1).value
        
        if not cell_val:
            continue
        
        # Extract clean Application ID
        app_id_raw = str(cell_val).strip().split()[0]
        app_id = re.sub(r'\D', '', app_id_raw)[:10]
        
        if not app_id or len(app_id) != 10:
            continue
        
        # Skip duplicates (keep first occurrence)
        if app_id in seen_app_ids:
            continue
        seen_app_ids.add(app_id)
        
        # Check if this App ID has PDF data
        if app_id not in pdf_data:
            continue
        
        data = pdf_data[app_id]
        
        # Get current values
        cur_cts = ws.cell(row=row, column=COL_CTS + 1).value
        cur_ats = ws.cell(row=row, column=COL_ATS + 1).value
        cur_dtl = ws.cell(row=row, column=COL_DTL + 1).value
        
        # === Update CTS ===
        if not cur_cts:
            cts_codes = set()
            
            # If we have annexure items, match each one
            if data['CTS_items']:
                for item in data['CTS_items']:
                    codes = find_element_codes(item, element_map)
                    cts_codes.update(codes)
            elif data['CTS'] and not data['CTS'].startswith('ANNEXURE:'):
                codes = find_element_codes(data['CTS'], element_map)
                cts_codes.update(codes)
            
            if cts_codes:
                ws.cell(row=row, column=COL_CTS + 1).value = ','.join(sorted(cts_codes))
                updates['CTS'] += 1
                print(f"  Row {row} ({app_id}): CTS = {','.join(sorted(cts_codes)[:3])}...")
        
        # === Update ATS ===
        if not cur_ats:
            ats_text = data['ATS']
            # Leave blank if NIL (as per requirements)
            if ats_text and ats_text != "NIL":
                ats_codes = find_element_codes(ats_text, element_map)
                if ats_codes:
                    ws.cell(row=row, column=COL_ATS + 1).value = ','.join(sorted(set(ats_codes)))
                    updates['ATS'] += 1
                    print(f"  Row {row} ({app_id}): ATS = {','.join(sorted(set(ats_codes)))}")
        
        # === Update DTL ===
        if not cur_dtl and data['DTL']:
            dtl_codes = find_element_codes(data['DTL'], element_map)
            if dtl_codes:
                ws.cell(row=row, column=COL_DTL + 1).value = ','.join(sorted(set(dtl_codes)))
                updates['DTL'] += 1
                print(f"  Row {row} ({app_id}): DTL = {','.join(sorted(set(dtl_codes)))}")
    
    print(f"\n[*] Saving workbook...")
    wb.save(EXCEL_PATH)
    print("[+] Saved successfully!")
    
    return updates

def main():
    print("=" * 70)
    print("CTU PDF Extractor - Element Code Mapping (v3 - with Annexures)")
    print("=" * 70)
    
    # Step 1: Load Element Status mapping
    element_map, code_to_desc = load_element_status()
    
    if len(element_map) == 0:
        print("[!] ERROR: No element codes loaded!")
        return
    
    # Step 2: Extract from PDF 33
    print("\n[*] Extracting from PDF 33 (33rd CMETS)...")
    text_33 = extract_pdf_text(PDF_33)
    annexures_33 = extract_annexures(text_33)
    print(f"    Found {len(annexures_33)} Annexure sections")
    
    for num, items in list(annexures_33.items())[:3]:
        print(f"      Annexure {num}: {len(items)} items")
    
    pdf_data_33 = extract_app_transmission_data(text_33, annexures_33)
    print(f"[+] Extracted data for {len(pdf_data_33)} Application IDs")
    
    # Step 3: Extract from PDF 34
    print("\n[*] Extracting from PDF 34 (34th CMETS)...")
    text_34 = extract_pdf_text(PDF_34)
    annexures_34 = extract_annexures(text_34)
    print(f"    Found {len(annexures_34)} Annexure sections")
    
    pdf_data_34 = extract_app_transmission_data(text_34, annexures_34)
    print(f"[+] Extracted data for {len(pdf_data_34)} Application IDs")
    
    # Merge data
    all_data = pdf_data_33.copy()
    for app_id, data in pdf_data_34.items():
        if app_id not in all_data:
            all_data[app_id] = data
        else:
            # Merge non-empty values
            for key in ['ATS', 'DTL', 'CTS']:
                if data[key] and not all_data[app_id][key]:
                    all_data[app_id][key] = data[key]
            if data['CTS_items'] and not all_data[app_id]['CTS_items']:
                all_data[app_id]['CTS_items'] = data['CTS_items']
    
    print(f"\n[*] Total unique Application IDs: {len(all_data)}")
    
    # Show sample
    print("\nSample extracted data:")
    for app_id in list(all_data.keys())[:2]:
        d = all_data[app_id]
        print(f"  App {app_id}:")
        print(f"    ATS: {d['ATS'][:50]}..." if d['ATS'] else "    ATS: (empty)")
        print(f"    DTL: {d['DTL'][:50]}..." if d['DTL'] else "    DTL: (empty)")
        print(f"    CTS: {d['CTS']}, items={len(d['CTS_items'])}")
    
    # Step 4: Update Excel
    updates = update_excel(all_data, element_map, code_to_desc)
    
    # Summary
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"  CTS updates: {updates['CTS']}")
    print(f"  ATS updates: {updates['ATS']}")
    print(f"  DTL updates: {updates['DTL']}")
    print(f"  TOTAL:       {sum(updates.values())}")
    print("=" * 70)

if __name__ == "__main__":
    main()
