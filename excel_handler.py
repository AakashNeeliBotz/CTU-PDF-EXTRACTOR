import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook
import shutil
from field_mappings import DATA_TO_BE_CAPTURED_FIELDS


def reindex_dataframe_to_template(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    """
    Reindex DataFrame columns to match template's canonical field order.
    
    Args:
        df: DataFrame with data records
        sheet_name: Name of target sheet
        
    Returns:
        Reindexed DataFrame with columns in template order
    """
    # Get expected fields for this sheet
    if sheet_name == "Data to be captured":
        expected_fields = DATA_TO_BE_CAPTURED_FIELDS
    else:
        # TODO: Add field lists for other sheets
        # For now, just return as-is
        return df
    
    # Reindex to match template order
    # Missing columns will be filled with NaN (converted to None later)
    # Extra columns are dropped
    df_reindexed = df.reindex(columns=expected_fields)
    
    return df_reindexed

def write_to_excel(data_records, template_path, output_path, sheet_name):
    """
    Writes a list of data records to a specific sheet in an Excel file.

    Args:
        data_records (list): A list of dictionaries, where each dictionary is a record.
        template_path (str): The path to the Excel template file.
        output_path (str): The path where the updated Excel file will be saved.
        sheet_name (str): The name of the sheet to write the data to.
    """
    if not data_records:
        print("[~] No data records to write to Excel. Skipping.")
        return

    try:
        # Create a DataFrame from the list of dictionaries
        new_data_df = pd.DataFrame(data_records)
        
        # Reindex to template order to ensure column alignment
        new_data_df = reindex_dataframe_to_template(new_data_df, sheet_name)
        
        print(f"[*] DataFrame reindexed to template order: {list(new_data_df.columns)[:10]}...")

        # Check if the output file already exists to append data
        if os.path.exists(output_path):
            # Use openpyxl engine to load the workbook and append
            with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                # Get the existing sheet to find the last row
                workbook = writer.book
                sheet = workbook[sheet_name]
                start_row = sheet.max_row
                
                # Write the new data, skipping the header, starting at column B (startcol=1 is 0-indexed)
                new_data_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, startcol=1, index=False, header=False)
            print(f"[+] Appended {len(new_data_df)} new records to '{sheet_name}' in '{output_path}'.")

        else:
            # If the output file doesn't exist, create it from the template
            # Copy the template file first to preserve formatting
            shutil.copy(template_path, output_path)
            
            # Now open it and write our data
            wb = load_workbook(output_path)
            
            # Write data to the target sheet
            # Data starts at row 3 (row 1 is empty, row 2 has headers), column B (column A is empty)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Write data starting at row 3, column 2 (B)
                start_row = 3
                start_col = 2
                
                # Write each record as a row
                for record_idx in range(len(new_data_df)):
                    for col_idx, col_name in enumerate(new_data_df.columns):
                        cell_value = new_data_df.iloc[record_idx][col_name]
                        # Convert NaN to None for Excel
                        if pd.isna(cell_value):
                            cell_value = None
                        ws.cell(row=start_row, column=start_col + col_idx, value=cell_value)
                    start_row += 1
                
                wb.save(output_path)
                wb.close()
                
                print(f"[+] Created new output file '{output_path}' and wrote {len(new_data_df)} records to '{sheet_name}'.")
            else:
                wb.close()
                print(f"[!] Sheet '{sheet_name}' not found in template.")


    except FileNotFoundError:
        print(f"[!] Error: The template file was not found at '{template_path}'.")
    except Exception as e:
        print(f"[!] An unexpected error occurred while writing to Excel: {e}")
        import traceback
        traceback.print_exc()

# Maintain backward-compatible interface
write_data_to_sheet = write_to_excel

if __name__ == '__main__':
    # --- For Testing ---
    # This block demonstrates how to use the write_to_excel function.
    
    # Define paths and sheet name
    template_file = "Connectivity Application Data.xlsx"
    output_file = "Connectivity_Application_Data_OUTPUT.xlsx"
    target_sheet = "Data to be captured"
    
    # Create some sample data that matches the Excel columns
    sample_data = [
        {
            "sr_no": 1,
            "region": "NR",
            "state": "Punjab",
            "substation": "Amritsar",
            "name_of_developers": "Brightstar Renewables",
            "application_id": "N-5821",
            "status_of_lta": "Under Implementation",
            "application_quantum_mw": 400
        },
        {
            "sr_no": 2,
            "region": "SR",
            "state": "Tamil Nadu",
            "substation": "Tirunelveli",
            "name_of_developers": "WindFlow Energy",
            "application_id": "S-9904",
            "status_of_lta": "Pending",
            "application_quantum_mw": 300
        }
    ]
    
    print(f"--- Testing Excel Writer ---")
    print(f"Template: {template_file}")
    print(f"Output: {output_file}")
    
    # Ensure the template exists before running
    if os.path.exists(template_file):
        write_to_excel(sample_data, template_file, output_file, target_sheet)
    else:
        print(f"\n[!] Please make sure the Excel template '{template_file}' is in the same directory.")
