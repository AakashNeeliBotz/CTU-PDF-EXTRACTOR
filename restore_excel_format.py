import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

def restore_excel_with_proper_formatting():
    """
    Restore the Excel file with proper formatting and headers preservation
    """
    original_file = 'Connectivity_Application_Data_backup.xlsx'
    corrupted_file = 'Connectivity_Application_Data_TEST_ALL_SHEETS38.xlsx'
    output_file = 'Connectivity_Application_Data_TEST_ALL_SHEETS38.xlsx'
    
    print("Restoring Excel file with proper formatting...")
    
    try:
        # Load the original file to get proper headers
        original_xl = pd.ExcelFile(original_file)
        
        # Load the corrupted file to get the updated data
        corrupted_xl = pd.ExcelFile(corrupted_file)
        
        # Create new workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Process each sheet
        for sheet_name in original_xl.sheet_names:
            print(f"Processing sheet: {sheet_name}")
            
            # Load original sheet to get headers and structure
            original_df = pd.read_excel(original_file, sheet_name=sheet_name, header=None)
            
            if sheet_name == 'Data to be captured':
                # For the main sheet, we need to preserve the headers and add our data
                # Load the corrupted data (our additions)
                corrupted_df = pd.read_excel(corrupted_file, sheet_name=sheet_name, header=None)
                
                # Get headers from original (first 2 rows)
                headers = original_df.iloc[:2]  # First 2 rows contain headers
                
                # Get the data portion from original (excluding headers)
                original_data = original_df.iloc[2:]
                
                # Get our new data (last few rows from corrupted file)
                # Find where our data starts by looking for the pattern
                start_row = len(original_data)
                new_data = corrupted_df.iloc[start_row:]
                
                # Combine everything: headers + original data + new data
                final_df = pd.concat([headers, original_data, new_data], ignore_index=True)
                
                print(f"  Original data rows: {len(original_data)}")
                print(f"  New data rows: {len(new_data)}")
                print(f"  Final data rows: {len(final_df)}")
                
            else:
                # For other sheets, just copy as-is from original
                final_df = original_df
                print(f"  Preserving {len(final_df)} rows unchanged")
            
            # Create worksheet
            ws = wb.create_sheet(title=sheet_name)
            
            # Write data to worksheet
            for r_idx, row in enumerate(dataframe_to_rows(final_df, index=False, header=False), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            print(f"  Sheet '{sheet_name}' restored with {final_df.shape[0]} rows and {final_df.shape[1]} columns")
        
        # Save the workbook
        wb.save(output_file)
        print(f"\nExcel file restored successfully: {output_file}")
        
        # Verify the restoration
        verify_restoration(output_file)
        
    except Exception as e:
        print(f"Error restoring Excel file: {e}")
        import traceback
        traceback.print_exc()

def verify_restoration(file_path):
    """Verify that the restoration worked properly"""
    print("\nVerifying restoration...")
    
    try:
        xl_file = pd.ExcelFile(file_path)
        print(f"Total sheets: {len(xl_file.sheet_names)}")
        print("Sheet names:", xl_file.sheet_names)
        
        # Check Data to be captured sheet specifically
        df_data = pd.read_excel(file_path, sheet_name='Data to be captured', nrows=5, header=None)
        print("\nData to be captured sheet first few rows:")
        print(df_data.head())
        
        # Check that we have proper headers in first 2 rows
        if len(df_data) >= 2:
            print("\nHeader rows:")
            print("Row 0:", df_data.iloc[0].tolist()[:10])  # First 10 columns
            print("Row 1:", df_data.iloc[1].tolist()[:10])  # First 10 columns
        
        print("Restoration verification completed successfully!")
        
    except Exception as e:
        print(f"Error during verification: {e}")

def main():
    print("Starting Excel restoration process...")
    restore_excel_with_proper_formatting()
    print("Process completed!")

if __name__ == "__main__":
    main()