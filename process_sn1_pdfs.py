import fitz
import re
import os
from openpyxl import load_workbook

def extract_meeting_date_from_pdf(pdf_path):
    """Extract the meeting date from the PDF."""
    filename = os.path.basename(pdf_path)
    
    # Try filename first
    filename_date_patterns = [
        r'(\d{1,2}[.\-]\d{1,2}[.\-]\d{2,4})',
    ]
    for pattern in filename_date_patterns:
        match = re.search(pattern, filename)
        if match:
            return match.group(1)
            
    # Try content
    try:
        doc = fitz.open(pdf_path)
        for i in range(min(10, len(doc))):
            text = doc[i].get_text()
            # Look for "held on"
            match = re.search(r'held on\s+(\d{1,2}[.\-/]\d{1,2}[.\-/]\d{2,4})', text, re.IGNORECASE)
            if match:
                return match.group(1)
            # Look for any date
            match = re.search(r'\b(\d{1,2}[.\-/]\d{1,2}[.\-/]\d{2,4})\b', text)
            if match:
                return match.group(1)
        doc.close()
    except Exception as e:
        print(f"Error extracting date from {pdf_path}: {e}")
    return None

def extract_data_from_pdfs(sn1_folder):
    """
    Scans PDFs and returns a dictionary:
    { 
      app_id: {
        'gna': [meeting_num1, ...],
        'lta': [meeting_num2, ...],
        'dates': {meeting_num: date, ...}
      }
    }
    """
    pdf_files = [f for f in os.listdir(sn1_folder) if f.lower().endswith('.pdf')]
    results = {}
    meeting_dates = {}

    for pdf_file in pdf_files:
        path = os.path.join(sn1_folder, pdf_file)
        
        # Determine meeting number
        match = re.search(r'(\d+)(?:st|nd|rd|th)', pdf_file, re.IGNORECASE)
        meeting_num = match.group(1) if match else None
        if not meeting_num:
            num_matches = re.findall(r'\b(\d{2})\b', pdf_file)
            if num_matches: meeting_num = num_matches[-1]
        
        if not meeting_num: continue
        meeting_num = str(int(meeting_num)) # normalize "033" to "33"
        
        date = extract_meeting_date_from_pdf(path)
        meeting_dates[meeting_num] = date
        
        print(f"Processing Meeting {meeting_num} ({pdf_file})...")
        
        try:
            doc = fitz.open(path)
            current_section = "UNKNOWN"
            
            for page in doc:
                text = page.get_text()
                lines = text.split('\n')
                
                for line in lines:
                    # Detect section
                    line_upper = line.upper()
                    if "GNA" in line_upper:
                        current_section = "GNA"
                    elif "LTA" in line_upper:
                        current_section = "LTA"
                        
                    # Find IDs
                    matches = re.findall(r'\b(220000\d{4})\b', line)
                    for app_id in matches:
                        if app_id not in results:
                            results[app_id] = {'gna': set(), 'lta': set()}
                        
                        if current_section == "LTA":
                            results[app_id]['lta'].add(meeting_num)
                        else:
                            # Default to GNA if not sure, or if GNA section
                            results[app_id]['gna'].add(meeting_num)
            doc.close()
        except Exception as e:
            print(f"Error processing {pdf_file}: {e}")
            
    return results, meeting_dates

def update_excel(excel_path, pdf_results, meeting_dates):
    wb = load_workbook(excel_path)
    # Target sheet
    target_sheet = None
    for name in wb.sheetnames:
        if "data to be capture" in name.lower():
            target_sheet = name
            break
            
    if not target_sheet:
        print("Sheet 'Data to be Capture' not found.")
        return
        
    ws = wb[target_sheet]
    
    # Columns (1-based index)
    COL_GNA_ID = 9
    COL_LTA_ID = 10
    COL_ENH_ID = 11
    COL_GNA_APP = 12
    COL_LTA_APP = 13
    COL_GNA_DATE = 14
    COL_LTA_DATE = 15
    
    update_count = 0
    
    for row in range(3, ws.max_row + 1):
        gna_id = str(ws.cell(row=row, column=COL_GNA_ID).value).strip() if ws.cell(row=row, column=COL_GNA_ID).value else None
        lta_id = str(ws.cell(row=row, column=COL_LTA_ID).value).strip() if ws.cell(row=row, column=COL_LTA_ID).value else None
        enh_id = str(ws.cell(row=row, column=COL_ENH_ID).value).strip() if ws.cell(row=row, column=COL_ENH_ID).value else None
        
        # Clean IDs
        def clean_id(val):
            if not val: return None
            val = val.split('.')[0] # remove .0
            m = re.search(r'(220000\d{4})', val)
            return m.group(1) if m else None
            
        gna_id = clean_id(gna_id)
        lta_id = clean_id(lta_id)
        enh_id = clean_id(enh_id)
        
        row_updated = False
        
        # Check GNA/Enhancement matches
        gna_meetings = set()
        for sid in [gna_id, enh_id]:
            if sid and sid in pdf_results:
                gna_meetings.update(pdf_results[sid]['gna'])
        
        if gna_meetings:
            # Get current values
            curr_app = str(ws.cell(row=row, column=COL_GNA_APP).value or "").split(',')
            curr_app = {x.strip() for x in curr_app if x.strip()}
            # Remove placeholder indices if they exist (1, 2, 3)
            curr_app = {x for x in curr_app if x not in ['1', '2', '3', '1.0', '2.0', '3.0']}
            
            # Combine
            new_app_set = curr_app.union(gna_meetings)
            if new_app_set != curr_app:
                sorted_meetings = sorted(list(new_app_set), key=lambda x: int(x))
                ws.cell(row=row, column=COL_GNA_APP).value = ", ".join(sorted_meetings)
                
                # Update dates
                dates = [meeting_dates.get(m) for m in sorted_meetings if meeting_dates.get(m)]
                if dates:
                    ws.cell(row=row, column=COL_GNA_DATE).value = ", ".join(dates)
                
                update_count += 1
                row_updated = True
                print(f"Updated Row {row} GNA: {gna_id or enh_id} -> {ws.cell(row=row, column=COL_GNA_APP).value}")

        # Check LTA matches
        lta_meetings = set()
        if lta_id and lta_id in pdf_results:
            lta_meetings.update(pdf_results[lta_id]['lta'])
            
        if lta_meetings:
            curr_app = str(ws.cell(row=row, column=COL_LTA_APP).value or "").split(',')
            curr_app = {x.strip() for x in curr_app if x.strip()}
            curr_app = {x for x in curr_app if x not in ['1', '2', '3', '1.0', '2.0', '3.0']}
            
            new_app_set = curr_app.union(lta_meetings)
            if new_app_set != curr_app:
                sorted_meetings = sorted(list(new_app_set), key=lambda x: int(x))
                ws.cell(row=row, column=COL_LTA_APP).value = ", ".join(sorted_meetings)
                
                # Update dates
                dates = [meeting_dates.get(m) for m in sorted_meetings if meeting_dates.get(m)]
                if dates:
                    ws.cell(row=row, column=COL_LTA_DATE).value = ", ".join(dates)
                
                if not row_updated: update_count += 1
                print(f"Updated Row {row} LTA: {lta_id} -> {ws.cell(row=row, column=COL_LTA_APP).value}")

    wb.save(excel_path)
    print(f"Finished! Total rows updated: {update_count}")

if __name__ == "__main__":
    SN1_DIR = r'c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\downloaded_pdfs\SN1'
    EXCEL_PATH = r'c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx'
    
    results, dates = extract_data_from_pdfs(SN1_DIR)
    update_excel(EXCEL_PATH, results, dates)
