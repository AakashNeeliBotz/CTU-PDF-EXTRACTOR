import openpyxl

wb = openpyxl.load_workbook('Connectivity_Application_Data_TEST_ALL_SHEETS28.xlsx')
ws = wb['Element Status']

print('=== Element Status Audit ===')
print()

# Check all records
records = []
for r in range(4, 150):
    code = ws.cell(row=r, column=2).value
    element = ws.cell(row=r, column=3).value
    scope = ws.cell(row=r, column=5).value
    awarded = ws.cell(row=r, column=15).value
    if not code and not element:
        break
    records.append({'row': r, 'code': code, 'element': element, 'scope': scope, 'awarded': awarded})

print(f'Total records: {len(records)}')
print()

# Group by developer
developers = {}
for rec in records:
    dev = rec['awarded'] or 'UNKNOWN'
    if dev not in developers:
        developers[dev] = []
    developers[dev].append(rec)

print('=== Records per Developer ===')
for dev, recs in sorted(developers.items(), key=lambda x: -len(x[1])):
    dev_display = str(dev).replace('\n', ' ')[:70]
    print(f'{len(recs):3d} records: {dev_display}')

print()
print('=== Potential Issues ===')

# Check for empty/missing data
empty_element = [r for r in records if not r['element'] or str(r['element']).strip() == '']
if empty_element:
    print(f'- Empty element descriptions: {len(empty_element)} records')
    for rec in empty_element[:3]:
        print(f'    Row {rec["row"]}: code={rec["code"]}')

empty_awarded = [r for r in records if not r['awarded'] or str(r['awarded']).strip() == '']
if empty_awarded:
    print(f'- Empty awarded_to: {len(empty_awarded)} records')

# Check for very short elements (likely parsing errors)
short_elements = [r for r in records if r['element'] and len(str(r['element']).strip()) < 20]
if short_elements:
    print(f'- Very short element descriptions (<20 chars): {len(short_elements)} records')
    for rec in short_elements[:5]:
        print(f'    Row {rec["row"]}: "{rec["element"]}"')

# Check for duplicate element descriptions
from collections import Counter
element_counts = Counter(str(r['element'])[:100] for r in records if r['element'])
duplicates = [(elem, cnt) for elem, cnt in element_counts.items() if cnt > 1]
if duplicates:
    print(f'- Duplicate element descriptions: {len(duplicates)} unique elements appear multiple times')
    for elem, cnt in duplicates[:5]:
        print(f'    {cnt}x: "{elem[:60]}..."')

# Check scope distribution
scope_counts = Counter(r['scope'] for r in records)
print()
print('=== Scope Distribution ===')
for scope, cnt in scope_counts.most_common():
    print(f'  {scope}: {cnt} records')

# Check Data to be captured sheet for withdrawn developers with elements
print()
print('=== Cross-checking with Status ===')
ws2 = wb['Data to be captured']
withdrawn_devs = set()
for r in range(6, 500):
    status = ws2.cell(row=r, column=18).value  # status_of_application column
    name = ws2.cell(row=r, column=7).value  # name column
    if status and 'withdrawn' in str(status).lower() and name:
        withdrawn_devs.add(str(name).replace('\n', ' ')[:50])

print(f'Withdrawn developers in Data sheet: {len(withdrawn_devs)}')
for wd in list(withdrawn_devs)[:5]:
    print(f'  - {wd}')

# Check if any withdrawn developer has elements
issues = []
for dev, recs in developers.items():
    dev_str = str(dev).replace('\n', ' ')
    for wd in withdrawn_devs:
        if wd[:20].lower() in dev_str.lower():
            issues.append(f'{dev_str[:50]} has {len(recs)} elements but appears withdrawn')
            break

if issues:
    print()
    print('=== BUGS FOUND ===')
    for issue in issues:
        print(f'  ! {issue}')
else:
    print('No withdrawn developers with elements found!')

wb.close()
