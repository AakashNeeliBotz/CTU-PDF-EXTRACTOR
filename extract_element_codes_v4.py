"""
Improved extraction - finds Details of Transmission sections by looking BEFORE Application IDs
"""
import pandas as pd
import fitz
import re
from openpyxl import load_workbook
from difflib import SequenceMatcher
import warnings
warnings.filterwarnings('ignore')

EXCEL_PATH = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx"
PDF_33 = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\downloaded_pdfs\SN1\172381548953Minutes of 33rd CMETS NR meeting held on 05.08.2024.pdf"
PDF_34 = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\downloaded_pdfs\SN1\172838877090Minutes of meeting 34th CMETS NR Meeting held on 20-9-24.pdf"

COL_APP_ID = 8
COL_CTS = 39
COL_ATS = 40
COL_DTL = 41
DATA_START_ROW = 3

def extract_pdf_text(pdf_path):
    doc = fitz.open(pdf_path)
    text = ""
    for page in doc:
        text += page.get_text("text")
    doc.close()
    return text

def normalize_text(text):
    if not text:
        return ""
    text = str(text).lower()
    text = ' '.join(text.split())
    text = text.replace('–', '-').replace('—', '-')
    text = re.sub(r'[^\w\s\-/]', ' ', text)
    return text.strip()

def similarity_score(s1, s2):
    return SequenceMatcher(None, normalize_text(s1), normalize_text(s2)).ratio()

def load_element_status():
    print("[*] Loading Element Status sheet...")
    df = pd.read_excel(EXCEL_PATH, sheet_name='Element Status', header=None)
    
    element_map = {}
    code_to_desc = {}
    
    for i in range(len(df)):
        for j in range(min(6, len(df.columns))):
            cell = df.iloc[i, j]
            if pd.notna(cell):
                cell_str = str(cell).strip()
                if re.match(r'^EL-[A-Z0-9]{5}$', cell_str):
                    desc_col = 3 if j < 3 else 4
                    if desc_col < len(df.columns):
                        desc = df.iloc[i, desc_col]
                        if pd.notna(desc) and len(str(desc).strip()) > 10:
                            desc_str = str(desc).strip()
                            element_map[normalize_text(desc_str)] = cell_str
                            code_to_desc[cell_str] = desc_str
    
    print(f"[+] Loaded {len(code_to_desc)} Element Codes")
    return element_map, code_to_desc

def find_element_codes(description, element_map, threshold=0.55):
    if not description:
        return []
    
    desc_norm = normalize_text(description)
    matches = []
    
    if desc_norm in element_map:
        return [element_map[desc_norm]]
    
    for key, code in element_map.items():
        if len(key) > 12:
            if key in desc_norm or desc_norm in key:
                matches.append(code)
                continue
            key_words = set(key.split())
            desc_words = set(desc_norm.split())
            common = key_words & desc_words
            if len(common) >= len(key_words) * 0.5:
                matches.append(code)
    
    if matches:
        return list(set(matches))
    
    candidates = []
    for key, code in element_map.items():
        score = similarity_score(desc_norm, key)
        if score >= threshold:
            candidates.append((score, code))
    
    candidates.sort(reverse=True)
    return [code for score, code in candidates[:5]]

def extract_annexures(pdf_text):
    """Extract Annexure sections"""
    annexures = {}
    pattern = r'Annexure-([IVX]+|\d+)\s*(.*?)(?=Annexure-[IVX\d]+|$)'
    matches = re.findall(pattern, pdf_text, re.DOTALL | re.IGNORECASE)
    
    for num, content in matches:
        content = re.sub(r'Minutes of \d+.*?Page \d+ of \d+', '', content, flags=re.DOTALL)
        items = []
        item_pattern = r'(\d+)\.\s+(.+?)(?=\d+\.\s|\n\n|Additional|$)'
        for _, item_text in re.findall(item_pattern, content, re.DOTALL):
            cleaned = ' '.join(item_text.split())
            if len(cleaned) > 15:
                items.append(cleaned)
        
        if not items:
            for bullet in re.findall(r'[•\-]\s*(.+?)(?=[•\-]|$)', content, re.DOTALL):
                if len(bullet.strip()) > 15:
                    items.append(' '.join(bullet.split()))
        
        key = num.upper()
        annexures[key] = items
        if key == 'II':
            annexures['2'] = items
        elif key == 'I':
            annexures['1'] = items
    
    return annexures

def extract_app_data_improved(pdf_text, annexures):
    """
    Improved extraction that finds App ID detailed sections
    by looking for patterns like:
    - "it was agreed to grant connectivity" + App ID + "Details of Transmission"
    """
    results = {}
    
    # Pattern: Find "agreed to grant" sections that lead to transmission details
    # These are typically formatted as narrative paragraphs followed by "Details of Transmission system"
    
    # Split by "Details of Transmission system for Connectivity under GNA:"
    pattern = r'Details of Transmission system for Connectivity under GNA:\s*'
    parts = re.split(pattern, pdf_text, flags=re.IGNORECASE)
    
    for i, detail_section in enumerate(parts[1:], 1):
        # Get the text BEFORE this "Details of Transmission" section
        # to find which Application ID this belongs to
        prefix = parts[i-1]
        
        # Look for app IDs in the 3000 characters before "Details of Transmission"
        context = prefix[-3000:] if len(prefix) > 3000 else prefix
        
        # Find app IDs mentioned in context - prioritize those near "grant", "agreed", or in tables
        app_ids_in_context = list(re.findall(r'\b22\d{8}\b', context))
        
        # Also look for specific patterns like "connectivity of M/s [Company] (App ID)"
        company_pattern = r'connectivity.*?(\d{10})'
        company_matches = re.findall(company_pattern, context, re.IGNORECASE)
        app_ids_in_context.extend(company_matches)
        
        # Find the most recent/relevant App ID (last one mentioned is usually the subject)
        if not app_ids_in_context:
            continue
        
        # Parse ATS
        ats_text = ""
        ats_match = re.search(
            r'A\.\s*Associated\s*Transmission\s*System\s*\(ATS\)[:\s]*(.+?)(?=B\.\s*Transmission|$)',
            detail_section[:1500], re.DOTALL | re.IGNORECASE
        )
        if ats_match:
            ats_raw = ' '.join(ats_match.group(1).split())
            ats_text = "NIL" if "NIL" in ats_raw.upper() else ats_raw[:400]
        
        # Parse DTL
        dtl_text = ""
        dtl_match = re.search(
            r'B\.\s*Transmission\s*System\s*under\s*applicant\s*scope[:\s]*(.+?)(?=C\.\s*Transmission|$)',
            detail_section[:2000], re.DOTALL | re.IGNORECASE
        )
        if dtl_match:
            dtl_raw = dtl_match.group(1)
            items = re.findall(r'\([ivx]+\)\.*\s*(.+?)(?=\([ivx]+\)|C\.|Minutes|$)',
                             dtl_raw, re.DOTALL | re.IGNORECASE)
            if items:
                dtl_text = ' '.join([' '.join(it.split()) for it in items])[:500]
            else:
                dtl_text = ' '.join(dtl_raw.split())[:500]
        
        # Parse CTS
        cts_text = ""
        cts_items = []
        cts_match = re.search(
            r'C\.\s*Transmission\s*system\s*for\s*Connectivity\s*under\s*GNA[:\s]*(.+?)(?=Start Date|Sl\.|Minutes|$)',
            detail_section[:2500], re.DOTALL | re.IGNORECASE
        )
        if cts_match:
            cts_raw = cts_match.group(1).strip()
            annex_ref = re.search(r'Annexure[- ]?([IVX]+|\d+)', cts_raw, re.IGNORECASE)
            if annex_ref:
                annex_key = annex_ref.group(1).upper()
                if annex_key in annexures:
                    cts_items = annexures[annex_key]
                    cts_text = f"ANNEXURE:{annex_key}"
            else:
                cts_text = ' '.join(cts_raw.split())[:500]
        
        # Associate with the relevant App IDs
        # Use the last few App IDs mentioned (most relevant)
        relevant_app_ids = list(set(app_ids_in_context[-5:]))
        
        for app_id in relevant_app_ids:
            if app_id not in results:
                results[app_id] = {'ATS': '', 'DTL': '', 'CTS': '', 'CTS_items': []}
            
            if ats_text and not results[app_id]['ATS']:
                results[app_id]['ATS'] = ats_text
            if dtl_text and not results[app_id]['DTL']:
                results[app_id]['DTL'] = dtl_text
            if cts_text and not results[app_id]['CTS']:
                results[app_id]['CTS'] = cts_text
            if cts_items and not results[app_id]['CTS_items']:
                results[app_id]['CTS_items'] = cts_items
    
    return results

def update_excel(pdf_data, element_map):
    print("\n[*] Updating Excel workbook...")
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Data to be captured']
    
    updates = {'CTS': 0, 'ATS': 0, 'DTL': 0}
    seen = set()
    
    for row in range(DATA_START_ROW, ws.max_row + 1):
        cell = ws.cell(row=row, column=COL_APP_ID + 1).value
        if not cell:
            continue
        
        app_id = re.sub(r'\D', '', str(cell).split()[0])[:10]
        if not app_id or len(app_id) != 10 or app_id in seen:
            continue
        seen.add(app_id)
        
        if app_id not in pdf_data:
            continue
        
        data = pdf_data[app_id]
        
        cur_cts = ws.cell(row=row, column=COL_CTS + 1).value
        cur_ats = ws.cell(row=row, column=COL_ATS + 1).value
        cur_dtl = ws.cell(row=row, column=COL_DTL + 1).value
        
        # Update CTS
        if not cur_cts:
            cts_codes = set()
            if data['CTS_items']:
                for item in data['CTS_items']:
                    codes = find_element_codes(item, element_map)
                    cts_codes.update(codes)
            elif data['CTS'] and not data['CTS'].startswith('ANNEXURE:'):
                codes = find_element_codes(data['CTS'], element_map)
                cts_codes.update(codes)
            
            if cts_codes:
                ws.cell(row=row, column=COL_CTS + 1).value = ','.join(sorted(cts_codes))
                updates['CTS'] += 1
                print(f"  Row {row} ({app_id}): CTS = {len(cts_codes)} codes")
        
        # Update ATS
        if not cur_ats and data['ATS'] and data['ATS'] != "NIL":
            ats_codes = find_element_codes(data['ATS'], element_map)
            if ats_codes:
                ws.cell(row=row, column=COL_ATS + 1).value = ','.join(sorted(set(ats_codes)))
                updates['ATS'] += 1
        
        # Update DTL
        if not cur_dtl and data['DTL']:
            dtl_codes = find_element_codes(data['DTL'], element_map)
            if dtl_codes:
                ws.cell(row=row, column=COL_DTL + 1).value = ','.join(sorted(set(dtl_codes)))
                updates['DTL'] += 1
    
    print(f"\n[*] Saving...")
    wb.save(EXCEL_PATH)
    print("[+] Done!")
    return updates

def main():
    print("=" * 70)
    print("CTU PDF Extractor - v4 (Improved Section Detection)")
    print("=" * 70)
    
    element_map, _ = load_element_status()
    
    if not element_map:
        print("[!] ERROR: No element codes loaded!")
        return
    
    print("\n[*] Processing PDF 33...")
    text_33 = extract_pdf_text(PDF_33)
    annexures_33 = extract_annexures(text_33)
    print(f"    Annexures found: {len(annexures_33)}")
    data_33 = extract_app_data_improved(text_33, annexures_33)
    print(f"    App IDs with data: {len(data_33)}")
    
    print("\n[*] Processing PDF 34...")
    text_34 = extract_pdf_text(PDF_34)
    annexures_34 = extract_annexures(text_34)
    print(f"    Annexures found: {len(annexures_34)}")
    data_34 = extract_app_data_improved(text_34, annexures_34)
    print(f"    App IDs with data: {len(data_34)}")
    
    # Merge
    all_data = data_33.copy()
    for app_id, data in data_34.items():
        if app_id not in all_data:
            all_data[app_id] = data
        else:
            for key in ['ATS', 'DTL', 'CTS']:
                if data[key] and not all_data[app_id][key]:
                    all_data[app_id][key] = data[key]
            if data['CTS_items'] and not all_data[app_id]['CTS_items']:
                all_data[app_id]['CTS_items'] = data['CTS_items']
    
    print(f"\n[*] Total: {len(all_data)} unique App IDs from PDFs")
    
    # Show some samples with CTS items
    print("\nSample App IDs with CTS items:")
    for app_id in list(all_data.keys())[:3]:
        d = all_data[app_id]
        if d['CTS_items']:
            print(f"  {app_id}: {len(d['CTS_items'])} CTS items")
            for item in d['CTS_items'][:2]:
                print(f"    - {item[:80]}...")
    
    # Update Excel
    updates = update_excel(all_data, element_map)
    
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"  CTS updates: {updates['CTS']}")
    print(f"  ATS updates: {updates['ATS']}")
    print(f"  DTL updates: {updates['DTL']}")
    print(f"  TOTAL:       {sum(updates.values())}")
    print("=" * 70)

if __name__ == "__main__":
    main()
