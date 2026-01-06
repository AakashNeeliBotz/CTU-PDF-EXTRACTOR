"""
Core module for PDF table extraction using camelot
"""
import camelot
import pandas as pd
import os
import openpyxl.styles
import re
from loguru import logger
from .config import CAMELOT_SETTINGS, OUTPUT_DIR, PDFS_DIR

def extract_tables(pdf_path):
    """
    Extract tables from a PDF file using camelot library
    
    Args:
        pdf_path (str): Path to the PDF file
        
    Returns:
        list: List of camelot Table objects
    """
    logger.info(f"Processing file: {pdf_path}")
    
    # Check if file exists
    if not os.path.exists(pdf_path):
        logger.error(f"PDF file not found: {pdf_path}")
        raise FileNotFoundError(f"PDF file not found: {pdf_path}")
    
    try:
        # Use camelot to extract tables
        tables = camelot.read_pdf(pdf_path, **CAMELOT_SETTINGS)
        
        # Post-process tables to fix common alignment issues
        processed_tables = []
        for table in tables:
            processed_table = fix_column_alignment(table)
            processed_tables.append(processed_table)
        
        # Merge related tables
        merged_tables = merge_related_tables(processed_tables)
        
        logger.success(f"Found {len(tables)} tables in the PDF, merged into {len(merged_tables)} logical tables")
        return merged_tables
        
    except Exception as e:
        logger.error(f"Error processing {pdf_path}: {str(e)}")
        raise Exception(f"Error processing {pdf_path}: {str(e)}")

def save_tables(tables, output_format='csv'):
    """
    Save extracted tables to files
    
    Args:
        tables (list): List of camelot Table objects
        output_format (str): Output format ('csv', 'excel', 'json')
    """
    # Create output directory
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        logger.info(f"Created output directory: {OUTPUT_DIR}")
    
    saved_count = 0
    for i, table in enumerate(tables):
        # Set correct file extension
        if output_format == 'excel':
            filename = f"table_{i+1}.xlsx"
        else:
            filename = f"table_{i+1}.{output_format}"
        filepath = os.path.join(OUTPUT_DIR, filename)
        
        try:
            if output_format == 'csv':
                table.df.to_csv(filepath, index=False)
            elif output_format == 'excel':
                table.df.to_excel(filepath, index=False, engine='openpyxl')
            elif output_format == 'json':
                table.df.to_json(filepath, orient='records')
            else:
                raise ValueError(f"Unsupported output format: {output_format}")
                
            logger.info(f"Saved table {i+1} to {filepath}")
            saved_count += 1
        except Exception as e:
            logger.error(f"Failed to save table {i+1} as {output_format}: {str(e)}")
    
    logger.success(f"Successfully saved {saved_count} out of {len(tables)} tables as {output_format}")


def save_consolidated_excel(tables, pdf_filename):
    """
    Save all extracted tables from a PDF into a single Excel file with gaps between tables
    
    Args:
        tables (list): List of camelot Table objects
        pdf_filename (str): Name of the PDF file (without path)
    """
    # Create output directory if it doesn't exist
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        logger.info(f"Created output directory: {OUTPUT_DIR}")
    
    # Generate Excel filename based on PDF name
    excel_filename = os.path.splitext(pdf_filename)[0] + '_consolidated.xlsx'
    excel_filepath = os.path.join(OUTPUT_DIR, excel_filename)
    
    try:
        # Create a single dataframe with gaps between tables
        combined_data = []
        
        for i, table in enumerate(tables):
            # Add table data
            table_data = table.df.copy()
            
            # Add table identifier
            identifier_row = pd.DataFrame([f"--- TABLE {i+1} ---"] + [""] * (len(table_data.columns) - 1)).T
            identifier_row.columns = table_data.columns
            combined_data.append(identifier_row)
            
            # Add the table data
            combined_data.append(table_data)
            
            # Add gap rows (2 empty rows)
            gap_rows = pd.DataFrame([[""] * len(table_data.columns)] * 2, columns=table_data.columns)
            combined_data.append(gap_rows)
        
        # Combine all dataframes
        combined_df = pd.concat(combined_data, ignore_index=True)
        
        # Save to Excel with formatting
        with pd.ExcelWriter(excel_filepath, engine='openpyxl') as writer:
            combined_df.to_excel(writer, sheet_name='All_Tables', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['All_Tables']
            
            # Adjust column widths and add borders
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                # Find the maximum length in the column
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set column width (with a minimum and maximum)
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add borders to all cells
                for cell in column:
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )
        
        logger.success(f"Consolidated Excel file saved: {excel_filepath}")
        
    except Exception as e:
        logger.error(f"Failed to save consolidated Excel file: {str(e)}")


def are_tables_related(table1, table2):
    """
    Check if two tables are related and should be merged
    
    Args:
        table1: First camelot Table object
        table2: Second camelot Table object
        
    Returns:
        bool: True if tables are related, False otherwise
    """
    try:
        df1 = table1.df
        df2 = table2.df
        
        logger.debug(f"Checking if tables are related: table1 shape {df1.shape}, table2 shape {df2.shape}")
        
        # Check if tables have the same number of columns
        if df1.shape[1] != df2.shape[1]:
            logger.debug("Tables have different number of columns")
            return False
            
        # Check if table2 starts with a serial number that continues from table1
        if df1.shape[0] > 0 and df2.shape[0] > 0:
            # Get last serial number from table1 (first column)
            last_row_df1 = df1.iloc[-1, 0]
            first_row_df2 = df2.iloc[0, 0]
            
            logger.debug(f"Last row of table1 first column: '{last_row_df1}', First row of table2 first column: '{first_row_df2}'")
            
            # Check if both are numeric
            if (isinstance(last_row_df1, str) and last_row_df1.strip().isdigit() and
                isinstance(first_row_df2, str) and first_row_df2.strip().isdigit()):
                last_serial = int(last_row_df1.strip())
                first_serial = int(first_row_df2.strip())
                
                # Check if serial numbers are consecutive
                if first_serial == last_serial + 1:
                    logger.debug(f"Found consecutive serial numbers: {last_serial} -> {first_serial}")
                    return True
                else:
                    logger.debug(f"Serial numbers are not consecutive: {last_serial} -> {first_serial}")
        
        # Additional check: If table1 ends with data and table2 starts with data (both have serial numbers)
        # This handles cases where there might be descriptive text between data rows
        if df1.shape[0] > 0 and df2.shape[0] > 0:
            # Look for serial numbers in the last few rows of table1
            for i in range(min(3, df1.shape[0])):  # Check last 3 rows
                row_idx = df1.shape[0] - 1 - i
                cell_value = df1.iloc[row_idx, 0]
                if isinstance(cell_value, str) and cell_value.strip().isdigit():
                    last_serial = int(cell_value.strip())
                    
                    # Look for serial numbers in the first few rows of table2
                    for j in range(min(5, df2.shape[0])):  # Check first 5 rows (increased from 3)
                        cell_value2 = df2.iloc[j, 0]
                        if isinstance(cell_value2, str) and cell_value2.strip().isdigit():
                            first_serial = int(cell_value2.strip())
                            # Check if serial numbers are consecutive
                            logger.debug(f"Checking serial numbers: {last_serial} -> {first_serial}")
                            if first_serial == last_serial + 1:
                                logger.debug(f"Found consecutive serial numbers in deeper check: {last_serial} -> {first_serial}")
                                return True
                            break
                    break
        
        # Check if tables have similar column names (header rows)
        if df1.shape[0] > 0 and df2.shape[0] > 0:
            # Check if first row of table2 looks like a header (similar to table1 header)
            header1 = df1.iloc[0].tolist()
            first_row_df2 = df2.iloc[0].tolist()
            
            # Check if both rows have similar structure
            if len(header1) == len(first_row_df2):
                non_empty_header1 = sum(1 for cell in header1 if str(cell).strip())
                non_empty_first_row_df2 = sum(1 for cell in first_row_df2 if str(cell).strip())
                
                # If both rows have similar number of non-empty cells, they might be headers
                if abs(non_empty_header1 - non_empty_first_row_df2) <= 2:
                    # But if table2 has a serial number in first column, it's likely data, not header
                    first_cell_df2 = str(first_row_df2[0]).strip() if len(first_row_df2) > 0 else ""
                    if first_cell_df2.isdigit():
                        logger.debug(f"Found data row (digit in first column) after header: {first_cell_df2}")
                        return True  # It's data, not header, so tables are related
        
        # Additional check: If table2 starts with a header row but the next row has a serial number
        # that continues from table1, then they should be merged
        if df1.shape[0] > 0 and df2.shape[0] > 1:
            # Check if first row of table2 is a header (by checking if it looks like column names)
            first_row_df2 = df2.iloc[0].tolist()
            # Look for the first data row in table2 (skip header rows)
            second_row_df2 = None
            for i in range(1, min(6, df2.shape[0])):  # Check first 5 rows after header
                cell_value = df2.iloc[i, 0]
                if isinstance(cell_value, str) and cell_value.strip().isdigit():
                    second_row_df2 = cell_value
                    break
            
            logger.debug(f"Second row of table2 first column (first data row): '{second_row_df2}'")
            
            # If the first row looks like a header and the second row has a serial number
            if second_row_df2 is not None:
                # Find the last serial number in table1
                last_serial = None
                for i in range(df1.shape[0] - 1, -1, -1):  # Iterate backwards
                    cell_value = df1.iloc[i, 0]
                    if isinstance(cell_value, str) and cell_value.strip().isdigit():
                        last_serial = int(cell_value.strip())
                        break
                
                if last_serial is not None:
                    first_serial = int(second_row_df2.strip())
                    # Check if serial numbers are consecutive
                    logger.debug(f"Checking serial numbers from end of table1 to start of table2: {last_serial} -> {first_serial}")
                    if first_serial == last_serial + 1:
                        logger.debug(f"Found consecutive serial numbers in header check: {last_serial} -> {first_serial}")
                        return True
        
        logger.debug("Tables are not related")
        return False
    except Exception as e:
        logger.warning(f"Error checking if tables are related: {str(e)}")
        return False


def remove_repeated_headers(df):
    """
    Remove repeated header rows from a dataframe
    
    Args:
        df: pandas DataFrame
        
    Returns:
        pandas DataFrame: DataFrame with repeated headers removed
    """
    if df.shape[0] <= 1:
        return df
    
    # Get the first row as reference header
    header_row = df.iloc[0].tolist()
    
    # Check if any subsequent rows are identical or very similar to the header
    rows_to_drop = []
    
    for i in range(1, df.shape[0]):
        current_row = df.iloc[i].tolist()
        
        # Check if current row is very similar to header row
        if len(current_row) == len(header_row):
            # Count matching non-empty cells
            matching_cells = 0
            total_non_empty = 0
            
            for j in range(len(header_row)):
                header_cell = str(header_row[j]).strip()
                current_cell = str(current_row[j]).strip()
                
                if header_cell:  # If header cell is not empty
                    total_non_empty += 1
                    if header_cell == current_cell:
                        matching_cells += 1
            
            # If most non-empty header cells match, consider it a repeated header
            if total_non_empty > 0 and (matching_cells / total_non_empty) > 0.8:
                rows_to_drop.append(i)
    
    # Drop the repeated header rows
    if rows_to_drop:
        df_cleaned = df.drop(df.index[rows_to_drop]).reset_index(drop=True)
        logger.info(f"Removed {len(rows_to_drop)} repeated header rows")
        return df_cleaned
    
    return df


def merge_related_tables(tables):
    """
    Merge related tables that belong together
    
    Args:
        tables (list): List of camelot Table objects
        
    Returns:
        list: List of merged camelot Table objects
    """
    if not tables:
        return []
    
    merged_tables = []
    i = 0
    
    while i < len(tables):
        current_table = tables[i]
        merged = False
        
        # Check if next table is related to current table
        j = i + 1
        while j < len(tables):
            if are_tables_related(current_table, tables[j]):
                # Merge tables
                try:
                    # Concatenate dataframes
                    merged_df = pd.concat([current_table.df, tables[j].df], ignore_index=True)
                    # Remove repeated headers
                    merged_df = remove_repeated_headers(merged_df)
                    # Update current table with merged data
                    current_table.df = merged_df
                    merged = True
                    j += 1
                except Exception as e:
                    logger.warning(f"Error merging tables {i} and {j}: {str(e)}")
                    break
            else:
                break
        
        # Also remove repeated headers from standalone tables
        try:
            current_table.df = remove_repeated_headers(current_table.df)
        except Exception as e:
            logger.warning(f"Error cleaning table {i}: {str(e)}")
        
        merged_tables.append(current_table)
        if merged:
            i = j
        else:
            i += 1
    
    logger.info(f"Merged {len(tables)} tables into {len(merged_tables)} logical tables")
    return merged_tables


def fix_column_alignment(table):
    """
    Fix common column alignment issues in extracted tables
    
    Args:
        table: camelot Table object
        
    Returns:
        table: processed camelot Table object with fixed alignment
    """
    try:
        # Get the dataframe
        df = table.df.copy()
        
        # Check if we have at least 2 columns
        if df.shape[1] >= 2:
            # Look for cases where serial numbers appear in the wrong column
            # Common pattern: Serial number in first column gets merged with Application ID in second column
            
            for i in range(len(df)):
                # Check if the first column is empty or contains unexpected data
                # and the second column contains what looks like a serial number + data
                first_col = str(df.iloc[i, 0]).strip() if not pd.isna(df.iloc[i, 0]) else ""
                second_col = str(df.iloc[i, 1]).strip() if not pd.isna(df.iloc[i, 1]) else ""
                
                # If first column is empty and second column contains a pattern like "1. 2200000788"
                # where the first part looks like a serial number
                if not first_col and "." in second_col:
                    parts = second_col.split(".", 1)
                    if len(parts) == 2 and parts[0].strip().isdigit():
                        # Move the serial number to the first column
                        df.iloc[i, 0] = parts[0].strip()
                        # Keep the rest in the second column
                        df.iloc[i, 1] = parts[1].strip()
                
                # Another case: serial number and data are in the same cell
                # Try to split them properly
                elif first_col and not second_col and "." in first_col:
                    parts = first_col.split(".", 1)
                    if len(parts) == 2 and parts[0].strip().isdigit():
                        # Move the serial number to the first column
                        df.iloc[i, 0] = parts[0].strip()
                        # Move the rest to the second column
                        df.iloc[i, 1] = parts[1].strip()
            
            # Update the table's dataframe
            table.df = df
        
        return table
    except Exception as e:
        logger.warning(f"Could not fix column alignment: {str(e)}")
        # Return original table if fixing fails
        return table


def extract_required_fields_from_excel(excel_filepath):
    """
    Extract required fields from the consolidated Excel file and save with exact template formatting
    
    Args:
        excel_filepath (str): Path to the consolidated Excel file
        
    Returns:
        pandas.DataFrame: DataFrame with the required fields
    """
    try:
        # Read the consolidated Excel file
        df = pd.read_excel(excel_filepath, sheet_name='All_Tables')
        
        # Load template from Connectivity Application Data.xlsx
        template_file = r"c:\AkashNeeli\data_files\Connectivity Application Data.xlsx"
        template_wb = openpyxl.load_workbook(template_file)
        template_sheet = template_wb["Data to be captured"]
        
        # Get the headers from the template file (row 2)
        template_headers = []
        for col_num in range(1, template_sheet.max_column + 1):
            cell_value = template_sheet.cell(row=2, column=col_num).value
            template_headers.append(str(cell_value) if cell_value is not None else "")
        
        # Initialize lists to store the extracted data
        # First create a dictionary with all template headers
        extracted_data = {}
        for header in template_headers:
            if header:  # Only add non-empty headers
                extracted_data[header] = []
        
        # Add our required data fields (in case they're not in the template)
        required_fields = [
            'Sr.no.', 'Region', 'State', 'GNA/ST II Application ID', 'LTA Application ID',
            'Application/Submission Date', 'Nature of Applicant', 'Application Quantum (MW)(ST II)',
            'Applied Start of Connectivity sought by developer date( start date of connectivity as per the application)',
            'Status of application(Withdrawn / granted. Revoked.)'
        ]
        
        for field in required_fields:
            if field not in extracted_data:
                extracted_data[field] = []
        
        # Keep track of the current table structure
        current_header_structure = None
        column_mapping = {}
        
        # Process each row in the dataframe
        i = 0
        while i < len(df):
            row = df.iloc[i]
            
            # Check if this is a table boundary
            first_cell = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ""
            if "--- TABLE" in first_cell:
                # Reset the header structure when we encounter a new table
                current_header_structure = None
                column_mapping = {}
                i += 1
                continue
            
            # Check if this is a header row (based on common header patterns)
            if current_header_structure is None and len(row) > 0:
                header_row = [str(cell).strip() if not pd.isna(cell) else "" for cell in row]
                
                # Check if this looks like a header row by looking for common header keywords
                header_keywords = ['sl', 'no', 'application', 'id', 'applicant', 'location', 'date', 'nature', 'quantum', 'connectivity', 'region', 'criterion', 'mode']
                header_matches = sum(1 for cell in header_row[:10] if any(keyword in cell.lower().replace('\n', ' ') for keyword in header_keywords))
                
                if header_matches >= 2:  # If at least 2 header keywords are found
                    # This is a header row, determine the column mapping
                    current_header_structure = header_row
                    
                    # Reset column mapping
                    column_mapping = {}
                    
                    # Create column mapping based on header content
                    for j, header in enumerate(header_row):
                        header_clean = header.lower().replace('\n', ' ')
                        if ('sl' in header_clean and 'no' in header_clean) or ('serial' in header_clean):
                            column_mapping['serial'] = j
                        elif 'application' in header_clean and 'id' in header_clean:
                            column_mapping['app_id'] = j
                        elif 'project' in header_clean and 'location' in header_clean:
                            column_mapping['location'] = j
                        elif 'submission' in header_clean and 'date' in header_clean:
                            column_mapping['submission_date'] = j
                        elif 'region' in header_clean and 'date' not in header_clean and 'gnare' not in header_clean:
                            # Only match "region" if it's not part of "GNARE" columns
                            column_mapping['region'] = j  # Region column
                        elif 'nature' in header_clean and 'applicant' in header_clean:
                            column_mapping['nature_applicant'] = j
                        elif 'quantum' in header_clean and 'mw' in header_clean:
                            column_mapping['quantum'] = j
                        elif ('start' in header_clean and 'date' in header_clean) or ('start' in header_clean and 'connectivity' in header_clean):
                            column_mapping['start_date'] = j
                        elif 'applicant' in header_clean and 'nature' not in header_clean:  # Check for 'Applicant' but not 'Nature of Applicant'
                            column_mapping['applicant'] = j
                        elif 'criterion' in header_clean or 'mode' in header_clean:
                            column_mapping['criterion'] = j  # Criterion/Mode column
                            logger.debug(f"Found criterion column at index {j}: {header}")
                    
                    i += 1
                    continue
            
            # Check if this is a data row (starts with a serial number)
            first_cell = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ""
            
            # Check for serial number in different formats:
            # 1. Just a digit (e.g., "1")
            # 2. Digit followed by period (e.g., "1.")
            is_serial_row = False
            serial_number = ""
            
            if first_cell.isdigit():
                is_serial_row = True
                serial_number = first_cell
            elif first_cell.endswith('.') and first_cell[:-1].isdigit():
                is_serial_row = True
                serial_number = first_cell[:-1]  # Remove the period
            
            if is_serial_row:
                # Extract Sr.no
                sr_no = serial_number
                # Find the correct field in our extracted_data
                for field_name in extracted_data.keys():
                    if 'sr' in field_name.lower() and 'no' in field_name.lower():
                        extracted_data[field_name].append(sr_no)
                        break
                else:
                    # Fallback - add to first matching field
                    for field_name in extracted_data.keys():
                        if field_name:  # Non-empty field
                            extracted_data[field_name].append(sr_no)
                            break
                
                # Extract data based on the current table structure
                if current_header_structure is not None and column_mapping:
                    # Use the column mapping for this table structure
                    
                    # Extract Application ID
                    app_id_col = column_mapping.get('app_id')
                    app_id = ""
                    if app_id_col is not None and len(row) > app_id_col and not pd.isna(row.iloc[app_id_col]):
                        app_id = str(row.iloc[app_id_col]).strip()
                    # Fallback to column 1 if no app_id column was identified
                    elif len(row) > 1 and not pd.isna(row.iloc[1]):
                        app_id = str(row.iloc[1]).strip()
                    
                    # Determine if it's GNA/ST II or LTA based on whether it's numeric
                    if app_id.replace('.', '', 1).isdigit():  # Handle decimal numbers too
                        # Find GNA field
                        for field_name in extracted_data.keys():
                            if 'gna' in field_name.lower() and 'application' in field_name.lower() and 'id' in field_name.lower():
                                extracted_data[field_name].append(app_id)
                                break
                        # Find LTA field and add empty value
                        for field_name in extracted_data.keys():
                            if 'lta' in field_name.lower() and 'application' in field_name.lower() and 'id' in field_name.lower():
                                extracted_data[field_name].append("")
                                break
                    else:
                        # Find LTA field
                        for field_name in extracted_data.keys():
                            if 'lta' in field_name.lower() and 'application' in field_name.lower() and 'id' in field_name.lower():
                                extracted_data[field_name].append(app_id)
                                break
                        # Find GNA field and add empty value
                        for field_name in extracted_data.keys():
                            if 'gna' in field_name.lower() and 'application' in field_name.lower() and 'id' in field_name.lower():
                                extracted_data[field_name].append("")
                                break
                    
                    # Extract Application/Submission Date
                    date_col = column_mapping.get('submission_date')
                    submission_date = ""
                    if date_col is not None and len(row) > date_col and not pd.isna(row.iloc[date_col]):
                        submission_date = str(row.iloc[date_col]).strip()
                    # Fallback to column 4 if no submission_date column was identified
                    elif len(row) > 4 and not pd.isna(row.iloc[4]):
                        submission_date = str(row.iloc[4]).strip()
                    # Find date field
                    for field_name in extracted_data.keys():
                        if 'application' in field_name.lower() and 'submission' in field_name.lower() and 'date' in field_name.lower():
                            extracted_data[field_name].append(submission_date)
                            break
                    
                    # Extract Region
                    region_col = column_mapping.get('region')  # This is for tables with explicit "Region" column
                    location_col = column_mapping.get('location')  # This is for tables with "Project Location" column
                    region = ""
                    
                    # First try the region column (for tables with explicit "Region" column)
                    if region_col is not None and len(row) > region_col and not pd.isna(row.iloc[region_col]):
                        region = str(row.iloc[region_col]).strip()
                    
                    # Then try the location column (for tables with "Project Location" column)
                    if not region and location_col is not None and len(row) > location_col and not pd.isna(row.iloc[location_col]):
                        region = str(row.iloc[location_col]).strip()
                    
                    # Fallback to column 3 if no region/location column was identified
                    if not region and len(row) > 3 and not pd.isna(row.iloc[3]):
                        region = str(row.iloc[3]).strip()
                    
                    # Find region field
                    for field_name in extracted_data.keys():
                        if 'region' in field_name.lower() and field_name.lower() != 'integration':
                            extracted_data[field_name].append(region)
                            break
                    
                    # Extract State (ending part of Project Location)
                    state = ""
                    if ',' in region:
                        state = region.split(',')[-1].strip()
                    else:
                        state = region
                    
                    # Check if the region contains any Rajasthan district
                    rajasthan_districts = [
                        "Ajmer", "Alwar", "Banswara", "Baran", "Barmer", "Bharatpur", "Bhilwara", 
                        "Bikaner", "Bundi", "Chittorgarh", "Churu", "Dausa", "Dholpur", "Dungarpur", 
                        "Hanumangarh", "Jaipur", "Jaisalmer", "Jalore", "Jhalawar", "Jhunjhunu", 
                        "Jodhpur", "Karauli", "Kota", "Nagaur", "Pali", "Pratapgarh", "Rajsamand", 
                        "Sawai Madhopur", "Sikar", "Sirohi", "Sri Ganganagar", "Tonk", "Udaipur", 
                        "Balotra", "Beawar", "Kotputli-Behror", "Deeg/Didwana-Kuchaman", 
                        "Khairthal-Tijara", "Phalodi", "Salumbar"
                    ]
                    
                    # If any Rajasthan district is found in the region, set state to Rajasthan
                    for district in rajasthan_districts:
                        if district.lower() in region.lower():
                            state = "Rajasthan"
                            break
                    
                    # Find state field
                    for field_name in extracted_data.keys():
                        if 'state' in field_name.lower():
                            extracted_data[field_name].append(state)
                            break
                    
                    # Extract Nature of Applicant
                    nature_col = column_mapping.get('nature_applicant')
                    nature_applicant = ""
                    if nature_col is not None and len(row) > nature_col and not pd.isna(row.iloc[nature_col]):
                        nature_applicant = str(row.iloc[nature_col]).strip()
                    # Fallback to column 5 if no nature_applicant column was identified
                    elif len(row) > 5 and not pd.isna(row.iloc[5]):
                        nature_applicant = str(row.iloc[5]).strip()
                    # Find nature field
                    for field_name in extracted_data.keys():
                        if 'nature' in field_name.lower() and 'applicant' in field_name.lower():
                            extracted_data[field_name].append(nature_applicant)
                            break
                    
                    # Extract Application Quantum (MW)
                    quantum_col = column_mapping.get('quantum')
                    quantum = ""
                    if quantum_col is not None and len(row) > quantum_col and not pd.isna(row.iloc[quantum_col]):
                        quantum = str(row.iloc[quantum_col]).strip()
                    # Fallback to column 7 if no quantum column was identified
                    elif len(row) > 7 and not pd.isna(row.iloc[7]):
                        quantum = str(row.iloc[7]).strip()
                    
                    # Process quantum value to separate the initial number from parenthetical content
                    import re
                    app_quantum_value = quantum
                    granted_quantum_value = ""
                    
                    # First, normalize newlines and extra whitespace
                    normalized = re.sub(r'\s+', ' ', quantum.strip())
                    
                    # Pattern to match "NUMBER (text)" format
                    # This will extract the initial number and the parenthetical content
                    match = re.search(r'^([.\d]+(?:\s*[.\d]+)*)\s*\(([^)]+)\).*$', normalized)
                    if match:
                        app_quantum_value = match.group(1).strip()  # Initial number part
                        granted_quantum_value = "(" + match.group(2).strip() + ")"  # Parenthetical part with parentheses
                    else:
                        # If no parentheses, keep the original value in app_quantum and leave granted_quantum empty
                        app_quantum_value = quantum
                    
                    # Find application quantum field
                    for field_name in extracted_data.keys():
                        if 'quantum' in field_name.lower() and 'mw' in field_name.lower() and 'application' in field_name.lower():
                            extracted_data[field_name].append(app_quantum_value)
                            break
                    
                    # Find granted quantum field
                    for field_name in extracted_data.keys():
                        if 'granted' in field_name.lower() and 'quantum' in field_name.lower() and 'gna/lta' in field_name.lower():
                            extracted_data[field_name].append(granted_quantum_value)
                            break
                    # Extract Start Date of Connectivity
                    start_date_col = column_mapping.get('start_date')
                    start_date = ""
                    if start_date_col is not None and len(row) > start_date_col and not pd.isna(row.iloc[start_date_col]):
                        start_date = str(row.iloc[start_date_col]).strip()
                    # Fallback to column 8 if no start_date column was identified
                    elif len(row) > 8 and not pd.isna(row.iloc[8]):
                        start_date = str(row.iloc[8]).strip()
                    # Find start date field
                    for field_name in extracted_data.keys():
                        if 'start' in field_name.lower() and 'connectivity' in field_name.lower():
                            extracted_data[field_name].append(start_date)
                            break
                    
                    # Extract Criterion/Mode
                    criterion_col = column_mapping.get('criterion')
                    criterion = ""
                    if criterion_col is not None and len(row) > criterion_col and not pd.isna(row.iloc[criterion_col]):
                        criterion = str(row.iloc[criterion_col]).strip()
                        logger.debug(f"Extracted criterion data: {criterion} from column {criterion_col}")
                    # Fallback: Check columns 6-7 for criterion data if not found in mapped column
                    if not criterion:
                        # Check columns 6-7 for criterion data as fallback
                        for col_idx in range(6, min(8, len(row))):
                            if not pd.isna(row.iloc[col_idx]):
                                cell_value = str(row.iloc[col_idx]).strip()
                                # Simple heuristic: if it contains typical criterion values
                                if 'land' in cell_value.lower() or 'route' in cell_value.lower() or 'bg' in cell_value.lower():
                                    criterion = cell_value
                                    logger.debug(f"Fallback: Found criterion data: {criterion} from column {col_idx}")
                                    break
                    # Find criterion field (Mode(Criteria for applying))
                    for field_name in extracted_data.keys():
                        if 'mode' in field_name.lower() and 'criteria' in field_name.lower() and 'applying' in field_name.lower():
                            extracted_data[field_name].append(criterion)
                            logger.debug(f"Assigned criterion '{criterion}' to field '{field_name}'")
                            break
                    
                    # Extract Applicant/Name of Developers
                    applicant_col = column_mapping.get('applicant')
                    applicant = ""
                    if applicant_col is not None and len(row) > applicant_col and not pd.isna(row.iloc[applicant_col]):
                        applicant = str(row.iloc[applicant_col]).strip()
                    # Fallback: Check column 3 (Applicant column in consolidated file) if no applicant column was identified
                    elif len(row) > 2 and not pd.isna(row.iloc[2]):  # Column 3 in 0-indexed is row.iloc[2]
                        applicant = str(row.iloc[2]).strip()
                    # Find Name of Developers field
                    for field_name in extracted_data.keys():
                        if 'name' in field_name.lower() and 'developer' in field_name.lower():
                            extracted_data[field_name].append(applicant)
                            logger.debug(f"Assigned applicant '{applicant}' to field '{field_name}'")
                            break
                    
                else:
                    # Fallback to the original fixed column structure for backward compatibility
                    # Extract Region (Project Location) - column 3 (0-indexed)
                    region = str(row.iloc[3]).strip() if len(row) > 3 and not pd.isna(row.iloc[3]) else ""
                    # Find region field
                    for field_name in extracted_data.keys():
                        if 'region' in field_name.lower() and field_name.lower() != 'integration':
                            extracted_data[field_name].append(region)
                            break
                    
                    # Extract State (ending part of Project Location)
                    state = ""
                    if ',' in region:
                        state = region.split(',')[-1].strip()
                    else:
                        state = region
                    
                    # Check if the region contains any Rajasthan district
                    rajasthan_districts = [
                        "Ajmer", "Alwar", "Banswara", "Baran", "Barmer", "Bharatpur", "Bhilwara", 
                        "Bikaner", "Bundi", "Chittorgarh", "Churu", "Dausa", "Dholpur", "Dungarpur", 
                        "Hanumangarh", "Jaipur", "Jaisalmer", "Jalore", "Jhalawar", "Jhunjhunu", 
                        "Jodhpur", "Karauli", "Kota", "Nagaur", "Pali", "Pratapgarh", "Rajsamand", 
                        "Sawai Madhopur", "Sikar", "Sirohi", "Sri Ganganagar", "Tonk", "Udaipur", 
                        "Balotra", "Beawar", "Kotputli-Behror", "Deeg/Didwana-Kuchaman", 
                        "Khairthal-Tijara", "Phalodi", "Salumbar"
                    ]
                    
                    # If any Rajasthan district is found in the region, set state to Rajasthan
                    for district in rajasthan_districts:
                        if district.lower() in region.lower():
                            state = "Rajasthan"
                            break
                    
                    # Find state field
                    for field_name in extracted_data.keys():
                        if 'state' in field_name.lower():
                            extracted_data[field_name].append(state)
                            break                    
                    # Extract Application ID - column 1 (0-indexed)
                    app_id = str(row.iloc[1]).strip() if len(row) > 1 and not pd.isna(row.iloc[1]) else ""
                    
                    # Determine if it's GNA/ST II or LTA based on whether it's numeric
                    if app_id.isdigit():
                        # Find GNA field
                        for field_name in extracted_data.keys():
                            if 'gna' in field_name.lower() and 'application' in field_name.lower() and 'id' in field_name.lower():
                                extracted_data[field_name].append(app_id)
                                break
                        # Find LTA field and add empty value
                        for field_name in extracted_data.keys():
                            if 'lta' in field_name.lower() and 'application' in field_name.lower() and 'id' in field_name.lower():
                                extracted_data[field_name].append("")
                                break
                    else:
                        # Find LTA field
                        for field_name in extracted_data.keys():
                            if 'lta' in field_name.lower() and 'application' in field_name.lower() and 'id' in field_name.lower():
                                extracted_data[field_name].append(app_id)
                                break
                        # Find GNA field and add empty value
                        for field_name in extracted_data.keys():
                            if 'gna' in field_name.lower() and 'application' in field_name.lower() and 'id' in field_name.lower():
                                extracted_data[field_name].append("")
                                break
                    
                    # Extract Application/Submission Date - column 4 (0-indexed)
                    submission_date = str(row.iloc[4]).strip() if len(row) > 4 and not pd.isna(row.iloc[4]) else ""
                    # Find date field
                    for field_name in extracted_data.keys():
                        if 'application' in field_name.lower() and 'submission' in field_name.lower() and 'date' in field_name.lower():
                            extracted_data[field_name].append(submission_date)
                            break
                    
                    # Extract Nature of Applicant - column 5 (0-indexed)
                    nature_applicant = str(row.iloc[5]).strip() if len(row) > 5 and not pd.isna(row.iloc[5]) else ""
                    # Find nature field
                    for field_name in extracted_data.keys():
                        if 'nature' in field_name.lower() and 'applicant' in field_name.lower():
                            extracted_data[field_name].append(nature_applicant)
                            break
                    
                    # Extract Application Quantum (MW) - column 7 (0-indexed)
                    quantum = str(row.iloc[7]).strip() if len(row) > 7 and not pd.isna(row.iloc[7]) else ""
                    # Find quantum field
                    for field_name in extracted_data.keys():
                        if 'quantum' in field_name.lower() and 'mw' in field_name.lower() and 'application' in field_name.lower():
                            extracted_data[field_name].append(quantum)
                            break
                    
                    # Extract Granted Quantum from Application Quantum
                    granted_quantum = ""
                    # Use the same logic as the main section for consistency
                    app_quantum_value = quantum
                    granted_quantum_value = ""
                    
                    # First, normalize newlines and extra whitespace
                    normalized = re.sub(r'\s+', ' ', quantum.strip())
                    
                    # Pattern to match "NUMBER (text)" format
                    # This will extract the initial number and the parenthetical content
                    match = re.search(r'^([.\d]+(?:\s*[.\d]+)*)\s*$$([^)]+)$$.*$', normalized)
                    if match:
                        app_quantum_value = match.group(1).strip()  # Initial number part
                        granted_quantum_value = "(" + match.group(2).strip() + ")"  # Parenthetical part with parentheses
                    else:
                        # If no parentheses, keep the original value in app_quantum and leave granted_quantum empty
                        app_quantum_value = quantum
                        granted_quantum_value = ""
                    
                    # Find application quantum field
                    for field_name in extracted_data.keys():
                        if 'quantum' in field_name.lower() and 'mw' in field_name.lower() and 'application' in field_name.lower():
                            extracted_data[field_name].append(app_quantum_value)
                            break
                    
                    # Find granted quantum field
                    for field_name in extracted_data.keys():
                        if 'granted' in field_name.lower() and 'quantum' in field_name.lower() and 'gna/lta' in field_name.lower():
                            extracted_data[field_name].append(granted_quantum_value)
                            break
                    # Extract Start Date of Connectivity - column 8 (0-indexed)
                    start_date = str(row.iloc[8]).strip() if len(row) > 8 and not pd.isna(row.iloc[8]) else ""
                    # Find start date field
                    for field_name in extracted_data.keys():
                        if 'start' in field_name.lower() and 'connectivity' in field_name.lower():
                            extracted_data[field_name].append(start_date)
                            break
                    
                    # Extract Criterion/Mode - try to find in nearby columns
                    criterion = ""
                    # Check columns 6-7 for criterion data as fallback
                    for col_idx in range(6, min(8, len(row))):
                        if not pd.isna(row.iloc[col_idx]):
                            cell_value = str(row.iloc[col_idx]).strip()
                            # Simple heuristic: if it contains typical criterion values
                            if 'land' in cell_value.lower() or 'route' in cell_value.lower() or 'bg' in cell_value.lower():
                                criterion = cell_value
                                logger.debug(f"Fallback: Found criterion data: {criterion} from column {col_idx}")
                                break
                    # Find criterion field (Mode(Criteria for applying))
                    for field_name in extracted_data.keys():
                        if 'mode' in field_name.lower() and 'criteria' in field_name.lower() and 'applying' in field_name.lower():
                            extracted_data[field_name].append(criterion)
                            logger.debug(f"Assigned criterion '{criterion}' to field '{field_name}' (fallback)")
                            break
                    
                    # Extract Applicant/Name of Developers (fallback)
                    applicant = ""
                    # Fallback: Check column 3 (Applicant column in consolidated file)
                    if len(row) > 2 and not pd.isna(row.iloc[2]):  # Column 3 in 0-indexed is row.iloc[2]
                        applicant = str(row.iloc[2]).strip()
                    # Find Name of Developers field
                    for field_name in extracted_data.keys():
                        if 'name' in field_name.lower() and 'developer' in field_name.lower():
                            extracted_data[field_name].append(applicant)
                            logger.debug(f"Assigned applicant '{applicant}' to field '{field_name}' (fallback)")
                            break
                
                # Extract Status of application (check the next few rows for status keywords)
                status = ""
                # Look in the next few rows for status information
                for j in range(1, min(8, len(df) - i)):  # Check next 7 rows (increased from 5)
                    next_row = df.iloc[i + j]
                    next_first_cell = str(next_row.iloc[0]).strip() if not pd.isna(next_row.iloc[0]) else ""
                    
                    # If we find a row with status keywords
                    if any(keyword in next_first_cell.lower() for keyword in ['withdrawn', 'granted', 'revoked', 'closed', 'agreed']):
                        if 'withdrawn' in next_first_cell.lower() or 'closed' in next_first_cell.lower():
                            status = "Withdrawn"
                        elif 'granted' in next_first_cell.lower() or 'agreed' in next_first_cell.lower():
                            status = "Granted"
                        elif 'revoked' in next_first_cell.lower():
                            status = "Revoked"
                        break
                
                # Find status field
                for field_name in extracted_data.keys():
                    if 'status' in field_name.lower() and 'application' in field_name.lower():
                        extracted_data[field_name].append(status)
                        break
                
                logger.debug(f"Extracted data for serial {sr_no}: Status = '{status}'")
                
            i += 1
        
        # Create the output file with exact template formatting
        output_filename = os.path.splitext(excel_filepath)[0] + '_required_fields.xlsx'
        
        # Create a new workbook with proper template structure
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        new_sheet.title = "Data to be captured"
        
        # Create structure with proper formatting
        max_row = max(100, template_sheet.max_row)  # Ensure we have enough rows
        max_col = template_sheet.max_column
        
        # Copy the header row formatting from template
        for col_num in range(1, max_col + 1):
            template_cell = template_sheet.cell(row=2, column=col_num)  # Header row in template
            new_cell = new_sheet.cell(row=2, column=col_num)
            
            # Copy header values and formatting
            new_cell.value = template_cell.value
            
            # Apply header formatting (copy exact formatting from template)
            new_cell.font = openpyxl.styles.Font(
                name=template_cell.font.name,
                sz=template_cell.font.sz,
                b=template_cell.font.b
            )
            
            # Apply exact fill pattern from template
            if template_cell.fill.patternType == 'solid':
                new_cell.fill = openpyxl.styles.PatternFill(
                    patternType='solid',
                    start_color=template_cell.fill.fgColor,
                    end_color=template_cell.fill.fgColor
                )
            
            # Apply alignment
            new_cell.alignment = openpyxl.styles.Alignment(
                horizontal=template_cell.alignment.horizontal,
                vertical=template_cell.alignment.vertical,
                wrap_text=template_cell.alignment.wrapText
            )
            
            # Apply borders
            new_cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin') if template_cell.border.left.style else openpyxl.styles.Side(style=None),
                right=openpyxl.styles.Side(style='thin') if template_cell.border.right.style else openpyxl.styles.Side(style=None),
                top=openpyxl.styles.Side(style='thin') if template_cell.border.top.style else openpyxl.styles.Side(style=None),
                bottom=openpyxl.styles.Side(style='thin') if template_cell.border.bottom.style else openpyxl.styles.Side(style=None)
            )
        
        # Apply data cell formatting to all other rows
        for row_num in range(3, max_row + 1):
            for col_num in range(1, max_col + 1):
                new_cell = new_sheet.cell(row=row_num, column=col_num)
                # Apply data cell formatting (matching template row 4 formatting)
                new_cell.font = openpyxl.styles.Font(name='Adani Regular', size=10, bold=False)
                new_cell.fill = openpyxl.styles.PatternFill(fill_type=None)
                new_cell.alignment = openpyxl.styles.Alignment(wrap_text=False)
                new_cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style=None),
                    right=openpyxl.styles.Side(style=None),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )        
        # Map the extracted data to the template columns
        old_headers = list(extracted_data.keys())  # Use our extracted data keys as "old headers"
        
        # Define mapping between our extracted fields and template headers
        header_mapping = {
            'Sr.no': 'Sr.no.',
            'Region': 'Region',
            'State': 'State',
            'GNA/ST II Application ID': 'GNA/ST II Application ID',
            'LTA Application ID': 'LTA Application ID',
            'Application/Submission Date': 'Application/Submission Date',
            'Nature of Applicant': 'Nature of Applicant',
            'Application Quantum (MW)(ST II)': 'Application Quantum (MW)(ST II)',
            'Applied Start of Connectivity sought by developer date( start date of connectivity as per the application)': 'Applied Start of Connectivity sought by developer date( start date of connectivity as per the application)',
            'Status of application(Withdrawn / granted. Revoked.)': 'Status of application(Withdrawn / granted. Revoked.)',
            'Mode(Criteria for applying)': 'Mode(Criteria for applying)'
        }
        
        # Find the positions of the template headers that match our extracted fields
        header_positions = {}
        for old_header, new_header in header_mapping.items():
            if new_header in template_headers:
                header_positions[old_header] = template_headers.index(new_header) + 1  # 1-indexed
        
        # Also map direct matches
        for field_name in extracted_data.keys():
            if field_name in template_headers:
                header_positions[field_name] = template_headers.index(field_name) + 1
        
        # Write the data rows starting from row 3
        data_start_row = 3
        
        # Get the number of data rows (assuming all lists have the same length)
        data_row_count = len(next(iter(extracted_data.values()))) if extracted_data else 0
        
        # Write each data row with proper data formatting
        for row_idx in range(data_row_count):
            target_row = data_start_row + row_idx
            for field_name, values in extracted_data.items():
                if field_name in header_positions and row_idx < len(values):
                    col_idx = header_positions[field_name]
                    cell_value = values[row_idx]
                    new_sheet.cell(row=target_row, column=col_idx).value = cell_value
                    
                    # Ensure data cell formatting (matching template row 4 formatting)
                    cell = new_sheet.cell(row=target_row, column=col_idx)
                    cell.font = openpyxl.styles.Font(name='Adani Regular', size=10, bold=False)
                    cell.fill = openpyxl.styles.PatternFill(fill_type=None)
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=False)
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style=None),
                        right=openpyxl.styles.Side(style=None),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )
        
        # Copy column widths from template
        for col_num in range(1, min(template_sheet.max_column + 1, new_sheet.max_column + 1)):
            column_letter = openpyxl.utils.get_column_letter(col_num)
            if column_letter in template_sheet.column_dimensions:
                new_sheet.column_dimensions[column_letter].width = template_sheet.column_dimensions[column_letter].width
        
        # Save the new workbook
        new_wb.save(output_filename)
        logger.success(f"Required fields extracted and saved to: {output_filename}")
        
        # Also return a DataFrame for compatibility
        # Create a DataFrame with the extracted data using template column order
        # First, we need to ensure all arrays have the same length
        if extracted_data:
            # Find the maximum length among all arrays
            max_length = max(len(values) for values in extracted_data.values()) if extracted_data else 0
            
            # Pad shorter arrays with empty strings to match max_length
            for key in extracted_data:
                current_length = len(extracted_data[key])
                if current_length < max_length:
                    extracted_data[key].extend([""] * (max_length - current_length))
        
        ordered_data = {}
        for header in template_headers:
            if header and header in extracted_data:
                ordered_data[header] = extracted_data[header]
        
        # Add any remaining fields that weren't in the template
        for key, value in extracted_data.items():
            if key not in ordered_data:
                ordered_data[key] = value
        
        return pd.DataFrame(ordered_data)
        
    except Exception as e:
        logger.error(f"Failed to extract required fields from Excel: {str(e)}")
        return None

