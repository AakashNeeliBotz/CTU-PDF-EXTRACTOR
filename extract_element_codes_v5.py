"""
Final corrected extraction script with proper Element Status column mapping
and improved Annexure parsing.
"""
import pandas as pd
import fitz
import re
from openpyxl import load_workbook
from difflib import SequenceMatcher
import warnings
warnings.filterwarnings('ignore')

EXCEL_PATH = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx"
PDF_33 = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\downloaded_pdfs\SN1\172381548953Minutes of 33rd CMETS NR meeting held on 05.08.2024.pdf"
PDF_34 = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\downloaded_pdfs\SN1\172838877090Minutes of meeting 34th CMETS NR Meeting held on 20-9-24.pdf"

COL_APP_ID = 8
COL_CTS = 39
COL_ATS = 40
COL_DTL = 41
DATA_START_ROW = 3

def extract_pdf_text(pdf_path):
    doc = fitz.open(pdf_path)
    text = ""
    for page in doc:
        text += page.get_text("text")
    doc.close()
    return text

def normalize_text(text):
    if not text:
        return ""
    text = str(text).lower()
    text = ' '.join(text.split())
    text = text.replace('–', '-').replace('—', '-').replace('−', '-')
    text = re.sub(r'[^\w\s\-/]', ' ', text)
    return text.strip()

def similarity_score(s1, s2):
    return SequenceMatcher(None, s1, s2).ratio()

def load_element_status():
    """Load Element Status with correct column mapping: Col 1 = Code, Col 4 = Description"""
    print("[*] Loading Element Status sheet...")
    df = pd.read_excel(EXCEL_PATH, sheet_name='Element Status', header=None)
    
    element_map = {}  # normalized_description -> code
    code_to_desc = {} # code -> original_description
    
    for i in range(len(df)):
        # Element Code is in column 1
        code_cell = df.iloc[i, 1] if len(df.columns) > 1 else None
        # Description (Transmission Scope) is in column 4
        desc_cell = df.iloc[i, 4] if len(df.columns) > 4 else None
        
        if pd.notna(code_cell):
            code_str = str(code_cell).strip()
            if re.match(r'^EL-[A-Z0-9]{5}$', code_str):
                if pd.notna(desc_cell) and len(str(desc_cell).strip()) > 10:
                    desc_str = str(desc_cell).strip()
                    norm_desc = normalize_text(desc_str)
                    element_map[norm_desc] = code_str
                    code_to_desc[code_str] = desc_str
    
    print(f"[+] Loaded {len(code_to_desc)} Element Codes")
    return element_map, code_to_desc

def find_element_codes(description, element_map, threshold=0.50):
    """Find matching Element Codes for a description"""
    if not description or len(description) < 15:
        return []
    
    desc_norm = normalize_text(description)
    matches = []
    
    # Exact match
    if desc_norm in element_map:
        return [element_map[desc_norm]]
    
    # Keyword-based matching
    for key, code in element_map.items():
        if len(key) > 15:
            # Extract key terms
            key_terms = set(re.findall(r'\b\w{4,}\b', key))
            desc_terms = set(re.findall(r'\b\w{4,}\b', desc_norm))
            
            if key_terms and desc_terms:
                overlap = len(key_terms & desc_terms) / max(len(key_terms), len(desc_terms))
                if overlap > 0.4:
                    matches.append(code)
    
    if matches:
        return list(set(matches))[:5]
    
    # Fuzzy matching
    candidates = []
    for key, code in element_map.items():
        if len(key) > 15:
            score = similarity_score(desc_norm[:100], key[:100])
            if score >= threshold:
                candidates.append((score, code))
    
    candidates.sort(reverse=True)
    return [code for score, code in candidates[:5]]

def extract_annexures(pdf_text):
    """
    Extract Annexure sections by looking for 'Annexure-XX' headers
    followed by numbered lists.
    """
    annexures = {}
    
    # Find each Annexure section
    # Pattern: "Annexure-XX" followed by content until next "Annexure-" or end
    pattern = r'Annexure-([IVX]+)\s*\n*([\s\S]*?)(?=\nAnnexure-[IVX]+\s*\n|Minutes of \d+\w* Consultation|$)'
    matches = re.findall(pattern, pdf_text, re.IGNORECASE)
    
    for num, content in matches:
        # Clean the content
        content = re.sub(r'Minutes of \d+.*?Page \d+ of \d+', '', content, flags=re.DOTALL)
        content = re.sub(r'Transmission system for Connectivity.*?:\s*\n*', '', content, flags=re.IGNORECASE)
        content = re.sub(r'For connectivity at.*?:\s*\n*', '', content, flags=re.IGNORECASE)
        
        # Extract numbered items (1. item, 2. item, etc.)
        items = []
        
        # Pattern: "1." or "1)" at start of line followed by content
        item_pattern = r'(?:^|\n)\s*(\d+)\.\s*\n*(.+?)(?=\n\s*\d+\.\s|\n\s*Additional|$)'
        item_matches = re.findall(item_pattern, content, re.DOTALL)
        
        for item_num, item_text in item_matches:
            cleaned = ' '.join(item_text.split())
            # Filter out non-transmission items
            if len(cleaned) > 20 and any(kw in cleaned.lower() for kw in ['kv', 'mva', 'line', 'substation', 'ps', 'lilo', 'hvdc', 'establishment', 'augmentation']):
                items.append(cleaned)
        
        # If no numbered items found, try bullet points
        if not items:
            for bullet in re.findall(r'[•]\s*(.+?)(?=[•]|\n\n|$)', content, re.DOTALL):
                cleaned = ' '.join(bullet.split())
                if len(cleaned) > 20:
                    items.append(cleaned)
        
        if items:
            key = num.upper()
            annexures[key] = items
            print(f"    Annexure-{key}: {len(items)} items")
    
    return annexures

def extract_app_transmission_data(pdf_text, annexures):
    """Extract ATS, DTL, CTS data for each Application ID"""
    results = {}
    
    # Split by "Details of Transmission system for Connectivity under GNA:"
    sections = re.split(
        r'Details of Transmission system for Connectivity under GNA:\s*',
        pdf_text,
        flags=re.IGNORECASE
    )
    
    for i, section in enumerate(sections[1:], 1):
        # Get context before this section
        prefix = sections[i-1][-3500:] if len(sections[i-1]) > 3500 else sections[i-1]
        
        # Find App IDs in context - prioritize recent ones
        app_ids = list(re.findall(r'\b22\d{8}\b', prefix))
        app_ids += list(re.findall(r'\b22\d{8}\b', section[:2000]))
        
        if not app_ids:
            continue
        
        # Parse ATS (Section A)
        ats_text = ""
        ats_match = re.search(
            r'A\.\s*Associated\s*Transmission\s*System\s*\(ATS\)[:\s]*(.+?)(?=B\.\s*Transmission|$)',
            section[:2000], re.DOTALL | re.IGNORECASE
        )
        if ats_match:
            ats_raw = ' '.join(ats_match.group(1).split())
            ats_text = "NIL" if 'NIL' in ats_raw.upper() else ats_raw[:500]
        
        # Parse DTL (Section B)
        dtl_text = ""
        dtl_match = re.search(
            r'B\.\s*Transmission\s*System\s*under\s*applicant\s*scope[:\s]*(.+?)(?=C\.\s*Transmission|$)',
            section[:2500], re.DOTALL | re.IGNORECASE
        )
        if dtl_match:
            dtl_raw = dtl_match.group(1)
            # Look for (i). item patterns
            items = re.findall(r'\([ivx]+\)\.*\s*(.+?)(?=\([ivx]+\)|C\.|Minutes|Sl\.|$)',
                             dtl_raw, re.DOTALL | re.IGNORECASE)
            if items:
                dtl_text = ' '.join([' '.join(it.split()) for it in items])[:600]
            else:
                dtl_text = ' '.join(dtl_raw.split())[:600]
        
        # Parse CTS (Section C)
        cts_text = ""
        cts_items = []
        cts_match = re.search(
            r'C\.\s*Transmission\s*system\s*for\s*Connectivity\s*under\s*GNA[:\s]*(.+?)(?=Start Date|Sl\.|Minutes|$)',
            section[:3000], re.DOTALL | re.IGNORECASE
        )
        if cts_match:
            cts_raw = cts_match.group(1).strip()
            
            # Check for Annexure reference
            annex_ref = re.search(r'Annexure[- ]?([IVX]+)', cts_raw, re.IGNORECASE)
            if annex_ref:
                annex_key = annex_ref.group(1).upper()
                if annex_key in annexures:
                    cts_items = annexures[annex_key]
                cts_text = f"ANNEXURE:{annex_key}"
            else:
                cts_text = ' '.join(cts_raw.split())[:500]
        
        # Associate with recent App IDs (last 5 are most relevant)
        relevant_ids = list(set(app_ids[-6:]))
        
        for app_id in relevant_ids:
            if app_id not in results:
                results[app_id] = {'ATS': '', 'DTL': '', 'CTS': '', 'CTS_items': []}
            
            if ats_text and not results[app_id]['ATS']:
                results[app_id]['ATS'] = ats_text
            if dtl_text and not results[app_id]['DTL']:
                results[app_id]['DTL'] = dtl_text
            if cts_text and not results[app_id]['CTS']:
                results[app_id]['CTS'] = cts_text
            if cts_items and not results[app_id]['CTS_items']:
                results[app_id]['CTS_items'] = cts_items
    
    return results

def update_excel(pdf_data, element_map, code_to_desc):
    """Update Excel with Element Codes"""
    print("\n[*] Updating Excel...")
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Data to be captured']
    
    updates = {'CTS': 0, 'ATS': 0, 'DTL': 0}
    seen = set()
    
    for row in range(DATA_START_ROW, ws.max_row + 1):
        cell = ws.cell(row=row, column=COL_APP_ID + 1).value
        if not cell:
            continue
        
        app_id = re.sub(r'\D', '', str(cell).split()[0])[:10]
        if not app_id or len(app_id) != 10 or app_id in seen:
            continue
        seen.add(app_id)
        
        if app_id not in pdf_data:
            continue
        
        data = pdf_data[app_id]
        
        cur_cts = ws.cell(row=row, column=COL_CTS + 1).value
        cur_ats = ws.cell(row=row, column=COL_ATS + 1).value
        cur_dtl = ws.cell(row=row, column=COL_DTL + 1).value
        
        # Update CTS
        if not cur_cts:
            cts_codes = set()
            
            # Match CTS items to Element Codes
            if data['CTS_items']:
                for item in data['CTS_items']:
                    codes = find_element_codes(item, element_map)
                    cts_codes.update(codes)
            elif data['CTS'] and not data['CTS'].startswith('ANNEXURE:'):
                codes = find_element_codes(data['CTS'], element_map)
                cts_codes.update(codes)
            
            if cts_codes:
                ws.cell(row=row, column=COL_CTS + 1).value = ','.join(sorted(cts_codes))
                updates['CTS'] += 1
                print(f"  Row {row} ({app_id}): CTS = {len(cts_codes)} codes")
        
        # Update ATS (leave blank if NIL)
        if not cur_ats and data['ATS'] and data['ATS'] != "NIL":
            ats_codes = find_element_codes(data['ATS'], element_map)
            if ats_codes:
                ws.cell(row=row, column=COL_ATS + 1).value = ','.join(sorted(set(ats_codes)))
                updates['ATS'] += 1
                print(f"  Row {row} ({app_id}): ATS = {len(ats_codes)} codes")
        
        # Update DTL
        if not cur_dtl and data['DTL']:
            dtl_codes = find_element_codes(data['DTL'], element_map)
            if dtl_codes:
                ws.cell(row=row, column=COL_DTL + 1).value = ','.join(sorted(set(dtl_codes)))
                updates['DTL'] += 1
                print(f"  Row {row} ({app_id}): DTL = {len(dtl_codes)} codes")
    
    print(f"\n[*] Saving...")
    wb.save(EXCEL_PATH)
    print("[+] Saved!")
    
    return updates

def main():
    print("=" * 70)
    print("CTU PDF Extractor - v5 (Corrected Column Mapping)")
    print("=" * 70)
    
    element_map, code_to_desc = load_element_status()
    
    if not element_map:
        print("[!] ERROR: No element codes!")
        return
    
    # Show sample Element descriptions
    print("\nSample Element descriptions (Transmission Scope):")
    for code in list(code_to_desc.keys())[:3]:
        desc = code_to_desc[code]
        print(f"  {code}: {desc[:60]}...")
    
    print("\n[*] Processing PDF 33...")
    text_33 = extract_pdf_text(PDF_33)
    print(f"    Text extracted: {len(text_33)} chars")
    annexures_33 = extract_annexures(text_33)
    print(f"    Total Annexures: {len(annexures_33)}")
    data_33 = extract_app_transmission_data(text_33, annexures_33)
    print(f"    App IDs extracted: {len(data_33)}")
    
    print("\n[*] Processing PDF 34...")
    text_34 = extract_pdf_text(PDF_34)
    print(f"    Text extracted: {len(text_34)} chars")
    annexures_34 = extract_annexures(text_34)
    print(f"    Total Annexures: {len(annexures_34)}")
    data_34 = extract_app_transmission_data(text_34, annexures_34)
    print(f"    App IDs extracted: {len(data_34)}")
    
    # Merge
    all_data = data_33.copy()
    for app_id, data in data_34.items():
        if app_id not in all_data:
            all_data[app_id] = data
        else:
            for key in ['ATS', 'DTL', 'CTS']:
                if data[key] and not all_data[app_id][key]:
                    all_data[app_id][key] = data[key]
            if data['CTS_items'] and not all_data[app_id]['CTS_items']:
                all_data[app_id]['CTS_items'] = data['CTS_items']
    
    print(f"\n[*] Total unique App IDs: {len(all_data)}")
    
    # Show samples
    print("\nSample extracted data:")
    for app_id in list(all_data.keys())[:2]:
        d = all_data[app_id]
        print(f"\n  {app_id}:")
        print(f"    ATS: {d['ATS'][:40]}..." if d['ATS'] else "    ATS: (empty)")
        print(f"    DTL: {d['DTL'][:40]}..." if d['DTL'] else "    DTL: (empty)")
        print(f"    CTS: {d['CTS']}, {len(d['CTS_items'])} items")
        if d['CTS_items']:
            print(f"    First CTS item: {d['CTS_items'][0][:50]}...")
    
    # Try matching for a sample
    print("\nTest matching for first App ID's CTS items:")
    first_app = list(all_data.keys())[0]
    first_data = all_data[first_app]
    for item in first_data['CTS_items'][:3]:
        codes = find_element_codes(item, element_map)
        print(f"  Item: {item[:50]}...")
        print(f"    Matched codes: {codes}")
    
    # Update Excel
    updates = update_excel(all_data, element_map, code_to_desc)
    
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"  CTS updates: {updates['CTS']}")
    print(f"  ATS updates: {updates['ATS']}")
    print(f"  DTL updates: {updates['DTL']}")
    print(f"  TOTAL:       {sum(updates.values())}")
    print("=" * 70)

if __name__ == "__main__":
    main()
