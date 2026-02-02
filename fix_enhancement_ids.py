"""
Fix Enhancement IDs in the Excel file
- Move IDs like "2200000661 (Enh)" from LTA column to GNA/ST II Application ID column
- Keep the "(Enh)" or "(Enhancement)" notation WITH the ID
- Handle line breaks and formatting issues
"""

import openpyxl
import re

FILE_PATH = r'c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx'

def clean_text(text):
    """Clean text by normalizing whitespace and newlines"""
    if not text:
        return text
    text = str(text).replace('\n', '').replace('\r', '').replace(' ', '')
    return text

def extract_id_with_enhancement(text):
    """
    Extract ID and enhancement marker from mixed values
    Examples:
    - "2200000661 (Enh)" -> ("2200000661 (Enh)", "")
    - "2200000564 (Enhancement)" -> ("2200000564 (Enhancement)", "")
    - "2\n200000661 \n(Enh)" -> ("2200000661 (Enh)", "")
    - "2200000772  ACC Limited" -> ("2200000772", "ACC Limited")
    """
    if not text:
        return None, None
    
    # First, normalize the text by removing line breaks within numbers
    text = str(text).strip()
    
    # Remove line breaks to see the full pattern
    normalized = text.replace('\n', '').replace('\r', '')
    
    # Pattern for ID with Enhancement: 10-digit ID followed by (Enh) or (Enhancement)
    pattern_enh = r'(\d{10})\s*\(?(Enh(?:ancement)?)\)?'
    match = re.search(pattern_enh, normalized, re.IGNORECASE)
    if match:
        app_id = match.group(1)
        enh_marker = match.group(2)
        # Standardize the enhancement marker
        if enh_marker.lower() == 'enh':
            full_id = f"{app_id} (Enh)"
        else:
            full_id = f"{app_id} (Enhancement)"
        return full_id, None
    
    # Pattern for ID with name (no enhancement)
    pattern_name = r'^(\d{10})\s*(.*)$'
    match = re.match(pattern_name, normalized, re.DOTALL)
    if match:
        app_id = match.group(1)
        remaining = match.group(2).strip()
        # If remaining contains Enhancement, include it with ID
        if remaining and ('Enh' in remaining or 'Enhancement' in remaining):
            return f"{app_id} ({remaining.strip('() ')})", None
        return app_id, remaining if remaining else None
    
    return None, text

def process_excel():
    print("Loading workbook...")
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb['Data to be captured']
    
    header_row = 2
    data_start_row = 3
    
    COL_DEVELOPER = 7    # G - Name of Developers
    COL_GNA_ID = 9       # I - GNA/ST II Application ID
    COL_LTA_ID = 10      # J - LTA Application ID
    COL_ENHANCEMENT = 11 # K - Application ID under Enhancement
    
    changes = []
    
    print(f"Processing rows from {data_start_row} to {ws.max_row}...")
    
    for row in range(data_start_row, ws.max_row + 1):
        lta_val = ws.cell(row=row, column=COL_LTA_ID).value
        gna_val = ws.cell(row=row, column=COL_GNA_ID).value
        
        if lta_val:
            lta_str = str(lta_val).strip()
            
            # Check if LTA contains an ID (10 digits somewhere)
            normalized = lta_str.replace('\n', '').replace('\r', '').replace(' ', '')
            
            # Check if it has a 10-digit ID pattern
            if re.search(r'\d{10}', normalized):
                extracted_id, remaining = extract_id_with_enhancement(lta_str)
                
                if extracted_id:
                    old_gna = gna_val
                    old_lta = lta_val
                    
                    # Move the ID (with enhancement marker) to GNA column if empty
                    if not gna_val:
                        ws.cell(row=row, column=COL_GNA_ID).value = extracted_id
                        changes.append(f"Row {row}: Moved '{extracted_id}' from LTA to GNA")
                    
                    # Handle remaining (developer name) if any
                    if remaining and remaining not in ['(Enhancement)', 'Enhancement', '(Enh)', 'Enh']:
                        current_dev = ws.cell(row=row, column=COL_DEVELOPER).value
                        if not current_dev:
                            ws.cell(row=row, column=COL_DEVELOPER).value = remaining
                            changes.append(f"Row {row}: Moved developer name '{remaining}' to Developer column")
                    
                    # Clear LTA column (ID moved to GNA)
                    # Only keep if it's just "(Enhancement)" without ID
                    if '(Enhancement)' in lta_str or '(Enh)' in lta_str:
                        # Check if the Enhancement is already captured with the ID
                        if 'Enh' in extracted_id or 'Enhancement' in extracted_id:
                            ws.cell(row=row, column=COL_LTA_ID).value = None
                        else:
                            ws.cell(row=row, column=COL_LTA_ID).value = "(Enhancement)"
                    else:
                        ws.cell(row=row, column=COL_LTA_ID).value = None
    
    print(f"\n=== CHANGES MADE ===")
    for change in changes:
        print(change)
    print(f"\nTotal changes: {len(changes)}")
    
    print(f"\nSaving workbook...")
    wb.save(FILE_PATH)
    print("Done!")
    
    return changes

if __name__ == "__main__":
    process_excel()
