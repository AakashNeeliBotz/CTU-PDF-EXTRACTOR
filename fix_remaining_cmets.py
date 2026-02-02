"""
Fix remaining 32 rows with no Application ID - default to 33 (earlier meeting)
"""
import pandas as pd
from openpyxl import load_workbook
import sys

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

excel_path = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx"

# Load workbook
wb = load_workbook(excel_path)
ws = wb["Data to be captured"]

print("="*100)
print("FIXING REMAINING CMETS ROWS (NO APPLICATION ID) - DEFAULTING TO 33")
print("="*100)

changes_made = []
cmets_col = 12  # Column L
srno_col = 2    # Column B
app_id_col = 9  # Column I

fixed_count = 0
for row in range(3, ws.max_row + 1):
    cmets_val = ws.cell(row=row, column=cmets_col).value
    if cmets_val:
        val_str = str(cmets_val)
        if '33' in val_str and '34' in val_str:
            # This row still has both 33 and 34
            srno = ws.cell(row=row, column=srno_col).value
            app_id = ws.cell(row=row, column=app_id_col).value
            
            # Default to 33 (earlier meeting - typically original grant)
            ws.cell(row=row, column=cmets_col).value = 33
            fixed_count += 1
            print(f"   Row {row} (Sr.No. {srno}, AppID: {app_id}): '{cmets_val}' -> 33")
            changes_made.append(f"Row {row} (Sr.No. {srno}): CMETS changed from '{cmets_val}' to '33'")

print(f"\n   Fixed {fixed_count} remaining rows")

# Save changes
print("\n" + "="*100)
print("SAVING CHANGES")
print("="*100)

wb.save(excel_path)
print(f"Changes saved to: {excel_path}")

# Verify final state
print("\n" + "="*100)
print("FINAL VERIFICATION")
print("="*100)

df = pd.read_excel(excel_path, sheet_name="Data to be captured", header=1)
cmets_col_name = 'CMETS GNA Approved'
count_both = sum(1 for val in df[cmets_col_name] if pd.notna(val) and '33' in str(val) and '34' in str(val))
print(f"\nRemaining rows with both 33 and 34 in CMETS: {count_both}")

if count_both == 0:
    print("\n✓ SUCCESS: All CMETS GNA Approved values have been fixed!")
else:
    print(f"\n⚠ WARNING: {count_both} rows still have both 33 and 34")
