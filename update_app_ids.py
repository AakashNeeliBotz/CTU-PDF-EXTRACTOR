import pandas as pd
import re
from openpyxl import load_workbook

def update_application_ids():
    """Update application IDs to extract only the numeric part"""
    file_path = 'Connectivity_Application_Data_TEST_ALL_SHEETS38.xlsx'
    
    print("Updating Application IDs to extract only numeric part...")
    
    # Load the raw data
    df_raw = pd.read_excel(file_path, sheet_name='Data to be captured', header=None)
    
    print(f"Loaded data with {len(df_raw)} rows")
    
    # Update the application IDs to extract only the numeric part
    updated_count = 0
    
    for i in range(len(df_raw)):
        row = df_raw.iloc[i]
        # Check if this is one of our extracted entries
        cell_value = str(row.values)
        if ('Serentica Renewables' in cell_value or 
            'ACME Heergarh' in cell_value or 
            'AM Green Energy' in cell_value):
            
            current_app_id = row.iloc[10] if pd.notna(row.iloc[10]) else ''
            
            if current_app_id and isinstance(current_app_id, str):
                print(f"Found entry in row {i}, current App ID: {repr(current_app_id)}")
                
                # Check if the application ID contains a newline and date pattern
                if '\n(' in current_app_id:
                    # Extract the numeric part before the newline
                    parts = current_app_id.split('\n')
                    if parts:
                        numeric_part = parts[0]  # Take the first part which should be the numeric ID
                        # Further extract just the digits in case there are non-digit characters
                        digit_match = re.search(r'(\d+)', numeric_part)
                        if digit_match:
                            new_app_id = digit_match.group(1)
                            df_raw.iloc[i, 10] = new_app_id
                            print(f"  Updated: {current_app_id} -> {new_app_id}")
                            updated_count += 1
                        else:
                            print(f"  Could not extract digits from: {numeric_part}")
                else:
                    print(f"  No newline found in: {current_app_id}")
    
    print(f"\nSuccessfully updated {updated_count} Application IDs.")
    
    # Load the workbook and update the 'Data to be captured' sheet
    wb = load_workbook(file_path)
    ws = wb['Data to be captured']
    
    # Clear the worksheet
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
    
    # Write the updated data back to the worksheet
    for r_idx in range(len(df_raw)):
        for c_idx in range(len(df_raw.columns)):
            value = df_raw.iloc[r_idx, c_idx]
            if pd.isna(value):
                value = None  # Convert NaN to None for Excel
            ws.cell(row=r_idx+1, column=c_idx+1, value=value)
    
    # Save the workbook
    wb.save(file_path)
    print(f"Updated file saved: {file_path}")

def main():
    print("Starting Application ID update process...")
    update_application_ids()
    print("Process completed!")

if __name__ == "__main__":
    main()