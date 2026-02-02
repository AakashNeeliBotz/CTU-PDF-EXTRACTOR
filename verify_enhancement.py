"""Verify Enhancement rows after merging"""
import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx')
ws = wb['Data to be captured']
print('=== All Enhancement rows (now merged) ===')
count = 0
for row_idx in range(3, ws.max_row + 1):
    gna_id = str(ws.cell(row=row_idx, column=9).value or '')
    lta_id = str(ws.cell(row=row_idx, column=10).value or '')
    
    if 'Enh' in gna_id or 'Enhancement' in gna_id:
        count += 1
        lta_display = lta_id if lta_id else "(empty)"
        print(f'Row {row_idx}: GNA_ID = {gna_id}, LTA = {lta_display}')

print(f'\nTotal Enhancement rows: {count}')

print('\n=== Check for any remaining Enhancement in LTA column ===')
lta_enh = 0
for row_idx in range(3, ws.max_row + 1):
    lta_id = str(ws.cell(row=row_idx, column=10).value or '')
    if 'Enh' in lta_id or 'Enhancement' in lta_id:
        lta_enh += 1
        print(f'Row {row_idx}: LTA still has Enhancement: {lta_id}')
print(f'LTA Enhancement count: {lta_enh}')
