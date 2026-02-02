"""
Extract "Type" from "Nature of Applicant" brackets and update Type column.

RULES:
- Only modify "Data to be captured" sheet
- Extract text inside brackets: Generator (Solar) -> Solar
- Write to "Type" column (Col 16)
- DO NOT overwrite existing Type values
- DO NOT modify Nature of Applicant column
"""
from openpyxl import load_workbook
import re

EXCEL_PATH = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx"

# Column positions (1-indexed for openpyxl)
COL_TYPE = 16  # "Type" column
COL_NATURE = 36  # "Nature of Applicant" column
DATA_START_ROW = 3  # Data starts at row 3 (after headers)

def extract_type_from_brackets(nature_text):
    """
    Extract the value inside brackets from Nature of Applicant text.
    Examples:
    - Generator (Solar) -> Solar
    - Renewable Power Park developer (Solar) -> Solar
    - Generator (Wind) -> Wind
    - Generator (Hybrid) -> Hybrid
    - Generator (Solar) with ESS -> Solar
    """
    if not nature_text:
        return None
    
    text = str(nature_text).strip()
    
    # Find text inside parentheses
    match = re.search(r'\(([^)]+)\)', text)
    if match:
        extracted = match.group(1).strip()
        return extracted
    
    return None

def main():
    print("=" * 60)
    print("Excel Data Cleaning: Extract Type from Nature of Applicant")
    print("=" * 60)
    
    print(f"\n[*] Loading workbook: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH)
    
    # Only work with "Data to be captured" sheet
    ws = wb['Data to be captured']
    print(f"[*] Working on sheet: 'Data to be captured'")
    print(f"[*] Total rows: {ws.max_row}")
    
    updates = 0
    skipped_existing = 0
    no_brackets = 0
    
    print(f"\n[*] Processing rows {DATA_START_ROW} to {ws.max_row}...")
    
    for row in range(DATA_START_ROW, ws.max_row + 1):
        # Get current Type value
        type_cell = ws.cell(row=row, column=COL_TYPE)
        current_type = type_cell.value
        
        # Skip if Type already has a value (DO NOT overwrite)
        if current_type and str(current_type).strip():
            skipped_existing += 1
            continue
        
        # Get Nature of Applicant value
        nature_cell = ws.cell(row=row, column=COL_NATURE)
        nature_value = nature_cell.value
        
        if not nature_value:
            continue
        
        # Extract type from brackets
        extracted_type = extract_type_from_brackets(nature_value)
        
        if extracted_type:
            # Write to Type column
            type_cell.value = extracted_type
            updates += 1
            
            if updates <= 10:  # Show first 10 updates
                print(f"  Row {row}: '{str(nature_value)[:40]}...' -> Type: '{extracted_type}'")
        else:
            no_brackets += 1
    
    print(f"\n[*] Saving workbook...")
    wb.save(EXCEL_PATH)
    print("[+] Saved successfully!")
    
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"  Type values updated: {updates}")
    print(f"  Skipped (already has value): {skipped_existing}")
    print(f"  No brackets found: {no_brackets}")
    print("=" * 60)
    
    wb.close()

if __name__ == "__main__":
    main()
