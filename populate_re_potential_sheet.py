"""
Populate RE Potential sheet in Excel with extracted SN10a data
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
import os
import shutil

excel_path = 'Connectivity Application Data.xlsx'
mapped_data_csv = 'extraction_output/RE_Potential_SN10a_mapped.csv'

print("="*80)
print("POPULATING RE POTENTIAL SHEET")
print("="*80)

# Create backup
backup_path = 'Connectivity Application Data_backup.xlsx'
if not os.path.exists(backup_path):
    shutil.copy(excel_path, backup_path)
    print(f"✓ Created backup: {backup_path}")
else:
    print(f"✓ Backup already exists: {backup_path}")

# Load mapped data
print(f"\nLoading mapped data from: {mapped_data_csv}")
df_data = pd.read_csv(mapped_data_csv)
print(f"✓ Loaded {len(df_data)} records")

# Load Excel workbook
print(f"\nOpening Excel file: {excel_path}")
wb = openpyxl.load_workbook(excel_path)
sheet = wb['RE Potential']
print(f"✓ Opened sheet: {sheet.title}")

# Get headers from row 2
print("\nReading sheet headers from row 2...")
headers = {}
for col_idx, cell in enumerate(sheet[2], 1):
    if cell.value:
        headers[col_idx] = str(cell.value).strip()
print(f"✓ Found {len(headers)} columns")

# Create column mapping
column_mapping = {}
for col_idx, header in headers.items():
    header_lower = header.lower()
    if 'region' in header_lower:
        column_mapping['region'] = col_idx
    elif 'state' in header_lower:
        column_mapping['state'] = col_idx
    elif 'district' in header_lower:
        column_mapping['district'] = col_idx
    elif 'complex' in header_lower and 'complex' not in column_mapping:
        column_mapping['complex'] = col_idx
    elif 's/s' in header_lower:
        column_mapping['s_s'] = col_idx
    elif 'location' in header_lower or 'village' in header_lower:
        column_mapping['location'] = col_idx
    elif 'solar' in header_lower:
        column_mapping['solar'] = col_idx
    elif 'wind' in header_lower:
        column_mapping['wind'] = col_idx
    elif 'hybrid' in header_lower:
        column_mapping['hybrid'] = col_idx
    elif 'others' in header_lower or 'ctuil' in header_lower:
        column_mapping['others'] = col_idx
    elif 're potential' in header_lower:
        column_mapping['re_potential'] = col_idx
    elif 'installed capacity' in header_lower:
        column_mapping['installed_capacity'] = col_idx
    elif 'u/c' in header_lower or 'granted' in header_lower:
        column_mapping['uc_granted'] = col_idx
    elif 'transmission' in header_lower and 'scheme' in header_lower:
        column_mapping['transmission_scheme'] = col_idx
    elif 'remark' in header_lower:
        column_mapping['remarks'] = col_idx

print(f"\nColumn mapping:")
for field, col_idx in sorted(column_mapping.items(), key=lambda x: x[1]):
    print(f"  {field:20s} -> Column {col_idx} ({headers[col_idx]})")

# Find starting row (first empty row after header)
start_row = 3  # Data starts from row 3
print(f"\nData will be written starting from row {start_row}")

# Write data to sheet
print("\n" + "="*80)
print("WRITING DATA TO SHEET")
print("="*80)

records_written = 0
for idx, record in df_data.iterrows():
    current_row = start_row + idx
    
    # Write each field to corresponding column
    if 'region' in column_mapping and pd.notna(record.get('region')):
        sheet.cell(current_row, column_mapping['region'], record['region'])
    
    if 'state' in column_mapping and pd.notna(record.get('state')):
        sheet.cell(current_row, column_mapping['state'], record['state'])
    
    if 'district' in column_mapping and pd.notna(record.get('district')):
        sheet.cell(current_row, column_mapping['district'], record['district'])
    
    if 'complex' in column_mapping and pd.notna(record.get('complex')):
        sheet.cell(current_row, column_mapping['complex'], record['complex'])
    
    if 's_s' in column_mapping and pd.notna(record.get('s_s')):
        sheet.cell(current_row, column_mapping['s_s'], record['s_s'])
    
    if 'location' in column_mapping and pd.notna(record.get('location_village_tehsil')):
        sheet.cell(current_row, column_mapping['location'], record['location_village_tehsil'])
    
    if 'solar' in column_mapping and pd.notna(record.get('solar')):
        cell = sheet.cell(current_row, column_mapping['solar'], float(record['solar']))
        cell.number_format = '0.00'
    
    if 'wind' in column_mapping and pd.notna(record.get('wind')):
        cell = sheet.cell(current_row, column_mapping['wind'], float(record['wind']))
        cell.number_format = '0.00'
    
    if 'hybrid' in column_mapping and pd.notna(record.get('hybrid')):
        cell = sheet.cell(current_row, column_mapping['hybrid'], float(record['hybrid']))
        cell.number_format = '0.00'
    
    if 'others' in column_mapping and pd.notna(record.get('others_ctuil')):
        cell = sheet.cell(current_row, column_mapping['others'], float(record['others_ctuil']))
        cell.number_format = '0.00'
    
    if 're_potential' in column_mapping and pd.notna(record.get('re_potential_gw')):
        cell = sheet.cell(current_row, column_mapping['re_potential'], float(record['re_potential_gw']))
        cell.number_format = '0.00'
    
    if 'installed_capacity' in column_mapping and pd.notna(record.get('installed_capacity')):
        cell = sheet.cell(current_row, column_mapping['installed_capacity'], float(record['installed_capacity']))
        cell.number_format = '0.00'
    
    if 'uc_granted' in column_mapping and pd.notna(record.get('uc_granted_capacity')):
        cell = sheet.cell(current_row, column_mapping['uc_granted'], float(record['uc_granted_capacity']))
        cell.number_format = '0.00'
    
    if 'transmission_scheme' in column_mapping and pd.notna(record.get('transmission_scheme')):
        sheet.cell(current_row, column_mapping['transmission_scheme'], record['transmission_scheme'])
    
    if 'remarks' in column_mapping and pd.notna(record.get('remarks')):
        sheet.cell(current_row, column_mapping['remarks'], record['remarks'])
    
    records_written += 1
    
    if (idx + 1) % 20 == 0:
        print(f"  Written {idx + 1}/{len(df_data)} records...")

print(f"✓ Written {records_written} records to sheet")

# Save workbook
print("\n" + "="*80)
print("SAVING EXCEL FILE")
print("="*80)
wb.save(excel_path)
print(f"✓ Saved: {excel_path}")

# Verify
wb_verify = openpyxl.load_workbook(excel_path)
sheet_verify = wb_verify['RE Potential']

# Count non-empty rows
non_empty_count = 0
for row in range(3, sheet_verify.max_row + 1):
    # Check if any cell in the row has data
    has_data = False
    for col in range(1, sheet_verify.max_column + 1):
        if sheet_verify.cell(row, col).value:
            has_data = True
            break
    if has_data:
        non_empty_count += 1

print(f"\n✓ Verification: {non_empty_count} rows with data in RE Potential sheet")

# Show sample data
print("\n" + "="*80)
print("SAMPLE DATA FROM SHEET (First 5 rows)")
print("="*80)
for row_idx in range(3, min(8, sheet_verify.max_row + 1)):
    row_data = []
    for col_name, col_idx in sorted(column_mapping.items(), key=lambda x: x[1])[:8]:
        value = sheet_verify.cell(row_idx, col_idx).value
        row_data.append(f"{col_name}: {value}")
    print(f"Row {row_idx}: {', '.join(row_data[:5])}")

wb_verify.close()

print("\n" + "="*80)
print("✅ RE POTENTIAL SHEET POPULATION COMPLETE")
print("="*80)
print(f"\nSummary:")
print(f"  Records written: {records_written}")
print(f"  Excel file: {excel_path}")
print(f"  Backup file: {backup_path}")
