"""
Debug script to understand why more rows aren't being updated
"""
import pandas as pd
from openpyxl import load_workbook
import re

EXCEL_PATH = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx"

COL_APP_ID = 8
COL_CTS = 39
COL_ATS = 40
COL_DTL = 41

def analyze_empty_rows():
    """Analyze rows that are empty and could be updated"""
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Data to be captured']
    
    # Collect statistics
    total_rows = 0
    rows_with_app_id = 0
    rows_with_empty_cts = 0
    rows_with_empty_ats = 0
    rows_with_empty_dtl = 0
    
    empty_rows = []  # Rows that have app_id but empty CTS/ATS/DTL
    
    for row in range(3, ws.max_row + 1):
        app_id = ws.cell(row=row, column=COL_APP_ID + 1).value
        cts = ws.cell(row=row, column=COL_CTS + 1).value
        ats = ws.cell(row=row, column=COL_ATS + 1).value
        dtl = ws.cell(row=row, column=COL_DTL + 1).value
        
        total_rows += 1
        
        if app_id:
            app_id_str = str(app_id).strip()
            app_id_clean = re.sub(r'\D', '', app_id_str.split()[0])[:10]
            
            if len(app_id_clean) == 10:
                rows_with_app_id += 1
                
                if not cts:
                    rows_with_empty_cts += 1
                if not ats:
                    rows_with_empty_ats += 1
                if not dtl:
                    rows_with_empty_dtl += 1
                
                # Track rows that need updating
                if not cts or not ats or not dtl:
                    empty_rows.append({
                        'row': row,
                        'app_id': app_id_clean,
                        'cts': cts,
                        'ats': ats,
                        'dtl': dtl
                    })
    
    print("=" * 60)
    print("EXCEL ANALYSIS")
    print("=" * 60)
    print(f"Total rows in sheet: {total_rows}")
    print(f"Rows with valid Application ID: {rows_with_app_id}")
    print(f"Rows with empty CTS: {rows_with_empty_cts}")
    print(f"Rows with empty ATS: {rows_with_empty_ats}")
    print(f"Rows with empty DTL: {rows_with_empty_dtl}")
    
    print(f"\nFirst 15 rows needing updates:")
    for r in empty_rows[:15]:
        print(f"  Row {r['row']}: AppID={r['app_id']}, CTS={'empty' if not r['cts'] else 'has data'}, ATS={'empty' if not r['ats'] else 'has data'}, DTL={'empty' if not r['dtl'] else 'has data'}")
    
    # Show application IDs that need updates
    print("\nApplication IDs needing CTS update (first 20):")
    cts_needed = [r['app_id'] for r in empty_rows if not r['cts']][:20]
    print(f"  {cts_needed}")
    
    wb.close()
    return empty_rows

def check_matching_issue():
    """Check if the PDF App IDs match the Excel App IDs that need updates"""
    import fitz
    
    PDF_33 = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\downloaded_pdfs\SN1\172381548953Minutes of 33rd CMETS NR meeting held on 05.08.2024.pdf"
    
    # Get App IDs from PDF
    doc = fitz.open(PDF_33)
    pdf_text = ""
    for page in doc:
        pdf_text += page.get_text("text")
    doc.close()
    
    pdf_app_ids = set(re.findall(r'\b22\d{8}\b', pdf_text))
    print(f"\n\nApplication IDs found in PDF 33: {len(pdf_app_ids)}")
    
    # Get Excel App IDs that need updates
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Data to be captured']
    
    excel_app_ids_needing_cts = set()
    for row in range(3, ws.max_row + 1):
        app_id = ws.cell(row=row, column=COL_APP_ID + 1).value
        cts = ws.cell(row=row, column=COL_CTS + 1).value
        
        if app_id and not cts:
            app_id_clean = re.sub(r'\D', '', str(app_id).split()[0])[:10]
            if len(app_id_clean) == 10:
                excel_app_ids_needing_cts.add(app_id_clean)
    
    wb.close()
    
    # Find overlap
    matching = pdf_app_ids & excel_app_ids_needing_cts
    print(f"Excel App IDs needing CTS update: {len(excel_app_ids_needing_cts)}")
    print(f"App IDs in PDF that need Excel update: {len(matching)}")
    print(f"\nMatching App IDs (first 20): {list(matching)[:20]}")
    
    return matching

if __name__ == "__main__":
    analyze_empty_rows()
    check_matching_issue()
