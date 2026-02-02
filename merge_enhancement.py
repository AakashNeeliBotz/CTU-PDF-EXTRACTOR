"""
Merge Enhancement markers with IDs in GNA column
- For rows where GNA has ID and LTA has "(Enhancement)", combine them as "ID (Enhancement)" in GNA
- Clear the LTA column after merging
"""

import openpyxl

FILE_PATH = r'c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx'

def process_excel():
    print("Loading workbook...")
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb['Data to be captured']
    
    data_start_row = 3
    
    COL_GNA_ID = 9       # I - GNA/ST II Application ID
    COL_LTA_ID = 10      # J - LTA Application ID
    
    changes = []
    
    print(f"Processing rows from {data_start_row} to {ws.max_row}...")
    
    for row in range(data_start_row, ws.max_row + 1):
        gna_val = ws.cell(row=row, column=COL_GNA_ID).value
        lta_val = ws.cell(row=row, column=COL_LTA_ID).value
        
        if gna_val and lta_val:
            gna_str = str(gna_val).strip()
            lta_str = str(lta_val).strip()
            
            # Check if LTA contains Enhancement marker and GNA doesn't already have it
            if ('Enhancement' in lta_str or 'Enh' in lta_str) and 'Enh' not in gna_str and 'Enhancement' not in gna_str:
                # Combine ID with Enhancement marker
                # Standardize to "(Enhancement)" format
                new_gna = f"{gna_str} (Enhancement)"
                
                ws.cell(row=row, column=COL_GNA_ID).value = new_gna
                ws.cell(row=row, column=COL_LTA_ID).value = None
                
                changes.append(f"Row {row}: '{gna_str}' + '{lta_str}' -> '{new_gna}'")
    
    print(f"\n=== CHANGES MADE ===")
    for change in changes:
        print(change)
    print(f"\nTotal changes: {len(changes)}")
    
    print(f"\nSaving workbook...")
    wb.save(FILE_PATH)
    print("Done!")
    
    return changes

if __name__ == "__main__":
    process_excel()
