"""
Fix Developer Names Script
Standardizes developer name variations in the "Data to be captured" sheet.
"""

import openpyxl
import re

FILE_PATH = r'c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx'

# Developer name corrections
# Format: (pattern_to_match, replacement)
DEVELOPER_CORRECTIONS = [
    # Ambuja variations
    ("Cements Limited / Ambuja", "Ambuja Cements Limited"),
    ("Ambuja  Cements  Limited", "Ambuja Cements Limited"),  # Extra spaces
    
    # Add more corrections here as needed
]


def normalize_spaces(text):
    """Remove extra whitespace and newlines."""
    if not text:
        return text
    text = str(text).replace('\n', ' ').replace('\r', ' ')
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def fix_developer_name(name):
    """Fix developer name based on known corrections."""
    if not name:
        return name
    
    original = str(name)
    normalized = normalize_spaces(original)
    
    # Apply corrections
    for pattern, replacement in DEVELOPER_CORRECTIONS:
        if normalized == pattern:
            return replacement
    
    # If no direct match, just return normalized version
    return normalized


def main():
    print(f"Loading workbook: {FILE_PATH}")
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb['Data to be captured']
    
    print(f"Processing 'Data to be captured' sheet ({ws.max_row} rows)")
    
    changes = 0
    col_developer = 7
    
    for row in range(3, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_developer)
        original = cell.value
        
        if original:
            fixed = fix_developer_name(original)
            if fixed != original:
                cell.value = fixed
                changes += 1
                if changes <= 15:
                    orig_short = str(original).replace('\n', ' ')[:40]
                    print(f"  Row {row}: '{orig_short}' -> '{fixed[:40]}'")
    
    print(f"\nTotal developer names fixed: {changes}")
    
    # Save
    print(f"\nSaving to: {FILE_PATH}")
    wb.save(FILE_PATH)
    print("Done!")


if __name__ == "__main__":
    main()
