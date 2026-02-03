"""
Fix data placement errors in Excel file.
1. Move dates from "Nature of Applicant" to "Application/Submission Date"
2. Swap values between Column Q and Column AD for rows 156, 157, 158
"""

import openpyxl
import re

FILE_PATH = r'c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx'

# Column indices
COL_Q = 17  # Application Quantum (MW)(ST II)
COL_AD = 30  # Application/Submission Date
COL_NATURE = 36  # Nature of Applicant

# Date pattern at the start of text (dd.mm.yyyy or dd-mm-yyyy or dd/mm/yyyy)
DATE_PATTERN = re.compile(r'^(\d{1,2}[.\-/]\d{1,2}[.\-/]\d{2,4})\s*')


def extract_date_from_nature(text):
    """
    Extract date from beginning of Nature of Applicant text.
    Returns (extracted_date, remaining_text)
    """
    if not text or not isinstance(text, str):
        return None, text
    
    text = str(text).strip()
    match = DATE_PATTERN.match(text)
    
    if match:
        date = match.group(1)
        remaining = text[match.end():].strip()
        return date, remaining
    
    return None, text


def main():
    print(f"Loading workbook: {FILE_PATH}")
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb['Data to be captured']
    
    print(f"Processing 'Data to be captured' sheet ({ws.max_row} rows)")
    print("="*80)
    
    changes = {
        'dates_moved': 0,
        'rows_swapped': 0
    }
    
    # TASK 1: Fix dates in "Nature of Applicant"
    print("\nTASK 1: Moving dates from Nature of Applicant to Application/Submission Date")
    print("-"*80)
    
    for row in range(3, ws.max_row + 1):
        nature_cell = ws.cell(row=row, column=COL_NATURE)
        app_date_cell = ws.cell(row=row, column=COL_AD)
        
        nature_value = nature_cell.value
        app_date_value = app_date_cell.value
        
        if nature_value:
            extracted_date, remaining_text = extract_date_from_nature(nature_value)
            
            if extracted_date:
                # Only move if Application/Submission Date is empty
                if not app_date_value or (isinstance(app_date_value, str) and not app_date_value.strip()):
                    app_date_cell.value = extracted_date
                    nature_cell.value = remaining_text if remaining_text else None
                    changes['dates_moved'] += 1
                    
                    if changes['dates_moved'] <= 10:
                        print(f"  Row {row}: Moved date '{extracted_date}' to App/Submission Date")
                        print(f"           Nature now: '{remaining_text[:50] if remaining_text else '(empty)'}'...")
                else:
                    # Don't overwrite existing date, but still clean the Nature column
                    nature_cell.value = remaining_text if remaining_text else None
                    if changes['dates_moved'] <= 10:
                        print(f"  Row {row}: Removed date from Nature (App Date already has: {app_date_value})")
    
    print(f"\n  Total dates moved: {changes['dates_moved']}")
    
    # TASK 2: Swap Column Q and Column AD for rows 156, 157, 158
    print("\nTASK 2: Swapping Column Q and Column AD for rows 156, 157, 158")
    print("-"*80)
    
    for row in [156, 157, 158]:
        cell_q = ws.cell(row=row, column=COL_Q)
        cell_ad = ws.cell(row=row, column=COL_AD)
        
        old_q = cell_q.value
        old_ad = cell_ad.value
        
        # Swap values
        cell_q.value = old_ad
        cell_ad.value = old_q
        
        changes['rows_swapped'] += 1
        print(f"  Row {row}: Q ({old_q}) <-> AD ({old_ad})")
    
    # Summary
    print("\n" + "="*80)
    print("SUMMARY:")
    print(f"  Dates moved from Nature to App/Submission Date: {changes['dates_moved']}")
    print(f"  Rows with Q/AD swapped: {changes['rows_swapped']}")
    
    # Save
    print(f"\nSaving to: {FILE_PATH}")
    wb.save(FILE_PATH)
    print("Done!")
    print("\nConfirmation: Only 'Data to be captured' sheet was modified. No formatting changes.")


if __name__ == "__main__":
    main()
