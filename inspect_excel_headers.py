import openpyxl

def inspect_excel():
    path = r'c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx'
    wb = openpyxl.load_workbook(path)
    ws = wb['Data to be captured']
    
    print(f"Total columns: {ws.max_column}")
    
    # Print the first few rows to understand structure
    for r in range(1, 4):
        print(f"Row {r}:")
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val:
                print(f"  Col {c}: {val}")

if __name__ == "__main__":
    inspect_excel()
