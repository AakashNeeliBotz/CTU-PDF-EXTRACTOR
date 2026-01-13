import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import re
import os

# --- Configuration ---
SOURCE_FILE = r'c:\ElementStatusv2\Output_Report_TBCB_UC.xlsx'
TARGET_FILE = r'c:\ElementStatusv2\Connectivity Application Data 1.xlsx'
TARGET_SHEET = 'Element Status'
START_ROW_TARGET = 4

# --- Parsing Logic ---
def parse_scope(scope):
    if not isinstance(scope, str):
        return "", "", ""
    
    # Try to find Part
    part_match = re.search(r'Part[\s-]*([A-Z0-9]+)', scope, re.IGNORECASE)
    part = part_match.group(0) if part_match else ""
    
    # Try to find Phase
    # Improved regex for Phase to catch more variants
    phase_match = re.search(r'(Phase|Ph)[\s-]*([IVX0-9]+)', scope, re.IGNORECASE)
    phase = phase_match.group(0) if phase_match else ""
    
    # Try to find AREA
    area = ""
    if "Rajasthan" in scope: area = "Rajasthan"
    elif "Khavda" in scope: area = "Khavda"
    elif "Narela" in scope: area = "Narela"
    elif "Ahmedabad" in scope: area = "Ahmedabad"
    elif "Bhadla" in scope: area = "Bhadla"
    elif "Sikar" in scope: area = "Sikar"
    elif "Lakadia" in scope: area = "Lakadia"
    
    if not area:
        area_match = re.search(r'(?:in|from|at)\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)', scope)
        if area_match:
            area = area_match.group(1)
        else:
            area_match = re.search(r'^([A-Z][A-Za-z0-9\s-]+)\s+(?:\d+kV|Line|substation)', scope)
            if area_match:
                area = area_match.group(1).strip()

    return area.strip(), part.strip(), phase.strip()

def safe_divide(numerator, denominator):
    try:
        n = float(numerator)
        d = float(denominator)
        if d == 0:
            return None
        return round(n / d, 2)
    except (ValueError, TypeError):
        return None

def main():
    print("Loading source data (Explicitly skipping Row 1 and 2)...")
    # Using header=1 (0-indexed) means Excel Row 2 is the header.
    # Data starts from Excel Row 3.
    df_source = pd.read_excel(SOURCE_FILE, header=1)
    
    print(f"Total source records (starting from Source Row 3): {len(df_source)}")
    print(f"First record (Source Row 3) Scope Snippet: {repr(str(df_source.iloc[0].get('Name of the Transmission Project &\nScope'))[:100])}")

    print("Opening target workbook...")
    wb = openpyxl.load_workbook(TARGET_FILE)
    if TARGET_SHEET not in wb.sheetnames:
        print(f"Error: Sheet '{TARGET_SHEET}' not found")
        return
    
    ws = wb[TARGET_SHEET]

    # Clear existing data rows to ensure a clean population mapping
    # ONLY clear columns C, D, E (3, 4, 5) and P to AD (16 to 30)
    max_r = ws.max_row
    if max_r >= START_ROW_TARGET:
        print(f"Clearing existing mapped data from Row {START_ROW_TARGET} to {max_r}...")
        for r in range(START_ROW_TARGET, max_r + 1):
            # Clear C, D, E
            for c in range(3, 6): 
                ws.cell(row=r, column=c).value = None
            # Clear P to AD
            for c in range(16, 31):
                ws.cell(row=r, column=c).value = None

    row_count = 0
    for idx, row in df_source.iterrows():
        target_row = START_ROW_TARGET + row_count
        
        scope_text = str(row.get('Name of the Transmission Project &\nScope', ''))
        area, part, phase = parse_scope(scope_text)
        
        # Transmission Scheme = AREA + Part + Phase
        # Join with single spaces, ignoring empty strings
        scheme_parts = [area, part, phase]
        scheme = " ".join([p for p in scheme_parts if p])
        
        # Mandatory Write Map
        # C -> Inter/Intra Tx. Element
        ws.cell(row=target_row, column=3, value=area)
        # D -> Transmission Scheme
        ws.cell(row=target_row, column=4, value=scheme)
        # E -> Transmission Scope
        ws.cell(row=target_row, column=5, value=scope_text)
        
        # I -> Mode (TBCB/RTM) - Set to TBCB for TBCB PDF data
        ws.cell(row=target_row, column=9, value="TBCB")
        
        # P -> SPV Transfer Date
        ws.cell(row=target_row, column=16, value=row.get('SPV\nTransfe\nr Date'))
        
        # Q -> Length
        length = row.get('Lengt\nh')
        ws.cell(row=target_row, column=17, value=length)
        
        # R -> Location
        location = row.get('Total\nLocs.\n(Nos)')
        ws.cell(row=target_row, column=18, value=location)
        
        # S -> Foundation
        foundation = row.get('Found\nation\n(Nos)')
        ws.cell(row=target_row, column=19, value=foundation)
        
        # T -> Erection
        erection = row.get('Erecti\non\n(Nos)')
        ws.cell(row=target_row, column=20, value=erection)
        
        # U -> Stringing
        stringing = row.get('Stringin\ng (ckm)')
        ws.cell(row=target_row, column=21, value=stringing)
        
        # Calculations (V, W, X)
        ws.cell(row=target_row, column=22, value=safe_divide(foundation, location))
        ws.cell(row=target_row, column=23, value=safe_divide(erection, location))
        ws.cell(row=target_row, column=24, value=safe_divide(stringing, length))
        
        # Y -> Civil Work (%)
        ws.cell(row=target_row, column=25, value=row.get('Civil\nworks\n(%)'))
        # Z -> Equipment Received (%)
        ws.cell(row=target_row, column=26, value=row.get('Eqpt.\nReceive\nd (%)'))
        # AA -> Equipment Erected (%)
        ws.cell(row=target_row, column=27, value=row.get('Eqpt.\nErectio\nn (%)'))
        
        # AB/AC -> SCOD
        ws.cell(row=target_row, column=28, value=row.get('Target -\nOrg'))
        ws.cell(row=target_row, column=29, value=row.get('Target -\nAnticipate\nd'))
        
        # AD -> Remarks
        ws.cell(row=target_row, column=30, value=row.get('Remarks / Constraints & assistance required'))
        
        if row_count == 0:
            print(f"Verified Row {START_ROW_TARGET} (Target) populated with Source Row 3 info: AREA='{area}', Scheme='{scheme}'")
        
        row_count += 1

    print(f"Saving to {TARGET_FILE}...")
    wb.save(TARGET_FILE)
    print(f"Population successful. Added {row_count} rows starting from Row {START_ROW_TARGET}.")

if __name__ == "__main__":
    main()
