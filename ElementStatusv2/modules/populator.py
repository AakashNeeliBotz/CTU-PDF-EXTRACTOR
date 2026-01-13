
import pandas as pd
import openpyxl
import warnings

warnings.filterwarnings("ignore")

class ExcelPopulator:
    def __init__(self):
        self.src_cols_map = {
            'Scope': ['Name', 'Scope'],
            'SPV': ['SPV', 'Transfe'],
            'Locs': ['Total', 'Locs'],
            'Found': ['Found', 'ation', 'Nos'],
            'Erect': ['Erecti', 'on', 'Nos'],
            'String': ['Stringin', 'g'],
            'Civil': ['Civil', 'works'],
            'EqptRec': ['Eqpt', 'Receive'],
            'EqptEre': ['Eqpt', 'Erectio'],
            'OrgSCOD': ['Target', 'Org'],
            'AntSCOD': ['Target', 'Anticipate'],
            'Remarks': ['Remarks'],
            'Length': ['Lengt', 'h']
        }
        
        # Absolute Column Indices (1-based)
        self.mapping_rules = {
            16: 'SPV', 17: 'Length', 18: 'Locs', 19: 'Found', 20: 'Erect',
            21: 'String', 22: 'CALC_FOUND', 23: 'CALC_ERECT', 24: 'CALC_STRING',
            25: 'Civil', 26: 'EqptRec', 27: 'EqptEre', 28: 'OrgSCOD',
            29: 'AntSCOD', 30: 'Remarks'
        }

    def clean(self, val):
        if pd.isna(val) or val is None: return None
        # Normalize: strip, replace newlines/tabs with space, collapse multiple spaces
        text = str(val).strip().replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
        # Standardize dashes and common special characters
        text = text.replace('û', '-').replace('–', '-').replace('—', '-')
        import re
        text = re.sub(r'\s+', ' ', text).strip()
        # Lowercase for consistent semantic matching
        return text.lower()

    def to_float(self, val):
        try: return float(val) if not pd.isna(val) else 0.0
        except: return 0.0

    def find_col(self, df, keywords):
        for col in df.columns:
            if all(k.lower() in str(col).lower() for k in keywords):
                return col
        return None

    def sync(self, source_path, target_path):
        print(f"Syncing {source_path} to {target_path}...")
        
        # Robust Header Detection: Try Row 1, then Row 0
        df_src = None
        for h_val in [1, 0]:
            try:
                temp_df = pd.read_excel(source_path, header=h_val)
                if self.find_col(temp_df, ['Name', 'Scope']):
                    df_src = temp_df
                    print(f"  Detected headers at Row {h_val}")
                    break
            except: continue
            
        if df_src is None:
            return "Error: Could not find 'Scope' column in source (tried header rows 0 & 1)."
            
        src_cols = {}
        for k, kw in self.src_cols_map.items():
            found = self.find_col(df_src, kw)
            if found: src_cols[k] = found
            else: print(f"  Warning: Missing source col for {k} ({kw})")

        src_data = {}
        for _, r in df_src.iterrows():
            k = self.clean(r[src_cols['Scope']])
            if k: src_data[k] = r
            
        print(f"  Loaded {len(src_data)} records from source.")
        if src_data:
            print(f"  Sample Source Key: '{list(src_data.keys())[0]}'")

        wb = openpyxl.load_workbook(target_path)
        ws = wb["Element Status"]
        
        wb = openpyxl.load_workbook(target_path)
        ws = wb["Element Status"]
        
        # Identify Scope Column (Search Row 2)
        scope_col = None
        for cell in ws[2]:
            if cell.value and "Transmission Scope" in str(cell.value):
                scope_col = cell.column
                break
        
        if not scope_col: return "Scope column not found"

        updates, skips = 0, 0
        target_rows = list(ws.iter_rows(min_row=4))
        print(f"  Iterating {len(target_rows)} target rows.")
        
        # We will match first, then clear ONLY the relevant columns for the matched row
        src_keys = list(src_data.keys())
        matched_src_indices = set()
        
        for r_idx, row_cells in enumerate(target_rows):
            raw_target_val = row_cells[scope_col-1].value
            target_key = self.clean(raw_target_val)
            
            # If target key is empty, check if we should do sequential population
            # (If the target was just cleared or is a new template)
            src_row = None
            if not target_key or len(target_key) < 5:
                # Sequential Fallback: If we have source records and this target row is empty
                if r_idx < len(src_keys):
                    sk = src_keys[r_idx]
                    src_row = src_data[sk]
                    matched_src_indices.add(r_idx)
                    print(f"    Sequential Fill: Row {row_cells[0].row} -> '{sk[:30]}...'")
            else:
                # Match Logic: Strict semantic key match
                if target_key in src_data:
                    src_row = src_data[target_key]
                    # Find index
                    try: matched_src_indices.add(src_keys.index(target_key))
                    except: pass
                else:
                    # Fuzzy match fallback
                    for idx, sk in enumerate(src_keys):
                        if target_key in sk or sk in target_key:
                            src_row = src_data[sk]
                            matched_src_indices.add(idx)
                            print(f"    Fuzzy Match: '{target_key[:20]}' <-> '{sk[:20]}'")
                            break

            if src_row is None:
                continue

            # CLEAR specific columns for this row before writing
            for col_idx in [3, 4, 5]:
                ws.cell(row=row_cells[0].row, column=col_idx).value = None
            for col_idx in self.mapping_rules.keys():
                ws.cell(row=row_cells[0].row, column=col_idx).value = None

            # Write Transmission Scope into Col 5 (E)
            # Find the actual source name (uncleaned)
            src_scope_val = src_row[src_cols['Scope']]
            ws.cell(row=row_cells[0].row, column=5).value = src_scope_val
            
            # Populate Other Mapping Rules
            for col_idx, rule in self.mapping_rules.items():
                cell = ws.cell(row=row_cells[0].row, column=col_idx)
                val = None
                if rule.startswith('CALC_'):
                    try:
                        k = rule.split('_')[1]
                        num = self.to_float(src_row[src_cols['Found' if k=='FOUND' else 'Erect' if k=='ERECT' else 'String']])
                        den = self.to_float(src_row[src_cols['Locs' if k!='STRING' else 'Length']])
                        if den: val = round(num/den, 2)
                    except: pass
                else: 
                    src_col_name = src_cols.get(rule)
                    if src_col_name:
                        val = src_row[src_col_name]
                
                if val is not None and not pd.isna(val):
                    cell.value = val
                    updates += 1

        # APPEND any source records that were NOT matched to any target row
        unmatched_indices = [i for i in range(len(src_keys)) if i not in matched_src_indices]
        if unmatched_indices:
            print(f"  Appending {len(unmatched_indices)} additional records...")
            for idx in unmatched_indices:
                src_row = src_data[src_keys[idx]]
                new_row = [None] * max(31, max(self.mapping_rules.keys()) + 1)
                
                # Set Transmission Scope into Col 5 (index 4)
                new_row[4] = src_row[src_cols['Scope']]
                
                for col_idx, rule in self.mapping_rules.items():
                    val = None
                    if rule.startswith('CALC_'):
                        try:
                            k = rule.split('_')[1]
                            num = self.to_float(src_row[src_cols['Found' if k=='FOUND' else 'Erect' if k=='ERECT' else 'String']])
                            den = self.to_float(src_row[src_cols['Locs' if k!='STRING' else 'Length']])
                            if den: val = round(num/den, 2)
                        except: pass
                    else:
                        src_col_name = src_cols.get(rule)
                        if src_col_name: val = src_row[src_col_name]
                    
                    if val is not None and not pd.isna(val):
                        new_row[col_idx-1] = val
                
                ws.append(new_row)
                updates += sum(1 for v in new_row if v is not None)

        wb.save(target_path)
        return f"Done: {updates} updates, {skips} skipped."
