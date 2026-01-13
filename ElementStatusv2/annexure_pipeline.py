
import os
import shutil
from openpyxl import load_workbook
from modules.pdf_processor import AnnexureExtractor

# Configuration
MINUTES_PDF_PATH = "172838877090Minutes of meeting 34th CMETS NR Meeting held on 20-9-24.pdf"
EXCEL_PATH = "Connectivity Application Data 1.xlsx"
OUTPUT_PATH = "Connectivity Application Data 1_Updated.xlsx"
SHEET_NAME = "Element Status"

def find_column(ws, search_text, max_rows=6):
    """Find column index containing search_text in header rows."""
    for row_idx in range(1, max_rows):
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val and search_text in str(cell_val):
                return col_idx, row_idx
    return None, None

def main():
    print("=" * 60)
    print("ANNEXURE EXTRACTION PIPELINE")
    print("Processing: Minutes of Meeting PDF")
    print("=" * 60)
    
    # 1. Check source files
    if not os.path.exists(MINUTES_PDF_PATH):
        print(f"Error: PDF not found at {MINUTES_PDF_PATH}")
        return
        
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: Excel not found at {EXCEL_PATH}")
        return

    # 2. Copy Original Excel (Preserves ALL headers, formatting, existing data)
    print(f"\nCopying {EXCEL_PATH} to {OUTPUT_PATH}...")
    shutil.copy2(EXCEL_PATH, OUTPUT_PATH)

    # 3. Open Copy with openpyxl
    print(f"Opening {OUTPUT_PATH}...")
    wb = load_workbook(OUTPUT_PATH)
    
    if SHEET_NAME not in wb.sheetnames:
        print(f"Error: Sheet '{SHEET_NAME}' not found.")
        return
        
    ws = wb[SHEET_NAME]
    
    # Find required column
    tx_col_idx, _ = find_column(ws, "Transmission Scope")
    
    print(f"Transmission Scope column: {tx_col_idx}")
    
    if not tx_col_idx:
        print("Error: Could not find 'Transmission Scope' column.")
        return

    # 4. Extract Annexures from Minutes of Meeting PDF
    print("\n" + "-" * 60)
    print("Extracting Annexures from Meeting Minutes PDF...")
    print("-" * 60)
    
    ann_extractor = AnnexureExtractor(MINUTES_PDF_PATH)
    annexures = ann_extractor.extract_annexures()
    
    print(f"Extracted {len(annexures)} annexures.")
    
    # 5. Append Annexures at the end (after existing data)
    current_row = ws.max_row + 1
    print(f"Starting append at Row {current_row}...")
    
    updates_count = 0
    
    for ann in annexures:
        # Title Row with Annexure number
        title_text = f"{ann['annexure_number']}: {ann['title']}"
        ws.cell(row=current_row, column=tx_col_idx, value=title_text)
        current_row += 1
        updates_count += 1
        
        # Body Rows
        body_lines = ann['body'].split('\n')
        for line in body_lines:
            line = line.strip()
            if not line:
                continue
            ws.cell(row=current_row, column=tx_col_idx, value=line)
            current_row += 1
            updates_count += 1
            
        # Spacer Row between annexures
        current_row += 1

    # 6. Save
    print("\n" + "=" * 60)
    print(f"Appended {updates_count} new rows from Annexures")
    print("=" * 60)
    
    wb.save(OUTPUT_PATH)
    print(f"\nSaved to {OUTPUT_PATH}")
    print("Success!")
    print("\nNote: Original TBCB data is preserved. Only Annexures were appended.")

if __name__ == "__main__":
    main()
