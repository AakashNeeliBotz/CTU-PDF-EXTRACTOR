import openpyxl

wb = openpyxl.load_workbook('Connectivity_Application_Data_TEST_SN3_betterPrompts2.xlsx')
ws = wb['Data to be captured']

print(f'Sheet: {ws.title}')
print(f'Max row: {ws.max_row}')
print(f'Max column: {ws.max_column}')

print(f'\nFirst data row (row 3):')
for col in range(2, min(12, ws.max_column + 1)):
    value = ws.cell(3, col).value
    print(f'  Col {col}: {value}')

print(f'\nRow count check: {ws.max_row} rows total')
print(f'Data rows: {ws.max_row - 2} (excluding header rows)')
