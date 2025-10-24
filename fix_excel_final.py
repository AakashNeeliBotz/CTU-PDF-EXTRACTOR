"""
Final complete fix for Excel file:
1. Remove "Unnamed" row 1
2. Keep headers in place (row 1, starting column B)
3. Shift ONLY data rows one column to the right
"""
import openpyxl
from openpyxl.cell.cell import MergedCell

input_file = "Connectivity_Application_Data_TEST_SN3_FIXED.xlsx"
output_file = "Connectivity_Application_Data_TEST_SN3_FINAL.xlsx"

print(f"Fixing: {input_file}")
print("="*70)

# Load the workbook
wb = openpyxl.load_workbook(input_file)
ws = wb["Data to be captured"]

print(f"\nBefore fix:")
print(f"Max Row: {ws.max_row}, Max Column: {ws.max_column}")

# Show current state
print(f"\nCurrent state (first 8 rows, first 7 columns):")
for row_num in range(1, 9):
    row_data = [ws.cell(row=row_num, column=i).value for i in range(1, 8)]
    print(f"Row {row_num}: {[str(x)[:15] if x else '---' for x in row_data]}")

print(f"\n{'='*70}")
print("Applying fixes...")
print(f"{'='*70}")

# Step 1: Delete row 1 (the "Unnamed" row)
print(f"\n[1/2] Deleting row 1 (Unnamed headers)...")
ws.delete_rows(1, 1)

# After deleting row 1, what was row 2 is now row 1 (headers)
# What was row 8 is now row 7 (first data row)

# Step 2: Shift ONLY data rows (row 7 onwards) one column to the right
print(f"[2/2] Shifting data rows one column to the right...")

# First data row is now at row 7 (was row 8 before deletion)
first_data_row = 7
max_row = ws.max_row
max_col = ws.max_column

# Shift data rows from right to left to avoid overwriting
for row_num in range(first_data_row, max_row + 1):
    # Move cells from right to left
    for col_num in range(max_col, 0, -1):
        source_cell = ws.cell(row=row_num, column=col_num)
        target_cell = ws.cell(row=row_num, column=col_num + 1)
        
        # Skip merged cells
        if isinstance(source_cell, MergedCell):
            continue
        
        # Copy value
        target_cell.value = source_cell.value  # type: ignore
    
    # Clear column A for data rows
    first_cell = ws.cell(row=row_num, column=1)
    if not isinstance(first_cell, MergedCell):
        first_cell.value = None

print(f"[+] Fixes applied successfully!")

# Save
wb.save(output_file)
print(f"\n[+] Saved to: {output_file}")

# Verify
print(f"\n{'='*70}")
print("Verification:")
print(f"{'='*70}")

wb2 = openpyxl.load_workbook(output_file)
ws2 = wb2["Data to be captured"]

print(f"\nAfter fix:")
print(f"Max Row: {ws2.max_row}, Max Column: {ws2.max_column}")

print(f"\nFirst 10 rows after fix (first 7 columns):")
for row_num in range(1, min(11, ws2.max_row + 1)):
    row_data = [ws2.cell(row=row_num, column=i).value for i in range(1, 8)]
    print(f"Row {row_num}: {[str(x)[:15] if x else '---' for x in row_data]}")

# Check alignment
print(f"\n{'='*70}")
print("Alignment Check:")
print(f"{'='*70}")
header_row = [ws2.cell(row=1, column=i).value for i in range(2, 8)]
data_row = [ws2.cell(row=7, column=i).value for i in range(2, 8)]

print(f"\nHeaders (Row 1, Col B-G): {header_row}")
print(f"Data    (Row 7, Col B-G): {data_row}")

print(f"\nExpected alignment:")
print(f"  Sr.no. → {data_row[0] if len(data_row) > 0 else 'N/A'}")
print(f"  Region → {data_row[1] if len(data_row) > 1 else 'N/A'}")  
print(f"  State  → {data_row[2] if len(data_row) > 2 else 'N/A'}")

wb2.close()
wb.close()

print(f"\n{'='*70}")
print("Fix complete!")
print(f"{'='*70}")
