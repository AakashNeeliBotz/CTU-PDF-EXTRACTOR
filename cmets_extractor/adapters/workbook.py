from __future__ import annotations

import os
import shutil
from datetime import datetime
from typing import Any

from openpyxl import load_workbook

from cmets_extractor.config import (
    BULK_CONSUMERS_COLUMNS,
    BULK_CONSUMERS_DATE_FIELDS,
    BULK_CONSUMERS_NUMERIC_FIELDS,
    BULK_CONSUMERS_SHEET,
    DATA_CAPTURE_DATE_FIELDS,
    DATA_CAPTURE_NUMERIC_FIELDS,
    EXCEL_COLUMNS,
    MARGIN_FIELDS,
    MARGIN_SHEET,
    OUTPUT_EXCEL,
    TARGET_SHEET,
    TEMPLATE_EXCEL,
)
from cmets_extractor.domain.common.dates import normalize_output_date_text, parse_date
from cmets_extractor.domain.common.numbers import (
    convert_to_numeric,
    parse_numeric_value,
    to_int_if_whole,
)
from cmets_extractor.domain.common.text import clean_text


def write_to_excel(records, output_excel_path=None, template_excel_path=None):
    """Write extracted Data to be captured records to the workbook."""
    output_excel_path = output_excel_path or OUTPUT_EXCEL
    template_excel_path = template_excel_path or TEMPLATE_EXCEL
    print(f"\nWriting {len(records)} records to Excel...")

    shutil.copy(template_excel_path, output_excel_path)
    print(f"  Copied template to {output_excel_path}")

    wb = load_workbook(output_excel_path)
    ws = wb[TARGET_SHEET]

    start_row = 5

    for idx, record in enumerate(records):
        row_num = start_row + idx
        sr_no = idx + 1

        ws.cell(row=row_num, column=EXCEL_COLUMNS["sr_no"], value=sr_no)

        for field, col_num in EXCEL_COLUMNS.items():
            if field == "sr_no":
                continue

            value = record.get(field)
            if value is None:
                continue

            cell = ws.cell(
                row=row_num,
                column=col_num,
                value=prepare_data_capture_excel_value(field, value),
            )
            if field in DATA_CAPTURE_DATE_FIELDS and isinstance(cell.value, datetime):
                cell.number_format = "DD.MM.YYYY"

    wb.save(output_excel_path)
    wb.close()

    print(f"  Successfully wrote {len(records)} records to {output_excel_path}")
    print(f"  Data written to sheet: '{TARGET_SHEET}'")
    print(f"  Rows: {start_row} to {start_row + len(records) - 1}")
    return output_excel_path


def write_margin_to_excel(output_excel_path, margin_records):
    """Write Margin records into the legacy Margin sheet layout."""
    if not margin_records:
        print("\nSkipping Margin sheet write: no Margin records extracted.")
        return
    if not os.path.exists(output_excel_path):
        print(f"\nSkipping Margin sheet write: workbook not found at '{output_excel_path}'")
        return

    print(f"\nWriting {len(margin_records)} Margin records to Excel...")
    wb = load_workbook(output_excel_path)
    if MARGIN_SHEET not in wb.sheetnames:
        wb.close()
        print(f"Skipping Margin sheet write: Sheet '{MARGIN_SHEET}' not found in workbook.")
        return

    ws = wb[MARGIN_SHEET]
    clear_start_row = 5
    if ws.max_row >= clear_start_row:
        ws.delete_rows(clear_start_row, ws.max_row - clear_start_row + 1)

    start_row = clear_start_row
    start_col = 2
    for record in margin_records:
        for col_idx, field_name in enumerate(MARGIN_FIELDS):
            cell_value = convert_to_numeric(record.get(field_name))
            ws.cell(row=start_row, column=start_col + col_idx, value=cell_value)
        start_row += 1

    wb.save(output_excel_path)
    wb.close()
    print(f"  Successfully wrote {len(margin_records)} records to sheet: '{MARGIN_SHEET}'")
    print(f"  Rows: {clear_start_row} to {start_row - 1}")


def write_bulk_consumers_to_excel(output_excel_path, bulk_records):
    """Write Bulk Consumers records into the dedicated workbook sheet."""
    if not os.path.exists(output_excel_path):
        print(f"\nSkipping Bulk Consumers sheet write: workbook not found at '{output_excel_path}'")
        return

    wb = load_workbook(output_excel_path)
    if BULK_CONSUMERS_SHEET not in wb.sheetnames:
        wb.close()
        print(
            f"Skipping Bulk Consumers sheet write: Sheet '{BULK_CONSUMERS_SHEET}' "
            "not found in workbook."
        )
        return

    ws = wb[BULK_CONSUMERS_SHEET]
    clear_start_row = 3
    if ws.max_row >= clear_start_row:
        ws.delete_rows(clear_start_row, ws.max_row - clear_start_row + 1)

    print(f"\nWriting {len(bulk_records)} Bulk Consumers records to Excel...")
    for idx, record in enumerate(bulk_records, start=1):
        row_num = clear_start_row + idx - 1
        ws.cell(row=row_num, column=BULK_CONSUMERS_COLUMNS["sr_no"], value=idx)

        for field, col_num in BULK_CONSUMERS_COLUMNS.items():
            if field == "sr_no":
                continue

            value = record.get(field)
            if value is None:
                continue

            cell = ws.cell(
                row=row_num,
                column=col_num,
                value=prepare_bulk_consumers_excel_value(field, value),
            )
            if field in BULK_CONSUMERS_DATE_FIELDS and isinstance(cell.value, datetime):
                cell.number_format = "DD.MM.YYYY"

    wb.save(output_excel_path)
    wb.close()
    print(
        f"  Successfully wrote {len(bulk_records)} records to sheet: "
        f"'{BULK_CONSUMERS_SHEET}'"
    )
    if bulk_records:
        print(f"  Rows: {clear_start_row} to {clear_start_row + len(bulk_records) - 1}")


def prepare_data_capture_excel_value(field: str, value: Any):
    """Preserve dates and numbers as native Excel types instead of raw strings."""
    if value is None:
        return None

    if field in DATA_CAPTURE_DATE_FIELDS:
        if isinstance(value, datetime):
            return value
        parsed = parse_date(normalize_output_date_text(value))
        if parsed:
            return parsed
        return clean_text(value)

    if field in DATA_CAPTURE_NUMERIC_FIELDS:
        numeric = parse_numeric_value(value)
        if numeric is not None:
            return to_int_if_whole(numeric)

    return value


def prepare_bulk_consumers_excel_value(field: str, value: Any):
    """Write Bulk Consumers dates as normalized text to avoid Excel serial display."""
    if value is None:
        return None

    if field in BULK_CONSUMERS_DATE_FIELDS:
        if isinstance(value, datetime):
            return value.strftime("%d.%m.%Y")
        normalized = normalize_output_date_text(value)
        return normalized or clean_text(value)

    if field in BULK_CONSUMERS_NUMERIC_FIELDS:
        numeric = parse_numeric_value(value)
        if numeric is not None:
            return to_int_if_whole(numeric)

    return value


__all__ = [
    "prepare_bulk_consumers_excel_value",
    "prepare_data_capture_excel_value",
    "write_bulk_consumers_to_excel",
    "write_margin_to_excel",
    "write_to_excel",
]
