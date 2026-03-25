from __future__ import annotations

import os

import pandas as pd
from openpyxl import load_workbook

from cmets_extractor.config import (
    ELEMENT_STATUS_SHEET,
    ELEMENT_STATUS_TARGET_TEXT,
    RTM_ELEMENT_STATUS_TARGET_TEXTS,
    RTM_PDF_PATH,
    TBCB_PDF_PATH,
)
from cmets_extractor.domain.element_status import (
    ELEMENT_STATUS_MAPPING_RULES,
    build_element_status_source_data,
    build_nct_element_status_source_data,
    es_find_existing_scope_row,
    es_is_number,
    es_normalize_code,
    es_register_scope_row_keys,
    es_rule_value,
    es_scope_match_keys,
    es_source_scope_and_code,
)


def es_build_existing_row_index(ws, start_row=4):
    """
    Build lookup indices from existing sheet rows.
    - by full key: (scope, element_code)
    - by scope: scope only
    """
    index_by_full = {}
    index_by_scope = {}

    for row_num in range(start_row, ws.max_row + 1):
        scope_value = ws.cell(row=row_num, column=5).value
        match_keys = es_scope_match_keys(scope_value)
        if not match_keys:
            continue

        code_key = es_normalize_code(ws.cell(row=row_num, column=2).value)
        es_register_scope_row_keys(match_keys, code_key, row_num, index_by_full, index_by_scope)

    return index_by_full, index_by_scope


def es_write_sheet_row(ws, row_num, src_row, src_cols, source_label="TBCB"):
    """Write one Element Status source record into a target Excel row."""
    updates = 0

    clear_cols = {3, 4, 5}
    clear_cols.update(ELEMENT_STATUS_MAPPING_RULES.keys())
    for col_num in clear_cols:
        ws.cell(row=row_num, column=col_num, value=None)

    scope_raw, _, _, _ = es_source_scope_and_code(src_row, src_cols)
    if scope_raw is not None and str(scope_raw).strip():
        ws.cell(row=row_num, column=5, value=scope_raw)
        updates += 1

    for col_num, rule in ELEMENT_STATUS_MAPPING_RULES.items():
        value = es_rule_value(src_row, src_cols, rule, source_label=source_label)
        if value is None or pd.isna(value):
            continue
        if isinstance(value, str):
            is_num, numeric = es_is_number(value)
            value = numeric if is_num else value
        ws.cell(row=row_num, column=col_num, value=value)
        updates += 1

    return updates


def es_ensure_element_status_headers(ws):
    """Keep Element Status column headers aligned with current writer behavior."""
    if ws.cell(row=2, column=2).value != "Element Code":
        ws.cell(row=2, column=2, value="Element Code")
    if ws.cell(row=2, column=1).value == "Element Code":
        ws.cell(row=2, column=1, value=None)
    if ws.cell(row=2, column=9).value != "Mode (TBCB/RTM/NCT)":
        ws.cell(row=2, column=9, value="Mode (TBCB/RTM/NCT)")


def es_populate_sheet_from_source_rows(
    ws,
    source_rows,
    src_cols,
    source_label="TBCB",
    overwrite_existing=False,
):
    """Update or append parsed Element Status source rows into one worksheet."""
    index_by_full, index_by_scope = es_build_existing_row_index(ws, start_row=4)
    cell_updates = 0
    updated_rows = 0
    appended_rows = 0
    total_source = 0

    row_iter = source_rows.values() if hasattr(source_rows, "values") else source_rows
    for src_row in row_iter:
        total_source += 1
        _, scope_key, code_key, match_keys = es_source_scope_and_code(src_row, src_cols)
        if not scope_key:
            continue

        target_row = es_find_existing_scope_row(match_keys, code_key, index_by_full, index_by_scope)

        if target_row is None:
            target_row = ws.max_row + 1
            appended_rows += 1
            cell_updates += es_write_sheet_row(
                ws,
                target_row,
                src_row,
                src_cols,
                source_label=source_label,
            )
        else:
            updated_rows += 1
            if overwrite_existing:
                cell_updates += es_write_sheet_row(
                    ws,
                    target_row,
                    src_row,
                    src_cols,
                    source_label=source_label,
                )

        es_register_scope_row_keys(match_keys, code_key, target_row, index_by_full, index_by_scope)

    return {
        "cell_updates": cell_updates,
        "updated_rows": updated_rows,
        "appended_rows": appended_rows,
        "total_source": total_source,
    }


def es_append_cmets_catalog_entries(ws, cmets_element_catalog):
    """Append CMETS ATS/DTL/CTS entries into the Element Status worksheet."""
    index_by_full, index_by_scope = es_build_existing_row_index(ws, start_row=4)
    appended_rows = 0
    updated_rows = 0

    entries = sorted(
        cmets_element_catalog.values(),
        key=lambda entry: (entry.get("code") or "", entry.get("scope") or ""),
    )
    for entry in entries:
        scope = entry.get("scope")
        code = entry.get("code")
        if not scope or not code:
            continue

        code_key = es_normalize_code(code)
        match_keys = es_scope_match_keys(scope, source_label="CMETS")
        row_num = es_find_existing_scope_row(match_keys, code_key, index_by_full, index_by_scope)
        if row_num is None:
            row_num = ws.max_row + 1
            appended_rows += 1
            ws.cell(row=row_num, column=2, value=code)
            ws.cell(row=row_num, column=5, value=scope)

            # Per requirement, keep Element Status columns C/D blank for CMETS rows.
            ws.cell(row=row_num, column=3, value=None)
            ws.cell(row=row_num, column=4, value=None)
            ws.cell(row=row_num, column=8, value=None)
            ws.cell(row=row_num, column=9, value=None)
        else:
            updated_rows += 1

        meetings = ", ".join(sorted(entry.get("meetings", [])))
        if meetings and not ws.cell(row=row_num, column=30).value:
            ws.cell(row=row_num, column=30, value=f"CMETS meeting(s): {meetings}")

        es_register_scope_row_keys(match_keys, code_key, row_num, index_by_full, index_by_scope)

    return {
        "updated_rows": updated_rows,
        "appended_rows": appended_rows,
        "catalog_size": len(cmets_element_catalog),
    }


def populate_element_status_sheet_from_source_pdf(
    output_excel_path,
    pdf_path,
    source_label,
    target_texts=None,
):
    """Populate Element Status sheet from one monitoring PDF source."""
    if not os.path.exists(pdf_path):
        print(f"\nSkipping Element Status: {source_label} PDF not found at '{pdf_path}'")
        return
    if not os.path.exists(output_excel_path):
        print(f"\nSkipping Element Status: Output Excel not found at '{output_excel_path}'")
        return

    src_data, src_cols = build_element_status_source_data(
        pdf_path,
        source_label=source_label,
        target_texts=target_texts,
    )
    if not src_data:
        print(f"Skipping Element Status: No source records parsed from {source_label} PDF.")
        return

    wb = load_workbook(output_excel_path)
    if ELEMENT_STATUS_SHEET not in wb.sheetnames:
        wb.close()
        print(f"Skipping Element Status: Sheet '{ELEMENT_STATUS_SHEET}' not found in workbook.")
        return

    ws = wb[ELEMENT_STATUS_SHEET]
    es_ensure_element_status_headers(ws)

    scope_col = None
    for cell in ws[2]:
        if cell.value and "Transmission Scope" in str(cell.value):
            scope_col = cell.column
            break
    if not scope_col:
        wb.close()
        print("Skipping Element Status: 'Transmission Scope' header not found in row 2.")
        return

    print(f"\nPopulating Element Status from {source_label} ({len(src_data)} source records)...")
    stats = es_populate_sheet_from_source_rows(
        ws,
        src_data.values(),
        src_cols,
        source_label=source_label,
        overwrite_existing=False,
    )

    wb.save(output_excel_path)
    wb.close()

    print(
        f"Element Status population complete ({source_label}): "
        f"cell updates={stats['cell_updates']}, updated rows={stats['updated_rows']}, "
        f"appended rows={stats['appended_rows']}, total source={stats['total_source']}"
    )


def populate_element_status_sheet_from_monitoring_pdfs(output_excel_path):
    """Populate Element Status sheet from the monitoring PDF sources."""
    source_configs = [
        ("TBCB", TBCB_PDF_PATH, (ELEMENT_STATUS_TARGET_TEXT,)),
        ("RTM", RTM_PDF_PATH, RTM_ELEMENT_STATUS_TARGET_TEXTS),
    ]
    for source_label, pdf_path, target_texts in source_configs:
        populate_element_status_sheet_from_source_pdf(
            output_excel_path,
            pdf_path,
            source_label=source_label,
            target_texts=target_texts,
        )


def populate_element_status_sheet_from_nct_pdf(
    output_excel_path,
    pdf_path,
    source_label="NCT",
    overwrite_existing=False,
):
    """Populate Element Status sheet from the dedicated NCT minutes PDF."""
    if not os.path.exists(pdf_path):
        print(f"\nSkipping Element Status: {source_label} PDF not found at '{pdf_path}'")
        return
    if not os.path.exists(output_excel_path):
        print(f"\nSkipping Element Status: Output Excel not found at '{output_excel_path}'")
        return

    src_data = build_nct_element_status_source_data(pdf_path)
    if not src_data:
        print(f"Skipping Element Status: No source records parsed from {source_label} PDF.")
        return

    wb = load_workbook(output_excel_path)
    if ELEMENT_STATUS_SHEET not in wb.sheetnames:
        wb.close()
        print(f"Skipping Element Status: Sheet '{ELEMENT_STATUS_SHEET}' not found in workbook.")
        return

    ws = wb[ELEMENT_STATUS_SHEET]
    es_ensure_element_status_headers(ws)
    src_cols = {
        "Scope": "Scope",
        "ElementCode": "ElementCode",
        "MVA": "MVA",
        "Length": "Length",
        "Remarks": "Remarks",
        "AwardedTo": "AwardedTo",
        "Mode": "Mode",
    }

    print(f"\nPopulating Element Status from {source_label} ({len(src_data)} source records)...")
    stats = es_populate_sheet_from_source_rows(
        ws,
        src_data.values(),
        src_cols,
        source_label=source_label,
        overwrite_existing=overwrite_existing,
    )

    wb.save(output_excel_path)
    wb.close()
    print(
        f"Element Status population complete ({source_label}): "
        f"cell updates={stats['cell_updates']}, updated rows={stats['updated_rows']}, "
        f"appended rows={stats['appended_rows']}, total source={stats['total_source']}"
    )


def populate_element_status_sheet_from_cmets(output_excel_path, cmets_element_catalog):
    """Append CMETS ATS/DTL/CTS elements into Element Status sheet with unique codes."""
    if not cmets_element_catalog:
        print("\nSkipping CMETS Element Status append: no CMETS elements collected.")
        return
    if not os.path.exists(output_excel_path):
        print(f"\nSkipping CMETS Element Status append: Output Excel not found at '{output_excel_path}'")
        return

    wb = load_workbook(output_excel_path)
    if ELEMENT_STATUS_SHEET not in wb.sheetnames:
        wb.close()
        print(f"Skipping CMETS Element Status append: Sheet '{ELEMENT_STATUS_SHEET}' not found.")
        return

    ws = wb[ELEMENT_STATUS_SHEET]
    stats = es_append_cmets_catalog_entries(ws, cmets_element_catalog)

    wb.save(output_excel_path)
    wb.close()
    print(
        f"CMETS Element Status append complete: "
        f"updated rows={stats['updated_rows']}, appended rows={stats['appended_rows']}, "
        f"catalog size={stats['catalog_size']}"
    )


__all__ = [
    "es_append_cmets_catalog_entries",
    "es_build_existing_row_index",
    "es_ensure_element_status_headers",
    "es_populate_sheet_from_source_rows",
    "es_write_sheet_row",
    "populate_element_status_sheet_from_cmets",
    "populate_element_status_sheet_from_monitoring_pdfs",
    "populate_element_status_sheet_from_nct_pdf",
    "populate_element_status_sheet_from_source_pdf",
]
