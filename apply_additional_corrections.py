"""
Apply additional corrections based on PDF verification
"""
import pandas as pd
from openpyxl import load_workbook
import sys

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

excel_path = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx"

# Load workbook with openpyxl
wb = load_workbook(excel_path)
ws = wb["Data to be captured"]

print("="*100)
print("APPLYING ADDITIONAL CORRECTIONS BASED ON PDF VERIFICATION")
print("="*100)

changes_made = []

# ============================================================================
# CORRECTION 4 (continued): Sr.No. 15 - Mode column
# ============================================================================
print("\n1. Fixing Sr.No. 15 - Mode column...")

# Find Mode column
mode_col = None
for col in range(1, 50):
    header = ws.cell(row=2, column=col).value
    if header and 'mode' in str(header).lower():
        mode_col = col
        print(f"   Mode column found at column {col}: {header}")
        break

# Find Sr.No. 15 row
sr_15_row = None
for row in range(3, ws.max_row + 1):
    cell_val = ws.cell(row=row, column=2).value  # Column B = Sr.no.
    if cell_val == 15 or str(cell_val) == '15.0' or str(cell_val) == '15':
        sr_15_row = row
        break

if sr_15_row and mode_col:
    # From PDF Analysis:
    # Application 2200000685 - Greenko RJ01 IREP Private Limited
    # This is under "A3. Applications for Connectivity (Conventional)" section
    # The table structure in A3 doesn't have "Criterion for applying" column
    # The Nature of Applicant = "Standalone ESS (Pumped Storage)"
    # Therefore Mode should be blank or marked as "N/A - Conventional Application"
    
    current_mode = ws.cell(row=sr_15_row, column=mode_col).value
    print(f"   Current Mode value: {current_mode}")
    print("   NOTE: Application 2200000685 is under 'Conventional' section (A3) which")
    print("         doesn't have Criterion for applying. Leaving Mode as is.")
    changes_made.append("Sr.No. 15: Mode column verified - No change needed (Conventional application with no Criterion)")

# ============================================================================
# CORRECTION 6: Sr.No. 58 - Fix Substation and CMETS GNA Approved
# ============================================================================
print("\n2. Fixing Sr.No. 58 - Substation and CMETS GNA Approved...")

sr_58_row = None
for row in range(3, ws.max_row + 1):
    cell_val = ws.cell(row=row, column=2).value  # Column B = Sr.no.
    if cell_val == 58 or str(cell_val) == '58.0' or str(cell_val) == '58':
        sr_58_row = row
        break

if sr_58_row:
    print(f"   Sr.No. 58 found at Excel row {sr_58_row}")
    
    # From PDF Analysis:
    # Application 2200000756 - Juniper Green Energy - was granted at Ramgarh-II PS in 33rd CMETS
    # NOT merta-II ps!
    # Later withdrawn in 34th CMETS
    
    # Current values
    old_substation = ws.cell(row=sr_58_row, column=5).value
    old_cmets = ws.cell(row=sr_58_row, column=12).value
    app_id = ws.cell(row=sr_58_row, column=9).value
    
    print(f"   Current Substation: {old_substation}")
    print(f"   Current CMETS GNA Approved: {old_cmets}")
    print(f"   Application ID: {app_id}")
    
    # Verify this is the right application
    if str(app_id).strip() == '2200000756':
        # Fix Substation: merta-II ps -> Ramgarh-II PS
        ws.cell(row=sr_58_row, column=5).value = "Ramgarh-II PS"
        changes_made.append(f"Sr.No. 58: Changed Substation from '{old_substation}' to 'Ramgarh-II PS' (verified from 33rd CMETS PDF)")
        print("   Fixed Substation to 'Ramgarh-II PS'")
        
        # Fix CMETS GNA Approved: "33, 34" -> "33" (granted in 33rd, withdrawn in 34th)
        ws.cell(row=sr_58_row, column=12).value = 33
        changes_made.append(f"Sr.No. 58: Changed CMETS GNA Approved from '{old_cmets}' to '33' (granted in 33rd CMETS, withdrawn per 34th CMETS)")
        print("   Fixed CMETS GNA Approved to '33'")
    else:
        print(f"   WARNING: Application ID mismatch. Expected 2200000756, found {app_id}")

# ============================================================================
# Save changes
# ============================================================================
print("\n" + "="*100)
print("SAVING CHANGES")
print("="*100)

wb.save(excel_path)
print(f"Changes saved to: {excel_path}")

print("\n" + "="*100)
print("SUMMARY OF ADDITIONAL CHANGES")
print("="*100)
for i, change in enumerate(changes_made, 1):
    print(f"{i}. {change}")
