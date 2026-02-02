import pdfplumber
import re
import pandas as pd
from openpyxl import load_workbook

def extract_application_ids_from_pdf(pdf_path, pdf_number):
    """Extract GNA/ST II and LTA Application IDs from PDF"""
    gna_ids = []
    lta_ids = []
    
    print(f'Processing PDF #{pdf_number}: {pdf_path.split("/")[-1]}')
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Extract text from all pages
            full_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"
            
            # Look for application IDs in the text
            # Pattern for 10-digit application IDs
            app_id_pattern = r'\b(\d{10})\b'
            matches = re.findall(app_id_pattern, full_text)
            
            # Filter for valid application IDs (starting with 220000)
            valid_app_ids = [match for match in matches if match.startswith('220000')]
            
            # Check context around each ID to determine if it's GNA or LTA
            for app_id in valid_app_ids:
                # Find position of the ID in text
                pos = full_text.find(app_id)
                if pos != -1:
                    # Get context (200 characters before and after)
                    start_pos = max(0, pos - 200)
                    end_pos = min(len(full_text), pos + 200)
                    context = full_text[start_pos:end_pos].lower()
                    
                    # Check if it's GNA/ST II or LTA based on context
                    if 'gna' in context or 'st ii' in context or 'grant' in context:
                        gna_ids.append((pdf_number, app_id))
                    elif 'lta' in context or 'long term' in context:
                        lta_ids.append((pdf_number, app_id))
    
    except Exception as e:
        print(f'Error processing PDF #{pdf_number}: {e}')
    
    print(f'  Found {len(gna_ids)} GNA IDs and {len(lta_ids)} LTA IDs')
    return gna_ids, lta_ids

def update_excel_with_cmets_approval():
    """Main function to update Excel with CMETS approval numbers"""
    
    # PDF file paths
    pdf33_path = 'c:/Users/Sree Charan/Desktop/fold2/CTU-PDF-EXTRACTOR/downloaded_pdfs/SN1/172381548953Minutes of 33rd CMETS NR meeting held on 05.08.2024.pdf'
    pdf34_path = 'c:/Users/Sree Charan/Desktop/fold2/CTU-PDF-EXTRACTOR/downloaded_pdfs/SN1/172838877090Minutes of meeting 34th CMETS NR Meeting held on 20-9-24.pdf'
    
    # Excel file path
    excel_path = 'c:/Users/Sree Charan/Desktop/fold2/CTU-PDF-EXTRACTOR/Connectivity_Application_Data_TEST_ALL_SHEETS38.xlsx'
    
    print("=== Extracting Application IDs from SN1 PDFs ===")
    
    # Extract IDs from both PDFs
    gna_ids_33, lta_ids_33 = extract_application_ids_from_pdf(pdf33_path, 33)
    gna_ids_34, lta_ids_34 = extract_application_ids_from_pdf(pdf34_path, 34)
    
    # Combine all IDs
    all_gna_ids = gna_ids_33 + gna_ids_34
    all_lta_ids = lta_ids_33 + lta_ids_34
    
    print(f"\nTotal extracted:")
    print(f"GNA/ST II IDs: {len(all_gna_ids)}")
    print(f"LTA IDs: {len(all_lta_ids)}")
    
    print("\nGNA IDs found:")
    for pdf_num, app_id in all_gna_ids:
        print(f"  PDF #{pdf_num}: {app_id}")
    
    print("\nLTA IDs found:")
    for pdf_num, app_id in all_lta_ids:
        print(f"  PDF #{pdf_num}: {app_id}")
    
    print("\n=== Processing Excel Data ===")
    
    # Load Excel file
    wb = load_workbook(excel_path)
    ws = wb['Data to be captured']
    
    # Find column indices (assuming row 2 contains headers)
    gna_col_idx = None
    lta_col_idx = None
    cmets_gna_col_idx = None
    cmets_lta_col_idx = None
    
    # Check row 2 for headers
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=2, column=col).value
        if cell_value:
            header_text = str(cell_value).strip().lower()
            if 'gna/st ii application id' in header_text:
                gna_col_idx = col
            elif 'lta application id' in header_text:
                lta_col_idx = col
            elif 'cmets gna approved' in header_text:
                cmets_gna_col_idx = col
            elif 'cmets lta approved' in header_text:
                cmets_lta_col_idx = col
    
    print(f"Column indices found:")
    print(f"  GNA/ST II Application ID: {gna_col_idx}")
    print(f"  LTA Application ID: {lta_col_idx}")
    print(f"  CMETS GNA Approved: {cmets_gna_col_idx}")
    print(f"  CMETS LTA Approved: {cmets_lta_col_idx}")
    
    # Process each row in the Excel sheet
    updated_gna_count = 0
    updated_lta_count = 0
    
    for row in range(3, ws.max_row + 1):  # Start from row 3 (data rows)
        # Get GNA and LTA IDs from current row
        gna_id_cell = ws.cell(row=row, column=gna_col_idx) if gna_col_idx else None
        lta_id_cell = ws.cell(row=row, column=lta_col_idx) if lta_col_idx else None
        
        gna_id = str(gna_id_cell.value).strip() if gna_id_cell and gna_id_cell.value else None
        lta_id = str(lta_id_cell.value).strip() if lta_id_cell and lta_id_cell.value else None
        
        # Check if we have valid IDs to match
        cmets_gna_number = None
        cmets_lta_number = None
        
        # Match GNA ID
        if gna_id and len(gna_id) == 10 and gna_id.startswith('220000'):
            for pdf_num, extracted_id in all_gna_ids:
                if extracted_id == gna_id:
                    cmets_gna_number = pdf_num
                    break
        
        # Match LTA ID
        if lta_id and len(lta_id) == 10 and lta_id.startswith('220000'):
            for pdf_num, extracted_id in all_lta_ids:
                if extracted_id == lta_id:
                    cmets_lta_number = pdf_num
                    break
        
        # Update CMETS approval columns if matches found
        if cmets_gna_number and cmets_gna_col_idx:
            ws.cell(row=row, column=cmets_gna_col_idx).value = cmets_gna_number
            updated_gna_count += 1
            
        if cmets_lta_number and cmets_lta_col_idx:
            ws.cell(row=row, column=cmets_lta_col_idx).value = cmets_lta_number
            updated_lta_count += 1
    
    print(f"\nUpdates made:")
    print(f"  CMETS GNA Approved updated: {updated_gna_count} rows")
    print(f"  CMETS LTA Approved updated: {updated_lta_count} rows")
    
    # Save the workbook
    wb.save(excel_path)
    print(f"\nExcel file updated successfully: {excel_path}")

if __name__ == "__main__":
    update_excel_with_cmets_approval()