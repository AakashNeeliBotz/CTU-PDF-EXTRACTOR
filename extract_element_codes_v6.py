"""
Final extraction with enhanced debugging for CTS matching
"""
import pandas as pd
import fitz
import re
from openpyxl import load_workbook
from difflib import SequenceMatcher
import warnings
warnings.filterwarnings('ignore')

EXCEL_PATH = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx"
PDF_33 = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\downloaded_pdfs\SN1\172381548953Minutes of 33rd CMETS NR meeting held on 05.08.2024.pdf"
PDF_34 = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\downloaded_pdfs\SN1\172838877090Minutes of meeting 34th CMETS NR Meeting held on 20-9-24.pdf"

COL_APP_ID = 8
COL_CTS = 39
COL_ATS = 40
COL_DTL = 41
DATA_START_ROW = 3

def normalize_text(text):
    if not text:
        return ""
    text = str(text).lower()
    text = ' '.join(text.split())
    text = text.replace('–', '-').replace('—', '-').replace('−', '-')
    text = re.sub(r'[^\w\s\-/]', ' ', text)
    return text.strip()

def load_element_status():
    print("[*] Loading Element Status sheet...")
    df = pd.read_excel(EXCEL_PATH, sheet_name='Element Status', header=None)
    
    element_map = {}
    code_to_desc = {}
    
    for i in range(len(df)):
        code_cell = df.iloc[i, 1] if len(df.columns) > 1 else None
        desc_cell = df.iloc[i, 4] if len(df.columns) > 4 else None
        
        if pd.notna(code_cell):
            code_str = str(code_cell).strip()
            if re.match(r'^EL-[A-Z0-9]{5}$', code_str):
                if pd.notna(desc_cell) and len(str(desc_cell).strip()) > 10:
                    desc_str = str(desc_cell).strip()
                    norm_desc = normalize_text(desc_str)
                    element_map[norm_desc] = code_str
                    code_to_desc[code_str] = desc_str
    
    print(f"[+] Loaded {len(code_to_desc)} Element Codes")
    return element_map, code_to_desc

def find_element_codes(description, element_map, threshold=0.45):
    if not description or len(description) < 15:
        return []
    
    desc_norm = normalize_text(description)
    matches = []
    
    if desc_norm in element_map:
        return [element_map[desc_norm]]
    
    # Keyword matching
    for key, code in element_map.items():
        if len(key) > 12:
            key_terms = set(re.findall(r'\b\w{4,}\b', key))
            desc_terms = set(re.findall(r'\b\w{4,}\b', desc_norm))
            
            if key_terms and desc_terms:
                overlap = len(key_terms & desc_terms) / max(len(key_terms), len(desc_terms))
                if overlap > 0.35:
                    matches.append(code)
    
    if matches:
        return list(set(matches))[:8]
    
    # Fuzzy matching
    candidates = []
    for key, code in element_map.items():
        if len(key) > 12:
            score = SequenceMatcher(None, desc_norm[:80], key[:80]).ratio()
            if score >= threshold:
                candidates.append((score, code))
    
    candidates.sort(reverse=True)
    return [code for score, code in candidates[:5]]

def extract_pdf_text(pdf_path):
    doc = fitz.open(pdf_path)
    text = ""
    for page in doc:
        text += page.get_text("text")
    doc.close()
    return text

def extract_annexures(pdf_text):
    """Extract all Annexure sections"""
    annexures = {}
    
    # Find each numbered Annexure
    pattern = r'Annexure-([IVX]+)\s*\n+(.*?)(?=\nAnnexure-[IVX]+\s*\n|\nMinutes of \d+|$)'
    matches = re.findall(pattern, pdf_text, re.DOTALL | re.IGNORECASE)
    
    for num, content in matches:
        content = re.sub(r'Minutes of \d+.*?Page \d+ of \d+', '', content, flags=re.DOTALL)
        
        items = []
        # Extract numbered items
        for match in re.finditer(r'(\d+)\.\s+(.+?)(?=\d+\.\s|\n\n|Additional|B\.\s+Additional|$)', content, re.DOTALL):
            item_text = match.group(2)
            cleaned = ' '.join(item_text.split())
            if len(cleaned) > 15 and any(kw in cleaned.lower() for kw in ['kv', 'mva', 'line', 'ps', 'hvdc', 'establishm', 'lilo', 'augment', 'substation']):
                items.append(cleaned)
        
        if items:
            annexures[num.upper()] = items
    
    return annexures

def extract_app_data(pdf_text, annexures):
    results = {}
    
    # Split by transmission section header
    sections = re.split(
        r'Details of Transmission system for Connectivity under GNA:\s*',
        pdf_text, flags=re.IGNORECASE
    )
    
    for i, section in enumerate(sections[1:], 1):
        prefix = sections[i-1][-4000:] if len(sections[i-1]) > 4000 else sections[i-1]
        
        app_ids = list(re.findall(r'\b22\d{8}\b', prefix))
        app_ids += list(re.findall(r'\b22\d{8}\b', section[:2500]))
        
        if not app_ids:
            continue
        
        # Parse sections A, B, C
        ats_text = ""
        ats_match = re.search(r'A\.\s*Associated\s*Transmission\s*System\s*\(ATS\)[:\s]*(.+?)(?=B\.\s*Transmission|$)',
                             section[:2000], re.DOTALL | re.IGNORECASE)
        if ats_match:
            ats_raw = ' '.join(ats_match.group(1).split())
            ats_text = "NIL" if 'NIL' in ats_raw.upper() else ats_raw[:500]
        
        dtl_text = ""
        dtl_match = re.search(r'B\.\s*Transmission\s*System\s*under\s*applicant\s*scope[:\s]*(.+?)(?=C\.\s*Transmission|$)',
                             section[:3000], re.DOTALL | re.IGNORECASE)
        if dtl_match:
            dtl_raw = dtl_match.group(1)
            items = re.findall(r'\([ivx]+\)\.*\s*(.+?)(?=\([ivx]+\)|C\.|Minutes|Sl\.|$)', dtl_raw, re.DOTALL | re.IGNORECASE)
            if items:
                dtl_text = ' '.join([' '.join(it.split()) for it in items])[:600]
            else:
                dtl_text = ' '.join(dtl_raw.split())[:600]
        
        cts_text = ""
        cts_items = []
        cts_match = re.search(r'C\.\s*Transmission\s*system\s*for\s*Connectivity\s*under\s*GNA[:\s]*(.+?)(?=Start Date|Sl\.|Minutes|$)',
                             section[:3500], re.DOTALL | re.IGNORECASE)
        if cts_match:
            cts_raw = cts_match.group(1).strip()
            annex_ref = re.search(r'Annexure[- ]?([IVX]+)', cts_raw, re.IGNORECASE)
            if annex_ref:
                annex_key = annex_ref.group(1).upper()
                if annex_key in annexures:
                    cts_items = annexures[annex_key]
                cts_text = f"ANNEXURE:{annex_key}"
            else:
                cts_text = ' '.join(cts_raw.split())[:500]
        
        # Associate with last 6 app IDs (most relevant)
        for app_id in set(app_ids[-6:]):
            if app_id not in results:
                results[app_id] = {'ATS': '', 'DTL': '', 'CTS': '', 'CTS_items': []}
            
            if ats_text and not results[app_id]['ATS']:
                results[app_id]['ATS'] = ats_text
            if dtl_text and not results[app_id]['DTL']:
                results[app_id]['DTL'] = dtl_text
            if cts_text and not results[app_id]['CTS']:
                results[app_id]['CTS'] = cts_text
            if cts_items and not results[app_id]['CTS_items']:
                results[app_id]['CTS_items'] = cts_items
    
    return results

def main():
    print("=" * 70)
    print("CTU PDF Extractor - v6 (Final)")
    print("=" * 70)
    
    element_map, code_to_desc = load_element_status()
    
    print("\n[*] Processing PDFs...")
    text_33 = extract_pdf_text(PDF_33)
    annexures_33 = extract_annexures(text_33)
    data_33 = extract_app_data(text_33, annexures_33)
    
    text_34 = extract_pdf_text(PDF_34)
    annexures_34 = extract_annexures(text_34)
    data_34 = extract_app_data(text_34, annexures_34)
    
    # Merge data
    all_data = data_33.copy()
    for app_id, data in data_34.items():
        if app_id not in all_data:
            all_data[app_id] = data
        else:
            for key in ['ATS', 'DTL', 'CTS']:
                if data[key] and not all_data[app_id][key]:
                    all_data[app_id][key] = data[key]
            if data['CTS_items'] and not all_data[app_id]['CTS_items']:
                all_data[app_id]['CTS_items'] = data['CTS_items']
    
    print(f"[+] Extracted data for {len(all_data)} App IDs")
    print(f"    Annexures in PDF 33: {len(annexures_33)}")
    print(f"    Annexures in PDF 34: {len(annexures_34)}")
    
    # Show annexure content
    print("\nAnnexure sample items:")
    for key in list(annexures_33.keys())[:3]:
        print(f"  {key}: {len(annexures_33[key])} items")
        if annexures_33[key]:
            print(f"    First: {annexures_33[key][0][:60]}...")
    
    # Update Excel
    print("\n[*] Updating Excel...")
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Data to be captured']
    
    updates = {'CTS': 0, 'ATS': 0, 'DTL': 0}
    seen = set()
    checked = 0
    matched = 0
    
    for row in range(DATA_START_ROW, ws.max_row + 1):
        cell = ws.cell(row=row, column=COL_APP_ID + 1).value
        if not cell:
            continue
        
        app_id = re.sub(r'\D', '', str(cell).split()[0])[:10]
        if not app_id or len(app_id) != 10 or app_id in seen:
            continue
        seen.add(app_id)
        
        if app_id in all_data:
            checked += 1
            data = all_data[app_id]
            
            cur_cts = ws.cell(row=row, column=COL_CTS + 1).value
            cur_ats = ws.cell(row=row, column=COL_ATS + 1).value
            cur_dtl = ws.cell(row=row, column=COL_DTL + 1).value
            
            # CTS Update
            if not cur_cts:
                cts_codes = set()
                if data['CTS_items']:
                    for item in data['CTS_items']:
                        codes = find_element_codes(item, element_map)
                        cts_codes.update(codes)
                elif data['CTS'] and not data['CTS'].startswith('ANNEXURE:'):
                    codes = find_element_codes(data['CTS'], element_map)
                    cts_codes.update(codes)
                
                if cts_codes:
                    ws.cell(row=row, column=COL_CTS + 1).value = ','.join(sorted(cts_codes))
                    updates['CTS'] += 1
                    matched += 1
                    print(f"  CTS: Row {row} ({app_id}) -> {len(cts_codes)} codes")
            
            # ATS Update
            if not cur_ats and data['ATS'] and data['ATS'] != "NIL":
                ats_codes = find_element_codes(data['ATS'], element_map)
                if ats_codes:
                    ws.cell(row=row, column=COL_ATS + 1).value = ','.join(sorted(set(ats_codes)))
                    updates['ATS'] += 1
            
            # DTL Update
            if not cur_dtl and data['DTL']:
                dtl_codes = find_element_codes(data['DTL'], element_map)
                if dtl_codes:
                    ws.cell(row=row, column=COL_DTL + 1).value = ','.join(sorted(set(dtl_codes)))
                    updates['DTL'] += 1
    
    print(f"\n[*] Checked {checked} App IDs from PDFs against Excel")
    print(f"[*] Saving...")
    wb.save(EXCEL_PATH)
    print("[+] Done!")
    
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"  CTS updates: {updates['CTS']}")
    print(f"  ATS updates: {updates['ATS']}")
    print(f"  DTL updates: {updates['DTL']}")
    print(f"  TOTAL:       {sum(updates.values())}")
    print("=" * 70)

if __name__ == "__main__":
    main()
