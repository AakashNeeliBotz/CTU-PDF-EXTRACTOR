"""
Comprehensive scan for all Application IDs mixed in Substation column
"""
import pandas as pd
from openpyxl import load_workbook
import sys
import re

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

excel_path = r"c:\Users\Sree Charan\Desktop\fold2\CTU-PDF-EXTRACTOR\Connectivity_Application_Data_TEST_ALL_SHEETS38 (2).xlsx"

# Load workbook
wb = load_workbook(excel_path)
ws = wb["Data to be captured"]

print("="*100)
print("COMPREHENSIVE SCAN FOR APPLICATION IDS IN SUBSTATION COLUMN")
print("="*100)

srno_col = 2  # Column B
substation_col = 5  # Column E
app_id_col = 9  # Column I

# Pattern to find: Substation name followed by numbers
issues_found = []

for row in range(3, ws.max_row + 1):
    substation = ws.cell(row=row, column=substation_col).value
    if substation:
        sub_str = str(substation).strip()
        
        # Check for patterns:
        # 1. Pure number (Application ID only)
        # 2. Text followed by number with spacing
        # 3. Number embedded in text
        
        # Pattern 1: Pure 9-10 digit number
        if re.match(r'^\d{9,10}$', sub_str):
            srno = ws.cell(row=row, column=srno_col).value
            issues_found.append({
                'row': row,
                'srno': srno,
                'issue': 'Pure Application ID',
                'value': sub_str
            })
        
        # Pattern 2: Text + spaces + number at end
        match = re.match(r'^(.+?)\s{2,}(\d{9,10})$', sub_str)
        if match:
            srno = ws.cell(row=row, column=srno_col).value
            issues_found.append({
                'row': row,
                'srno': srno,
                'issue': 'Mixed Substation+AppID',
                'value': sub_str
            })
        
        # Pattern 3: Number followed by text
        match = re.match(r'^(\d{9,10})\s+(.+)$', sub_str)
        if match:
            srno = ws.cell(row=row, column=srno_col).value
            issues_found.append({
                'row': row,
                'srno': srno,
                'issue': 'AppID + Text',
                'value': sub_str
            })

print(f"\nTotal issues found: {len(issues_found)}")
for item in issues_found:
    print(f"  Row {item['row']}, Sr.No. {item['srno']}: {item['issue']} - '{item['value']}'")

if len(issues_found) == 0:
    print("✓ No Application IDs found in Substation column")

print("\n" + "="*100)
print("FINAL STATE CHECK")
print("="*100)

# Load dataframe
df = pd.read_excel(excel_path, sheet_name="Data to be captured", header=1)
print(f"\nTotal rows: {len(df)}")

# Check Sr.No. 301, 302, 303
print("\nSr.No. 301, 302, 303 check:")
for target_srno in [301, 302, 303]:
    row = df[df['Sr.no.'] == target_srno]
    if len(row) > 0:
        sub = row.iloc[0]['Substation']
        app_id = row.iloc[0]['GNA/ST II Application ID']
        print(f"  Sr.No. {target_srno}: Substation='{sub}', AppID='{app_id}'")
